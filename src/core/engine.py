#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Главный движок анализа временных рядов.
Содержит класс BigMasterTool и все функции расчета метрик.

TODO: Этот файл требует дальнейшего рефакторинга - разнести метрики по отдельным модулям.
"""

import argparse
import base64
import datetime as _dt
import html as _html
import importlib
import importlib.util
import logging
import os
import shutil
import warnings
from collections import Counter
from dataclasses import dataclass
from io import BytesIO
from pathlib import Path
from itertools import chain, combinations, permutations
from typing import Dict, List, Optional, Tuple

import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import networkx as nx
import numpy as np
import pandas as pd
import scipy.signal as signal
from hurst import compute_Hc
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from scipy import stats
from scipy.fft import fft
from scipy.signal import coherence, find_peaks
from scipy.spatial import cKDTree
from scipy.special import digamma
from sklearn.linear_model import LinearRegression
from sklearn.preprocessing import StandardScaler
from statsmodels.graphics.tsaplots import plot_acf
from statsmodels.tsa.stattools import adfuller, grangercausalitytests
from statsmodels.tsa.vector_ar.var_model import VAR
from tqdm import tqdm

# --- Optional dependency: nolds ---
# nolds==0.5.2 imports pkg_resources, which is removed in setuptools>=81.
# On fresh Python 3.13 installs this often breaks with:
#   ModuleNotFoundError: No module named 'pkg_resources'
# We keep the tool usable (other metrics still work) and provide a clear message.
_NOLDS_IMPORT_ERROR: Optional[str] = None
try:
    import nolds  # type: ignore
except Exception as _e:  # pragma: no cover
    nolds = None  # type: ignore
    _NOLDS_IMPORT_ERROR = str(_e)

_NOLDS_WARNED = False


def _nolds_or_warn() -> bool:
    """Return True if nolds is available; otherwise log a clear hint once."""
    global _NOLDS_WARNED
    if nolds is not None:
        return True
    if not _NOLDS_WARNED:
        _NOLDS_WARNED = True
        msg = (
            "[nolds] недоступен. Метрики, зависящие от nolds (sampen/dfa/hurst_rs fallback), "
            "будут возвращать NaN.\n"
            "Частая причина на Python 3.13: setuptools>=81 удалил pkg_resources. "
            "Решение: установи setuptools<81 (см. requirements.txt)."
        )
        if _NOLDS_IMPORT_ERROR:
            msg += f"\nПричина импорта: {_NOLDS_IMPORT_ERROR}"
        logging.error(msg)
    return False

# Импорты из нашей новой структуры
from ..config import *
from .data_loader import load_or_generate
from .preprocessing import configure_warnings


# Используем константы из config.py
# Сохраняем только те, которых нет в config
# Функции загрузки данных перенесены в data_loader.py
# Используем load_or_generate из импортов

##############################################
# Функции-метрики
##############################################
def correlation_matrix(data: pd.DataFrame, **kwargs) -> np.ndarray:
    return data.corr().values

def partial_correlation_matrix(df: pd.DataFrame, control: list = None, **kwargs) -> np.ndarray:
    cols = [c for c in df.columns if pd.api.types.is_numeric_dtype(df[c])]
    n = len(cols)
    out = np.eye(n)
    for i in range(n):
        for j in range(i + 1, n):
            xi, xj = cols[i], cols[j]
            ctrl_vars = control if control is not None else [c for c in cols if c not in (xi, xj)]
            sub_cols = [xi, xj] + [c for c in ctrl_vars if c in cols and c not in (xi, xj)]
            sub = df[sub_cols].dropna()
            if sub.shape[0] < len(sub_cols) + 1:
                pcor = np.nan
            else:
                try:
                    R = sub.corr().values
                    P = np.linalg.pinv(R)
                    pcor = -P[0, 1] / np.sqrt(P[0, 0] * P[1, 1])
                except Exception:
                    pcor = np.nan
            out[i, j] = out[j, i] = pcor
    return out


def partial_h2_matrix(df: pd.DataFrame, control: list = None, **kwargs) -> np.ndarray:
    """Возвращает квадрат частичной корреляции (H^2) для заданного контроля."""
    pcor = partial_correlation_matrix(df, control=control, **kwargs)
    return pcor**2

def lagged_directed_correlation(df: pd.DataFrame, lag: int, **kwargs) -> np.ndarray:
    """Directed lagged correlation: M[src, tgt] = corr(src(t), tgt(t+lag)).

    Логический фикс: используем shift и совместный dropna, чтобы не рассинхронизировать время.
    """
    lag = int(max(1, lag))
    cols = list(df.columns)
    m = len(cols)
    out = np.full((m, m), np.nan, dtype=float)
    np.fill_diagonal(out, 0.0)

    for i, src in enumerate(cols):
        x = df[src]
        for j, tgt in enumerate(cols):
            if i == j:
                continue
            y = df[tgt].shift(-lag)  # y(t+lag) выравниваем с x(t)
            pair = pd.concat([x, y], axis=1).dropna()
            if pair.shape[0] <= 3:
                continue
            xv = pair.iloc[:, 0].values
            yv = pair.iloc[:, 1].values
            if np.nanstd(xv) == 0 or np.nanstd(yv) == 0:
                out[i, j] = np.nan
            else:
                out[i, j] = float(np.corrcoef(xv, yv)[0, 1])
    return out

def h2_matrix(df: pd.DataFrame, **kwargs) -> np.ndarray: return correlation_matrix(df)**2
def lagged_directed_h2(df: pd.DataFrame, lag: int, **kwargs) -> np.ndarray: return lagged_directed_correlation(df, lag)**2

def coherence_matrix(data: pd.DataFrame, **kwargs):
    """Средняя когерентность между всеми парами.

    ВАЖНО: ряды должны быть синхронизированы по общему индексу.
    Нельзя делать dropna() по каждому столбцу отдельно — иначе получишь "когерентность"
    между разными моментами времени при разнесённых NaN.
    """
    fs = float(kwargs.get("fs", 1.0))
    fs = fs if np.isfinite(fs) and fs > 0 else 1.0
    N = data.shape[1]
    coh = np.full((N, N), np.nan, dtype=float)
    np.fill_diagonal(coh, 1.0)

    for i in range(N):
        for j in range(i + 1, N):
            pair = data.iloc[:, [i, j]].dropna()
            if pair.shape[0] <= 3:
                continue

            s1 = _as_float64_1d(pair.iloc[:, 0].values)
            s2 = _as_float64_1d(pair.iloc[:, 1].values)
            n = int(min(s1.size, s2.size))
            if n <= 3:
                continue
            s1 = s1[:n]
            s2 = s2[:n]

            try:
                nperseg = int(max(8, min(64, n // 2)))
                _, Cxy = signal.coherence(s1, s2, fs=fs, nperseg=nperseg, detrend="constant")
                if Cxy.size == 0:
                    continue
                Cxy = np.clip(np.asarray(Cxy, dtype=np.float64), 0.0, 1.0)
                Cxy[~np.isfinite(Cxy)] = np.nan
                coh[i, j] = coh[j, i] = float(np.nanmean(Cxy)) if np.isfinite(Cxy).any() else np.nan
            except Exception:
                coh[i, j] = coh[j, i] = np.nan

    return coh

def _knn_entropy(X, k=DEFAULT_K_MI):
    """Вычисляет энтропию для 1D-массива с помощью KNN."""
    N = len(X)
    if N <= k: return 0.0
    tree = cKDTree(X.reshape(-1, 1))
    d, _ = tree.query(X.reshape(-1, 1), k=k + 1, p=np.inf)
    # Расстояние до k-го соседа
    r = d[:, k]
    # digamma(N) - digamma(k) + d*log(2*r_k) - это для d-мерного пространства
    # Для 1D: digamma(N) - digamma(k) + E[log(2r_k)]
    return digamma(N) - digamma(k) + np.mean(np.log(2 * r + 1e-10)) 


def _knn_mutual_info(X, Y, k=DEFAULT_K_MI):
    """KSG-оценка взаимной информации I(X;Y) через kNN (max-норма).

    Правки относительно наивных реализаций:
    - строгий eps (nextafter) для устойчивости к ties;
    - исключаем саму точку из подсчёта соседей;
    - используем ψ(nx+1), ψ(ny+1).
    """
    X = np.asarray(X, dtype=np.float64).ravel()
    Y = np.asarray(Y, dtype=np.float64).ravel()
    N = int(min(X.size, Y.size))
    if N <= k or N <= 3:
        return 0.0
    X = X[:N]
    Y = Y[:N]

    XY = np.c_[X, Y]
    tree_XY = cKDTree(XY)
    d, _ = tree_XY.query(XY, k=int(k) + 1, p=np.inf)
    eps = d[:, int(k)]
    # строгий радиус
    eps = np.nextafter(eps, 0.0)

    tree_X = cKDTree(X.reshape(-1, 1))
    tree_Y = cKDTree(Y.reshape(-1, 1))

    nx = np.fromiter(
        (max(0, len(tree_X.query_ball_point([X[i]], r=float(eps[i]), p=np.inf)) - 1) for i in range(N)),
        dtype=float,
        count=N,
    )
    ny = np.fromiter(
        (max(0, len(tree_Y.query_ball_point([Y[i]], r=float(eps[i]), p=np.inf)) - 1) for i in range(N)),
        dtype=float,
        count=N,
    )

    raw_mi = digamma(N) + digamma(int(k)) - np.mean(digamma(nx + 1.0) + digamma(ny + 1.0))
    if not np.isfinite(raw_mi):
        return float("nan")
    return float(max(0.0, raw_mi))

def mutual_info_matrix(data: pd.DataFrame, k=DEFAULT_K_MI, **kwargs):
    """Матрица взаимной информации (KSG kNN).

    Важно: используем совместный dropna по паре, чтобы не рассинхронизировать время.
    """
    cols = list(data.columns)
    n_vars = len(cols)
    mi = np.zeros((n_vars, n_vars), dtype=float)

    for i in range(n_vars):
        for j in range(i + 1, n_vars):
            pair = data.iloc[:, [i, j]].dropna()
            if pair.shape[0] <= k:
                mi[i, j] = mi[j, i] = np.nan
                continue
            s1 = pair.iloc[:, 0].values
            s2 = pair.iloc[:, 1].values
            val = _knn_mutual_info(s1, s2, k=k)
            mi[i, j] = mi[j, i] = val

    return mi

def _knn_conditional_mutual_info(X, Y, Z, k=DEFAULT_K_MI):
    """KSG-оценка условной взаимной информации I(X;Y|Z) через kNN (max-норма).

    Важно:
    - строгий eps (nextafter) для ties;
    - исключаем саму точку;
    - используем ψ(n+1).
    """
    X = np.asarray(X, dtype=np.float64).ravel()
    Y = np.asarray(Y, dtype=np.float64).ravel()
    Z = np.asarray(Z, dtype=np.float64)
    if Z.ndim == 1:
        Z = Z.reshape(-1, 1)
    N = int(min(X.size, Y.size, Z.shape[0]))
    if N <= k or N <= 3:
        return 0.0
    X = X[:N]
    Y = Y[:N]
    Z = Z[:N, :]

    XZ = np.c_[X, Z]
    YZ = np.c_[Y, Z]
    XYZ = np.c_[X, Y, Z]

    tree_XYZ = cKDTree(XYZ)
    d, _ = tree_XYZ.query(XYZ, k=int(k) + 1, p=np.inf)
    eps = d[:, int(k)]
    eps = np.nextafter(eps, 0.0)

    tree_XZ = cKDTree(XZ)
    tree_YZ = cKDTree(YZ)
    tree_Z = cKDTree(Z)

    nxz = np.fromiter(
        (max(0, len(tree_XZ.query_ball_point(XZ[i], r=float(eps[i]), p=np.inf)) - 1) for i in range(N)),
        dtype=float,
        count=N,
    )
    nyz = np.fromiter(
        (max(0, len(tree_YZ.query_ball_point(YZ[i], r=float(eps[i]), p=np.inf)) - 1) for i in range(N)),
        dtype=float,
        count=N,
    )
    nz = np.fromiter(
        (max(0, len(tree_Z.query_ball_point(Z[i], r=float(eps[i]), p=np.inf)) - 1) for i in range(N)),
        dtype=float,
        count=N,
    )

    cmi = digamma(int(k)) - np.mean(digamma(nxz + 1.0) + digamma(nyz + 1.0) - digamma(nz + 1.0))
    if not np.isfinite(cmi):
        return float("nan")
    return float(max(0.0, cmi))

def mutual_info_matrix_partial(
    data: pd.DataFrame,
    control: Optional[List[str]] = None,
    k=DEFAULT_K_MI,
    **kwargs,
):
    """Матрица условной взаимной информации I(X;Y|Z) (KSG kNN).

    Если control=None: Z = все остальные переменные.
    Важно: формируем sub-таблицу и делаем dropna ОДИН РАЗ (иначе рассинхрон).
    """
    cols = list(data.columns)
    N = len(cols)
    pmi = np.zeros((N, N), dtype=float)

    for i in range(N):
        for j in range(i + 1, N):
            xi, xj = cols[i], cols[j]
            Z_cols = control if control is not None else [c for c in cols if c not in (xi, xj)]
            Z_cols = [c for c in Z_cols if c in data.columns and c not in (xi, xj)]

            if not Z_cols:
                pair = data[[xi, xj]].dropna()
                if pair.shape[0] <= k:
                    val = np.nan
                else:
                    val = _knn_mutual_info(pair[xi].values, pair[xj].values, k=k)
            else:
                sub = data[[xi, xj] + Z_cols].dropna()
                if sub.shape[0] <= k:
                    val = np.nan
                else:
                    X = sub[xi].values
                    Y = sub[xj].values
                    Z = sub[Z_cols].values
                    val = _knn_conditional_mutual_info(X, Y, Z, k=k)

            pmi[i, j] = pmi[j, i] = val

    return pmi

def compute_granger_matrix(df: pd.DataFrame, lags: int = DEFAULT_MAX_LAG, **kwargs) -> np.ndarray:
    n = df.shape[1]
    G = np.full((n, n), 1.0)
    cols = df.columns.tolist()
    # matrix[src, tgt] = pvalue(src -> tgt)
    for src in range(n):
        for tgt in range(n):
            if src == tgt:
                G[src, tgt] = 0.0
                continue
            data_pair = df[[cols[tgt], cols[src]]].dropna()  # [target, source]
            if len(data_pair) > lags * 2 + 5:
                try:
                    tests = grangercausalitytests(data_pair, maxlag=lags, verbose=False)
                    G[src, tgt] = tests[lags][0]['ssr_ftest'][1]
                except (np.linalg.LinAlgError, ValueError):
                    G[src, tgt] = np.nan
    return G

def _load_pyinform():
    """Ленивая загрузка pyinform без падения всего приложения."""
    if not PYINFORM_AVAILABLE:
        logging.warning(
            "[PyInform] pyinform не установлен: TE будет посчитан через fallback (дискретизация + эмпирические вероятности).",
        )
        return None
    return importlib.import_module("pyinform")



def _transfer_entropy_discrete(source_d: np.ndarray, target_d: np.ndarray, k: int = 1) -> float:
    """Эмпирическая transfer entropy для дискретных целочисленных рядов.

    Конвенция: TE(source -> target).
    Реализация: sum p(x_{t}|x_{t-k:t-1}, y_{t-k:t-1}) * log( p(x_t|x_past,y_past) / p(x_t|x_past) )
    """
    try:
        k = int(k)
        if k < 1:
            k = 1
        source_d = np.asarray(source_d, dtype=int).ravel()
        target_d = np.asarray(target_d, dtype=int).ravel()
        n = min(source_d.size, target_d.size)
        if n <= k + 1:
            return float("nan")
        source_d = source_d[:n]
        target_d = target_d[:n]
        n_eff = n - k

        c_xyz = Counter()
        c_xx = Counter()
        c_xpast_ypast = Counter()
        c_xpast = Counter()

        for t in range(k, n):
            x_next = int(target_d[t])
            x_past = tuple(int(v) for v in target_d[t - k : t])
            y_past = tuple(int(v) for v in source_d[t - k : t])

            c_xyz[(x_next, x_past, y_past)] += 1
            c_xx[(x_next, x_past)] += 1
            c_xpast_ypast[(x_past, y_past)] += 1
            c_xpast[x_past] += 1

        te = 0.0
        for (x_next, x_past, y_past), c in c_xyz.items():
            denom_xy = c_xpast_ypast.get((x_past, y_past), 0)
            denom_x = c_xpast.get(x_past, 0)
            if denom_xy <= 0 or denom_x <= 0:
                continue
            p1 = c / denom_xy  # p(x_next | x_past, y_past)
            p0 = c_xx.get((x_next, x_past), 0) / denom_x  # p(x_next | x_past)
            if p0 <= 0 or p1 <= 0:
                continue
            te += (c / n_eff) * float(np.log(p1 / p0))

        # численная защита
        if not np.isfinite(te):
            return float("nan")
        return float(max(0.0, te))
    except Exception:
        return float("nan")


def compute_TE(source: np.ndarray, target: np.ndarray, lag: int = 1, bins: int = DEFAULT_BINS):
    """Transfer Entropy (дискретная).

    Семантика параметра lag: длина истории k (как в pyinform.transfer_entropy k=...).

    Логические фиксы:
    - дискретизация через квантильные бины (устойчивее, чем min/max линейка);
    - перед бинингом z-score (масштаб-инвариантность);
    - маленький детерминированный jitter при ties (квантизация/повторы значений).
    """
    try:
        def _zscore_1d(x: np.ndarray) -> np.ndarray:
            x = np.asarray(x, dtype=np.float64).ravel()
            if x.size == 0:
                return x
            m = np.nanmean(x)
            s = np.nanstd(x)
            if not np.isfinite(s) or s <= 0:
                return x - m
            return (x - m) / s

        def _add_tiny_jitter(x: np.ndarray) -> np.ndarray:
            # детерминированный jitter для борьбы с ties (важно для kNN/квантилей)
            x = np.asarray(x, dtype=np.float64)
            if x.size <= 3:
                return x
            # если много повторов — чуть шевелим
            uniq = np.unique(x[np.isfinite(x)])
            if uniq.size < max(3, int(0.2 * x.size)):
                rng = np.random.default_rng(0)
                scale = (np.nanstd(x) if np.nanstd(x) > 0 else 1.0) * 1e-10
                x = x + rng.normal(0.0, scale, size=x.shape)
            return x

        def discretize_quantile(series: np.ndarray, num_bins: int) -> np.ndarray:
            s = np.asarray(series, dtype=np.float64).ravel()
            if s.size == 0:
                return np.array([], dtype=int)
            s = _add_tiny_jitter(_zscore_1d(s))
            if not np.isfinite(s).all():
                s = s[np.isfinite(s)]
            if s.size == 0:
                return np.array([], dtype=int)
            if float(np.nanmin(s)) == float(np.nanmax(s)):
                return np.zeros(int(series.size), dtype=int)

            q = np.linspace(0.0, 1.0, int(num_bins) + 1)
            edges = np.quantile(s, q)
            # убираем дубли рёбер (если данных мало/много ties)
            edges = np.unique(edges)
            if edges.size <= 2:
                return np.zeros(int(series.size), dtype=int)

            # делаем края "открытыми" справа
            edges[-1] = np.nextafter(edges[-1], edges[-1] + 1.0)
            disc = np.digitize(_add_tiny_jitter(_zscore_1d(np.asarray(series))), bins=edges[1:-1], right=False)
            disc = np.clip(disc, 0, int(num_bins) - 1)
            return disc.astype(int)

        source_discrete = discretize_quantile(source, bins)
        target_discrete = discretize_quantile(target, bins)

        pyinform = _load_pyinform()
        k = int(max(1, lag))
        if pyinform is not None:
            return float(pyinform.transfer_entropy(source_discrete, target_discrete, k=k))

        return _transfer_entropy_discrete(source_discrete, target_discrete, k=k)

    except Exception as e:
        logging.error(f"[TE] Ошибка вычисления: {e}")
        return float("nan")

def TE_matrix(df: pd.DataFrame, lag: int = 1, bins: int = DEFAULT_BINS, **kwargs):
    """
    Строит матрицу Transfer Entropy для всех пар.

    Конвенция для направленных матриц: M[src, tgt] = мера src → tgt.
    """
    n = df.shape[1]
    te_matrix = np.zeros((n, n))

    for src in range(n):
        for tgt in range(n):
            if src == tgt:
                continue

            pair = df.iloc[:, [src, tgt]].dropna()
            if len(pair) <= lag:
                te_matrix[src, tgt] = np.nan
                continue
            s_src = pair.iloc[:, 0].values
            s_tgt = pair.iloc[:, 1].values
            te_matrix[src, tgt] = compute_TE(s_src, s_tgt, lag=lag, bins=bins)

    return te_matrix

def TE_matrix_partial(
    df: pd.DataFrame,
    lag: int = 1,
    control: Optional[List[str]] = None,
    bins: int = DEFAULT_BINS,
) -> np.ndarray:
    """
    Приближённая "partial" Transfer Entropy.

    Практичная аппроксимация:
      1) линейно вычитаем влияние контрольных переменных из src и tgt (остатки OLS);
      2) считаем обычный TE между остатками.

    Конвенция: M[src, tgt] = мера src → tgt.
    """
    cols = list(df.columns)
    N = len(cols)
    M = np.zeros((N, N))

    def _residualize(y: np.ndarray, X: np.ndarray) -> np.ndarray:
        if X.size == 0:
            return y
        X_aug = np.c_[np.ones(len(X)), X]
        beta, *_ = np.linalg.lstsq(X_aug, y, rcond=None)
        return y - X_aug @ beta

    for i, src in enumerate(cols):
        for j, tgt in enumerate(cols):
            if i == j:
                continue

            ctrl_cols = control if control is not None else [c for c in cols if c not in (src, tgt)]
            ctrl_cols = [c for c in ctrl_cols if c in df.columns]

            use_cols = [src, tgt] + ctrl_cols
            sub = df[use_cols].dropna()
            if sub.shape[0] <= lag + 1:
                M[i, j] = np.nan
                continue

            s_src = sub[src].values
            s_tgt = sub[tgt].values
            X_ctrl = sub[ctrl_cols].values if ctrl_cols else np.empty((len(sub), 0))

            try:
                s_src_r = _residualize(s_src, X_ctrl)
                s_tgt_r = _residualize(s_tgt, X_ctrl)
                M[i, j] = compute_TE(s_src_r, s_tgt_r, lag=lag, bins=bins)
            except Exception:
                M[i, j] = np.nan

    return M


def AH_matrix(df: pd.DataFrame, embed_dim=DEFAULT_EMBED_DIM, tau=DEFAULT_EMBED_TAU) -> np.ndarray:
    """
    Конвенция: M[src, tgt] = мера src → tgt.
    """
    df = df.dropna(axis=0, how='any')
    N = df.shape[1]
    out = np.zeros((N, N))
    arr = df.values
    for src in range(N):
        for tgt in range(N):
            if src == tgt:
                out[src, tgt] = 0.0
                continue

            # _H_ratio_direction(X, Y): интерпретируется как X → Y
            H_val = _H_ratio_direction(arr[:, src], arr[:, tgt], m=embed_dim, tau=tau)
            if H_val is None or H_val <= 0:
                AH = 0.0
            else:
                AH = 1.0 / H_val
                if AH > 1.0:
                    AH = 1.0
            out[src, tgt] = AH
    return out

def _H_ratio_direction(X, Y, m=DEFAULT_EMBED_DIM, tau=DEFAULT_EMBED_TAU):
    n = len(X)
    if len(Y) != n or n < 2:
        return None
    L = n - (m - 1) * tau
    if L < 2:
        return None
    
    X_state = np.zeros((L, m))
    Y_state = np.zeros((L, m))
    for j in range(m):
        X_state[:, j] = X[j*tau : j*tau + L]
        Y_state[:, j] = Y[j*tau : j*tau + L]
    
    valid_indices = ~np.isnan(X_state).any(axis=1) & ~np.isnan(Y_state).any(axis=1)
    if not np.any(valid_indices):
        return None
        
    X_state_valid = X_state[valid_indices]
    Y_state_valid = Y_state[valid_indices]
    
    if len(X_state_valid) < 2:
        return None

    tree_X = cKDTree(X_state_valid)
    tree_Y = cKDTree(Y_state_valid)
    
    dists_X, idx_X = tree_X.query(X_state_valid, k=2)
    
    
    if idx_X.shape[1] < 2: 
        return None 
    
    nn_idx = idx_X[:, 1] 
    diff = Y_state_valid - Y_state_valid[nn_idx]
    dY1 = np.sqrt(np.sum(diff**2, axis=1))
    
    dists_Y, _ = tree_Y.query(Y_state_valid, k=2)
    dY2 = dists_Y[:, 1]
    dY2 = np.where(dY2 == 0, 1e-10, dY2) 
    
    ratios = dY1 / dY2
    ratios = ratios[np.isfinite(ratios)] 
    
    if len(ratios) == 0:
        return None

    H_val = np.mean(ratios)
    return H_val

def compute_partial_AH_matrix(data: pd.DataFrame,
                               max_lag: int = DEFAULT_MAX_LAG,
                               embed_dim: int = DEFAULT_EMBED_DIM,
                               tau: int = DEFAULT_EMBED_TAU,
                               control: List[str] = None) -> np.ndarray:
    df = data.dropna(axis=0, how='any')
    N = df.shape[1]
    if N < 2:
        return np.zeros((N, N))

    if control and len(control) > 0:
        resid_df = pd.DataFrame(index=df.index)
        for col in df.columns:
            X_ctrl = df[control]
            y = df[col]
            if len(X_ctrl) > 0 and len(y) == len(X_ctrl) and not X_ctrl.isnull().any().any():
                try:
                    model = LinearRegression().fit(X_ctrl.values, y.values)
                    resid = y.values - model.predict(X_ctrl.values)
                    resid_df[col] = resid
                except ValueError: 
                    resid_df[col] = y 
            else:
                resid_df[col] = y
    else:
        try:
            model = VAR(df.values)
            res_full = model.fit(max_lag, ic=None)
            resid_df = pd.DataFrame(res_full.resid, columns=df.columns)
        except Exception as e:
            logging.error(f"VAR fit error (partial AH, fallback to raw): {e}")
            resid_df = df 

    return AH_matrix(resid_df, embed_dim=embed_dim, tau=tau)


def directional_AH_matrix(df: pd.DataFrame, maxlags: int = 5) -> np.ndarray:
    return AH_matrix(df, embed_dim=DEFAULT_EMBED_DIM, tau=DEFAULT_EMBED_TAU)

def granger_dict(df: pd.DataFrame, maxlag: int = 4) -> dict:
    results = {}
    cols = list(df.columns)
    for i, tgt in enumerate(cols):
        for j, src in enumerate(cols):
            if i == j:
                continue
            sub = df[[tgt, src]].dropna()
            if len(sub) < (maxlag + 10):
                results[f"{src}->{tgt}"] = None
                continue
            try:
                tests = grangercausalitytests(sub, maxlag=maxlag, verbose=False)
            except Exception as e:
                logging.error(f"[Granger] Ошибка Granger для {src}->{tgt}: {e}")
                results[f"{src}->{tgt}"] = None
                continue 
            results[f"{src}->{tgt}"] = tests # Сохраняем РЕЗУЬТАТ
    return results

# эта матрица НЕ ТА ЖЕ ЧТО В МАППИНГЕ
def _compute_granger_matrix_internal(df: pd.DataFrame, lags: int = DEFAULT_MAX_LAG) -> np.ndarray:
    n = df.shape[1]
    G = np.zeros((n, n))
    cols = df.columns.tolist()
    # matrix[src, tgt] = pvalue(src -> tgt)
    for src in range(n):
        for tgt in range(n):
            if src == tgt:
                G[src, tgt] = 0.0
                continue
            sub = df[[cols[tgt], cols[src]]].dropna()  # [target, source]
            try:
                tests = grangercausalitytests(sub, maxlag=lags, verbose=False)
                pvals = [tests[l][0]['ssr_ftest'][1] for l in tests]
                G[src, tgt] = min(pvals)
            except Exception as e:
                logging.error(f"[Granger-Internal] Ошибка Granger для {cols[src]}->{cols[tgt]}: {e}")
                G[src, tgt] = np.nan
    return G


def compute_partial_granger_matrix(data: pd.DataFrame, lags=DEFAULT_MAX_LAG) -> np.ndarray:
    """
    контроль остальных переменных, грейндж
    """
    df = data.dropna(axis=0, how='any')
    N = df.shape[1]
    if N < 3:
        return _compute_granger_matrix_internal(data, lags=lags)
    pg_matrix = np.zeros((N, N))
    T = len(df)
    p = lags
    if T <= p:
        return pg_matrix
    arr = df.values
    try:
        model_full = VAR(arr)
        res_full = model_full.fit(p, ic=None)
    except Exception as e:
        logging.error(f"VAR fit error (partial Granger): {e}")
        return pg_matrix
    sigma_full = np.cov(res_full.resid, rowvar=False)
    for i in range(N):
        reduced_arr = np.delete(arr, i, axis=1)
        try:
            model_red = VAR(reduced_arr)
            res_red = model_red.fit(p, ic=None)
            sigma_red = np.cov(res_red.resid, rowvar=False)
        except Exception as e:
            for j in range(N):
                if j != i:
                    pg_matrix[i, j] = np.nan 
            continue
        for j in range(N):
            if i == j:
                pg_matrix[i, j] = 0.0
            else:
                idx_j = j - 1 if i < j else j
                var_full = sigma_full[j, j] if sigma_full.shape[0] > j else np.var(res_full.resid[:, j])
                var_red = sigma_red[idx_j, idx_j] if sigma_red.shape[0] > idx_j else np.var(res_red.resid[:, idx_j])
                if var_full <= 0 or var_red <= 0:
                    gc_val = np.nan 
                else:
                    gc_val = np.log(var_red / var_full)
                    if gc_val < 0:
                        gc_val = 0.0
                pg_matrix[i, j] = gc_val
    return pg_matrix

def p_to_score(p: float, eps: float = 1e-300) -> float:
    """Convert p-value to comparable strength score: -log10(p). Higher = stronger."""
    if p is None or (isinstance(p, float) and np.isnan(p)):
        return np.nan
    p = float(p)
    if p <= 0:
        p = eps
    return float(-np.log10(p))


def granger_matrix(df: pd.DataFrame, granger_dict_result: dict) -> np.ndarray:
    cols = list(df.columns)
    n_vars = len(cols)
    G = np.ones((n_vars, n_vars))
    for i, tgt in enumerate(cols):
        for j, src in enumerate(cols):
            if i == j:
                G[i, j] = 0
            else:
                key = f"{src}->{tgt}"
                if granger_dict_result.get(key) is None:
                    G[i, j] = np.nan
                else:
                    test_dict = granger_dict_result[key]
                    bp = 1.0 
                    found_valid_p = False 
                    for lag_val, dct in test_dict.items():
                        if isinstance(dct, list) and len(dct) > 0 and 'ssr_ftest' in dct[0]:
                            F, pval, _, _ = dct[0]['ssr_ftest']
                            if not np.isnan(pval): 
                                bp = min(bp, pval)
                                found_valid_p = True
                    G[i, j] = bp if found_valid_p else np.nan 
    return G



def remove_linear_dependency(sub: pd.DataFrame, src: str, tgt: str, control_cols: list) -> tuple[np.ndarray, np.ndarray]:
    """Удаляет линейную компоненту контролей из src/tgt и возвращает остатки (src_res, tgt_res)."""
    if not control_cols:
        r1 = sub[src].to_numpy(dtype=float)
        r2 = sub[tgt].to_numpy(dtype=float)
        return r1, r2

    X = sub[control_cols].to_numpy(dtype=float)
    y_src = sub[src].to_numpy(dtype=float)
    y_tgt = sub[tgt].to_numpy(dtype=float)

    mdl_src = LinearRegression().fit(X, y_src)
    mdl_tgt = LinearRegression().fit(X, y_tgt)

    r1 = y_src - mdl_src.predict(X)
    r2 = y_tgt - mdl_tgt.predict(X)
    return r1, r2

def granger_matrix_partial(df: pd.DataFrame, maxlag=DEFAULT_MAX_LAG, control=None) -> np.ndarray:
    """Conditional Granger causality p-values via multivariate VAR.

    Конвенция: G[src, tgt] = pvalue( src -> tgt ), т.е. тест "src НЕ вызывает tgt".
    control: список контролей (без src/tgt). Если None — все остальные переменные.
    """
    columns = list(df.columns)
    n = len(columns)
    G = np.full((n, n), np.nan, dtype=float)
    np.fill_diagonal(G, 0.0)

    for src_i, src in enumerate(columns):
        for tgt_j, tgt in enumerate(columns):
            if src_i == tgt_j:
                continue

            control_cols = control if control is not None else [c for c in columns if c not in (src, tgt)]
            control_cols = [c for c in control_cols if c in df.columns and c not in (src, tgt)]
            use_cols = [tgt, src] + control_cols  # порядок важен только для удобства
            sub = df[use_cols].dropna()

            # минимум наблюдений: грубо (кол-во параметров VAR растёт с p и k)
            p = int(max(1, maxlag))
            k_vars = len(use_cols)
            if sub.shape[0] < max(30, 5 * p * k_vars):
                continue

            try:
                res = VAR(sub).fit(maxlags=p, ic=None, trend="c")
                test = res.test_causality(caused=tgt, causing=[src], kind="f")
                G[src_i, tgt_j] = float(test.pvalue) if np.isfinite(test.pvalue) else np.nan
            except Exception:
                G[src_i, tgt_j] = np.nan

    return G

def plt_fft_analysis(series: pd.Series, *, fs: float = 1.0):
    """Быстрый FFT-анализ.

    Важно: частотная шкала зависит от частоты дискретизации fs (Гц).
    По умолчанию fs=1.0 (частоты в "циклах на отсчёт").
    """
    arr = _as_float64_1d(series.dropna().values)
    if arr.size == 0:
        return np.array([]), np.array([]), np.array([]), np.array([])
    n = int(arr.size)
    fs = float(fs) if fs and np.isfinite(fs) and fs > 0 else 1.0
    dt = 1.0 / fs
    freqs = np.fft.fftfreq(n, d=dt)
    fft_vals = fft(arr)
    amplitude = np.abs(fft_vals)
    phase = np.angle(fft_vals)
    pos_mask = freqs >= 0
    freqs, amplitude, phase = freqs[pos_mask], amplitude[pos_mask], phase[pos_mask]
    peaks, _ = find_peaks(amplitude, height=(np.max(amplitude) * 0.2 if amplitude.size > 0 else 0))
    logging.debug(f"[FFT] Найдено пиков на частотах: {freqs[peaks] if peaks.size > 0 else 'Нет пиков'}")
    return freqs, amplitude, phase, peaks

def plot_amplitude_response(series: pd.Series, title: str) -> BytesIO:
    freqs, amplitude, phase, peaks = plt_fft_analysis(series)
    fig, ax = plt.subplots(figsize=(6, 4))
    ax.plot(freqs, amplitude, label="АЧХ")
    if peaks.size > 0:
        ax.plot(freqs[peaks], amplitude[peaks], "x", label="Пики")
    ax.set_title(title)
    ax.set_xlabel("Частота")
    ax.set_ylabel("Амплитуда")
    ax.legend()
    buf = BytesIO()
    plt.tight_layout()
    plt.savefig(buf, format="png", dpi=100)
    buf.seek(0)
    plt.close(fig)
    return buf

def plot_phase_response(series: pd.Series, title: str) -> BytesIO:
    freqs, amplitude, phase, peaks = plt_fft_analysis(series)
    fig, ax = plt.subplots(figsize=(6, 4))
    ax.plot(freqs, phase, label="ФЧХ", color="orange")
    ax.set_title(title)
    ax.set_xlabel("Частота")
    ax.set_ylabel("Фаза (рад)")
    ax.legend()
    buf = BytesIO()
    plt.tight_layout()
    plt.savefig(buf, format="png", dpi=100)
    buf.seek(0)
    plt.close(fig)
    return buf

def plot_combined_ac_fch(data: pd.DataFrame, title: str) -> BytesIO:
    fig, (ax1, ax2) = plt.subplots(2, 1, figsize=(8, 8), sharex=True)
    for col in data.columns:
        series = data[col]
        freqs, amplitude, phase, _ = plt_fft_analysis(series)
        if freqs.size > 0:
            ax1.plot(freqs, amplitude, label=col)
            ax2.plot(freqs, phase, label=col)
    ax1.set_title(title + " - АЧХ")
    ax1.set_ylabel("Амплитуда")
    ax1.legend()
    ax2.set_title(title + " - ФЧХ")
    ax2.set_xlabel("Частота")
    ax2.set_ylabel("Фаза (рад)")
    ax2.legend()
    buf = BytesIO()
    plt.tight_layout()
    plt.savefig(buf, format="png", dpi=100)
    buf.seek(0)
    plt.close(fig)
    return buf

##############################################
# индивидуальный AC & PH для каждого ряда
##############################################
def plot_individual_ac_ph(data: pd.DataFrame, title: str) -> dict:
    plots = {}
    for col in data.columns:
        series = data[col]
        # График АЧХ
        fig1, ax1 = plt.subplots(figsize=(6, 4))
        freqs, amplitude, _, peaks = plt_fft_analysis(series)
        ax1.plot(freqs, amplitude, label=f"АЧХ {col}")
        if peaks.size > 0:
            ax1.plot(freqs[peaks], amplitude[peaks], "x", label="Пики")
        ax1.set_title(f"АЧХ {col}")
        ax1.set_xlabel("Частота")
        ax1.set_ylabel("Амплитуда")
        ax1.legend()
        buf1 = BytesIO()
        plt.tight_layout()
        plt.savefig(buf1, format="png", dpi=100)
        buf1.seek(0)
        plt.close(fig1)
        # График ФЧХ
        fig2, ax2 = plt.subplots(figsize=(6, 4))
        freqs, _, phase, _ = plt_fft_analysis(series)
        ax2.plot(freqs, phase, label=f"ФЧХ {col}", color="orange")
        ax2.set_title(f"ФЧХ {col}")
        ax2.set_xlabel("Частота")
        ax2.set_ylabel("Фаза (рад)")
        ax2.legend()
        buf2 = BytesIO()
        plt.tight_layout()
        plt.savefig(buf2, format="png", dpi=100)
        buf2.seek(0)
        plt.close(fig2)
        plots[col] = {"AC": buf1, "PH": buf2}
    return plots

##############################################
# sample entropy
##############################################
def compute_sample_entropy(series: pd.Series) -> float:
    try:
        if not _nolds_or_warn():
            return np.nan
        return float(nolds.sampen(series.dropna().values))
    except Exception as ex:
        logging.error(f"[Sample Entropy] Ошибка: {ex}")
        return np.nan


def _as_float64_1d(x) -> np.ndarray:
    """Безопасно приводит вход к 1D float64 без NaN/inf."""
    arr = np.asarray(x, dtype=np.float64).reshape(-1)
    if arr.size == 0:
        return arr
    m = np.isfinite(arr)
    return arr[m]

##############################################
# Функции для частотного анализа когерентности
##############################################
def plot_coherence_vs_frequency(
    series1: pd.Series,
    series2: pd.Series,
    title: str,
    *,
    fs: float = 1.0,
    nperseg: Optional[int] = None,
) -> BytesIO:
    s1 = _as_float64_1d(series1.dropna().values)
    s2 = _as_float64_1d(series2.dropna().values)
    n = int(min(s1.size, s2.size))
    if n <= 3:
        return BytesIO()
    s1 = s1[:n]
    s2 = s2[:n]
    fs = float(fs) if fs and np.isfinite(fs) and fs > 0 else 1.0
    if nperseg is None:
        # для коротких рядов дефолт scipy (256) даёт вырожденную оценку
        nperseg = int(max(8, min(64, n // 2)))
    nperseg = int(max(8, min(nperseg, n)))
    freqs, cxy = coherence(s1, s2, fs=fs, nperseg=nperseg, detrend="constant")
    if cxy.size:
        cxy = np.clip(np.asarray(cxy, dtype=np.float64), 0.0, 1.0)
        cxy[~np.isfinite(cxy)] = np.nan
    fig, ax = plt.subplots(figsize=(6, 4))
    ax.plot(freqs, cxy, label="Когерентность")
    if cxy.size > 0:
        # np.nanargmax падает на all-nan
        max_idx = int(np.nanargmax(cxy)) if np.isfinite(cxy).any() else 0
        max_freq = freqs[max_idx]
        max_coh = cxy[max_idx]
        ax.plot(max_freq, max_coh, "ro", label=f"Макс. связь: {max_coh:.3f} на {max_freq:.3f}Hz")
        ax.annotate(f"{max_freq:.3f} Hz", xy=(max_freq, max_coh), xytext=(max_freq, max_coh+0.05),
                    arrowprops=dict(facecolor='black', shrink=0.05))
    ax.set_title(title)
    ax.set_xlabel("Частота (Hz)")
    ax.set_ylabel("Когерентность")
    ax.set_ylim(0, 1)
    ax.legend()
    buf = BytesIO()
    plt.tight_layout()
    plt.savefig(buf, format="png")
    buf.seek(0)
    plt.close(fig)
    return buf

##############################################
# Функции для экспорта данных в Excel
##############################################
def add_raw_data_sheet(wb: Workbook, df: pd.DataFrame) -> None:
    """Добавляет лист с исходными данными."""
    ws = wb.create_sheet("Raw Data")
    ws.append(list(df.columns))
    for _, row in df.iterrows():
        ws.append(list(row))

def plot_heatmap(matrix: np.ndarray, title: str, legend_text: str = "", annotate: bool = False, vmin=None, vmax=None) -> BytesIO:
    fig, ax = plt.subplots(figsize=(4, 3.2))
    
    if matrix is None or not isinstance(matrix, np.ndarray) or matrix.size == 0:
        ax.text(0.5, 0.5, "Error\n(No Data)", ha='center', va='center', color='red', fontsize=12)
        ax.set_xticks([])
        ax.set_yticks([])
    else:
        # Фиксируем шкалу, чтобы избежать авто-нормализации matplotlib.
        cax = ax.imshow(matrix, cmap="viridis", aspect="auto", vmin=vmin, vmax=vmax)
        fig.colorbar(cax, ax=ax)
        ax.set_title(title, fontsize=10)
        
        # Аннотации
        if annotate and matrix.shape[0] < 10:
            min_val = vmin if vmin is not None else np.nanmin(matrix)
            max_val = vmax if vmax is not None else np.nanmax(matrix)
            
            if np.isfinite(min_val) and np.isfinite(max_val) and max_val > min_val:
                threshold = min_val + (max_val - min_val) / 2.0
            else:
                threshold = 0.5 # запасной вариант

            for i in range(matrix.shape[0]):
                for j in range(matrix.shape[1]):
                    val = matrix[i, j]
                    if np.isnan(val):
                        display_val, color = "NaN", "red"
                    else:
                        display_val = f"{val:.2f}"
                        # Цвет текста зависит от порога, который выч отдельно. от 0,5 НЕ ВОЗВРАЩАТЬ
                        color = "white" if val < threshold else "black"
                    ax.text(j, i, display_val, ha="center", va="center", color=color, fontsize=8)

    if legend_text:
        ax.text(0.05, 0.95, legend_text, transform=ax.transAxes, fontsize=8,
                verticalalignment='top', bbox=dict(facecolor='white', alpha=0.5))
    
    buf = BytesIO()
    plt.tight_layout()
    plt.savefig(buf, format="png", dpi=150)
    buf.seek(0)
    plt.close(fig)
    return buf

def plot_connectome(matrix: np.ndarray, method_name: str, threshold: float = 0.2,
                    directed: bool = False, invert_threshold: bool = False, legend_text: str = "") -> BytesIO:
    n = matrix.shape[0]
    G = nx.DiGraph() if directed else nx.Graph()
    G.add_nodes_from(range(n))
    # ВАЖНО: договоримся об ориентации:
    # matrix[src, tgt] — влияние src -> tgt
    if directed:
        for src in range(n):
            for tgt in range(n):
                if src == tgt:
                    continue
                w = matrix[src, tgt]
                if w is None or np.isnan(w):
                    continue
                if invert_threshold:
                    # p-value: меньше => сильнее
                    if w < threshold:
                        G.add_edge(src, tgt, weight=float(w))
                else:
                    if abs(w) > threshold:
                        G.add_edge(src, tgt, weight=float(w))
    else:
        for i in range(n):
            for j in range(i + 1, n):
                w = matrix[i, j]
                if w is None or np.isnan(w):
                    continue
                if abs(w) > threshold:
                    G.add_edge(i, j, weight=float(w))
    pos = nx.circular_layout(G)
    fig, ax = plt.subplots(figsize=(4, 4))
    if directed:
        nx.draw_networkx_nodes(G, pos, ax=ax, node_color="lightblue", node_size=500)
        nx.draw_networkx_labels(G, pos, ax=ax)
        nx.draw_networkx_edges(G, pos, ax=ax, arrowstyle="->", arrowsize=10)
    else:
        nx.draw_networkx(G, pos, ax=ax, node_color="lightblue", node_size=500)
    ax.set_title(f"Connectome: {method_name}")
    if legend_text:
         ax.text(0.05, 0.05, legend_text, transform=ax.transAxes, fontsize=8,
                 verticalalignment='bottom', bbox=dict(facecolor='white', alpha=0.5))
    ax.axis("off")
    buf = BytesIO()
    plt.tight_layout()
    plt.savefig(buf, format="png")
    buf.seek(0)
    plt.close(fig)
    return buf

def add_method_to_sheet(ws, row: int, title: str, matrix: np.ndarray, directed: bool = False, legend_text: str = "") -> int:
    ws.append([title])
    if matrix is None:
        ws.append(["Метод не работает для этих данных."])
        return ws.max_row
    df_mat = pd.DataFrame(matrix)
    for r in dataframe_to_rows(df_mat, index=False, header=True):
        ws.append(r)
    buf_heat = plot_heatmap(matrix, title + " Heatmap", legend_text=legend_text)
    img_heat = Image(buf_heat)
    img_heat.width = 400
    img_heat.height = 300
    ws.add_image(img_heat, f"A{ws.max_row + 2}")
    buf_conn = plot_connectome(matrix, title + " Connectome", threshold=0.2, directed=directed, invert_threshold=False, legend_text=legend_text)
    img_conn = Image(buf_conn)
    img_conn.width = 400
    img_conn.height = 400
    ws.add_image(img_conn, f"G{ws.max_row + 2}")
    return ws.max_row

def fmt_val(v):
    try:
        f = float(v)
        if np.isnan(f):
            return "N/A"
        return f"{f:.3f}"
    except Exception:
        return "N/A"

def fdr_bh(pvals: np.ndarray) -> np.ndarray:
    """Benjamini–Hochberg FDR correction. Returns q-values (same shape)."""
    p = np.asarray(pvals, dtype=float)
    q = np.full(p.shape, np.nan, dtype=float)
    mask = np.isfinite(p)
    if mask.sum() == 0:
        return q
    pv = p[mask].ravel()
    m = pv.size
    order = np.argsort(pv)
    ranked = pv[order]
    q_raw = ranked * m / (np.arange(1, m + 1))
    # monotone
    q_mono = np.minimum.accumulate(q_raw[::-1])[::-1]
    q_mono = np.clip(q_mono, 0.0, 1.0)
    out = np.empty_like(pv)
    out[order] = q_mono
    q[mask] = out
    return q

def apply_pvalue_correction_matrix(mat: np.ndarray, directed: bool) -> np.ndarray:
    """Apply FDR correction to a p-value matrix (off-diagonal entries)."""
    M = np.array(mat, dtype=float, copy=True)
    n = M.shape[0]
    if n == 0:
        return M
    mask = np.isfinite(M)
    # ignore diagonal
    np.fill_diagonal(mask, False)
    if not directed:
        # use only upper triangle to avoid double-counting, then mirror
        tri = np.triu(mask, 1)
        q = fdr_bh(M[tri])
        M[tri] = q
        M = M + M.T
        np.fill_diagonal(M, 0.0)
        return M
    else:
        q = fdr_bh(M[mask])
        M[mask] = q
        np.fill_diagonal(M, 0.0)
        return M

#
# МАППИНГ
#
# МАППИНГ
###
method_mapping = {
    # ——— Correlation ———
    "correlation_full":     lambda data, lag=None, control=None: correlation_matrix(data),
    "correlation_partial":  lambda data, lag=None, control=None: partial_correlation_matrix(data, control),
    "correlation_directed": lambda data, lag, control=None: lagged_directed_correlation(data, lag),

    # ——— H² (squared corr) ———
    "h2_full":              lambda data, lag=None, control=None: correlation_matrix(data)**2,
    "h2_partial":           lambda data, lag=None, control=None: partial_h2_matrix(data, control), 
    "h2_directed":          lambda data, lag, control=None: lagged_directed_h2(data, lag),

    # ——— Mutual Information ———
    "mutinf_full":          lambda data, lag=0, control=None: mutual_info_matrix(data, k=DEFAULT_K_MI),
    "mutinf_partial":       lambda data, lag=0, control=None: mutual_info_matrix_partial(data, control, k=DEFAULT_K_MI),

    # ——— Coherence ———
    "coherence_full":       lambda data, lag=None, control=None: coherence_matrix(data),


# ...
    # ——— Granger causality ———
    "granger_full":         lambda data, lag, control=None: _compute_granger_matrix_internal(data, lags=lag),
    "granger_partial":      lambda data, lag, control=None: granger_matrix_partial(data, maxlag=lag, control=control), 
    "granger_directed":     lambda data, lag, control=None: _compute_granger_matrix_internal(data, lags=lag),
# ...

    # ——— Transfer entropy ———
    "te_full":    lambda data, lag, control=None: TE_matrix(data, lag=lag),
    "te_partial": lambda data, lag, control=None: TE_matrix_partial(data, lag=lag, control=control, bins=DEFAULT_BINS),
    "te_directed":lambda data, lag, control=None: TE_matrix(data, lag=lag, bins=DEFAULT_BINS),


    # ——— AH (non‑linear) ———
    "ah_full":              lambda data, lag=None, control=None: AH_matrix(data),
    "ah_partial":           lambda data, lag, control=None: compute_partial_AH_matrix(data, max_lag=lag, control=control),
    "ah_directed":          lambda data, lag, control=None:
                                (AH_matrix(data)
                                 if not control
                                 else compute_partial_AH_matrix(data, max_lag=lag, control=control)),
}



@dataclass(frozen=True)
class MethodSpec:
    directed: bool
    is_p_value: bool
    control_dependent: bool
    supports_lag: bool
    description: str = ""

# -----------------------------
# Метод-метаданные и логика
# -----------------------------

# p-value методы: меньше = сильнее свидетельство связи (нужно invert_threshold=True)
PVAL_METHODS = {
    "granger_full",
    "granger_partial",
    "granger_directed",
}

# directed методы: матрица A[i,j] интерпретируется как i -> j
DIRECTED_METHODS = {
    "correlation_directed",
    "h2_directed",
    "granger_full",
    "granger_partial",
    "granger_directed",
    "te_full",
    "te_partial",
    "te_directed",
    "ah_full",
    "ah_partial",
    "ah_directed",
}


def is_pvalue_method(variant: str) -> bool:
    return variant.lower() in PVAL_METHODS


def is_directed_method(variant: str) -> bool:
    return variant.lower() in DIRECTED_METHODS


def is_control_sensitive_method(variant: str) -> bool:
    # сейчас "partial" = методы с контролем; остальные control игнорируют
    return "_partial" in variant.lower()


METHOD_INFO: Dict[str, Dict[str, str]] = {
    "correlation_full": {
        "title": "Корреляция (полная)",
        "meaning": "Линейная связь. Значение в [-1, 1]. |value| ближе к 1 = сильнее.",
    },
    "correlation_partial": {
        "title": "Частичная корреляция",
        "meaning": "Линейная связь при контроле остальных переменных. [-1, 1].",
    },
    "correlation_directed": {
        "title": "Лаговая корреляция (directed)",
        "meaning": "Оценка направленной связи через сдвиг по лагу. Чем больше |value|, тем сильнее.",
    },
    "mutinf_full": {
        "title": "Взаимная информация (MI)",
        "meaning": "Нелинейная зависимость. >= 0. Больше = сильнее.",
    },
    "mutinf_partial": {
        "title": "Частичная MI",
        "meaning": "MI при контроле переменных. >= 0. Больше = сильнее.",
    },
    "coherence_full": {
        "title": "Когерентность",
        "meaning": "Частотная синхронизация. Обычно в [0, 1]. Больше = сильнее.",
    },
    "h2_full": {"title": "H2 (полная)", "meaning": "Нелинейная связность. Обычно в [0, 1]. Больше = сильнее."},
    "h2_partial": {"title": "H2 (partial)", "meaning": "H2 при контроле. Обычно в [0, 1]. Больше = сильнее."},
    "h2_directed": {"title": "H2 (directed)", "meaning": "Направленная H2. Больше = сильнее."},
    "granger_full": {"title": "Granger (p-values)", "meaning": "p-value теста. Меньше = сильнее свидетельство причинности."},
    "granger_partial": {"title": "Granger partial (p-values)", "meaning": "Granger partial (linear control; best lag up to L): p-value после удаления влияния control. Меньше = сильнее."},
    "granger_directed": {"title": "Granger directed (p-values)", "meaning": "То же семейство p-values. Меньше = сильнее."},
    "te_full": {"title": "Transfer Entropy", "meaning": "Направленный поток информации. Больше = сильнее."},
    "te_partial": {"title": "Transfer Entropy (partial)", "meaning": "TE при контроле. Больше = сильнее."},
    "te_directed": {"title": "Transfer Entropy (directed)", "meaning": "TE (directed). Больше = сильнее."},
    "ah_full": {"title": "AH (directed)", "meaning": "Нелинейная направленная мера. Больше = сильнее."},
    "ah_partial": {"title": "AH (partial)", "meaning": "AH при контроле. Больше = сильнее."},
    "ah_directed": {"title": "AH (directed)", "meaning": "AH (directed). Больше = сильнее."},
}


def _is_pvalue_method(variant: str) -> bool:
    return is_pvalue_method(variant)


def _is_directed_method(variant: str) -> bool:
    return is_directed_method(variant)

def _lag_quality(variant: str, mat: np.ndarray) -> float:
    """Скалярная метрика качества лага: больше => лучше (единая конвенция)."""
    if mat is None or not isinstance(mat, np.ndarray) or mat.size == 0:
        return np.nan
    n = mat.shape[0]
    if n < 2:
        return np.nan
    mask = ~np.eye(n, dtype=bool)
    vals = mat[mask]
    vals = vals[np.isfinite(vals)]
    if vals.size == 0:
        return np.nan
    if _is_pvalue_method(variant):
        # p-value: меньше лучше -> переводим в "evidence" (больше лучше)
        vals = np.clip(vals, 1e-12, 1.0)
        return float(np.mean(-np.log10(vals)))
    return float(np.mean(np.abs(vals)))


def get_method_spec(variant: str) -> MethodSpec:
    """Совместимость со старым API: возвращает MethodSpec на основе новых семантических множеств."""
    return MethodSpec(
        directed=is_directed_method(variant),
        is_p_value=is_pvalue_method(variant),
        control_dependent=is_control_sensitive_method(variant),
        supports_lag=is_directed_method(variant),
    )


def _residualize_series(y: np.ndarray, X: np.ndarray) -> np.ndarray:
    """Возвращает остатки регрессии y ~ X (с константой)."""
    y = np.asarray(y, dtype=float)
    if X is None or np.size(X) == 0:
        return y - np.nanmean(y)
    X = np.asarray(X, dtype=float)

    # синхронно выкидываем nan
    mask = np.isfinite(y)
    if X.ndim == 1:
        mask &= np.isfinite(X)
    else:
        mask &= np.all(np.isfinite(X), axis=1)

    y2 = y[mask]
    X2 = X[mask]
    if y2.size < 5:
        return y - np.nanmean(y)

    if X2.ndim == 1:
        X2 = X2.reshape(-1, 1)
    X2 = np.column_stack([np.ones(len(X2)), X2])

    try:
        beta, *_ = np.linalg.lstsq(X2, y2, rcond=None)
        resid = np.full_like(y, np.nan, dtype=float)
        resid[mask] = y2 - (X2 @ beta)
        m = np.nanmean(resid)
        resid = np.where(np.isfinite(resid), resid, m)
        return resid
    except Exception:
        return y - np.nanmean(y)


def _pairwise_partial_value(
    data: pd.DataFrame,
    i: int,
    j: int,
    controls: List[int],
    *,
    metric: str,
    lag: int = 1,
) -> float:
    """
    Частичная мера между (i,j) с индивидуальным набором контролей (по индексам колонок).
    metric: 'corr' | 'h2' | 'mi'
    Используем directed-сдвиг: x[t] -> y[t+lag].
    """
    x = data.iloc[:, i].to_numpy(dtype=float)
    y = data.iloc[:, j].to_numpy(dtype=float)

    lag = max(1, int(lag or 1))
    if lag >= len(x) or lag >= len(y):
        return np.nan

    x2 = x[:-lag]
    y2 = y[lag:]

    if controls:
        Z = data.iloc[:, controls].to_numpy(dtype=float)
        Z2 = Z[:-lag, :]
        rx = _residualize_series(x2, Z2)
        ry = _residualize_series(y2, Z2)
    else:
        rx = x2 - np.nanmean(x2)
        ry = y2 - np.nanmean(y2)

    if rx.size < 2 or ry.size < 2:
        return np.nan

    if metric == "corr":
        return float(np.corrcoef(rx, ry)[0, 1])
    if metric == "h2":
        c = float(np.corrcoef(rx, ry)[0, 1])
        return float(c * c)
    if metric == "mi":
        a = rx.reshape(-1, 1)
        b = ry.reshape(-1, 1)
        return float(_knn_mutual_info(a, b, k=DEFAULT_K_MI))
    return np.nan


def _pairwise_partial_matrix(
    data: pd.DataFrame,
    *,
    metric: str,
    lag: int,
    policy: str,
    custom_controls: Optional[List[str]] = None,
) -> np.ndarray:
    """
    NxN матрица частичной меры, где control-set выбирается ПО ПАРЕ.
    policy:
      - 'others': контролируем все остальные переменные (кроме пары)
      - 'custom': контролируем выбранный список custom_controls (пересечённый с остальными)
      - 'none'  : без контролей
    """
    cols = list(data.columns)
    n = len(cols)
    M = np.full((n, n), np.nan, dtype=float)
    for i in range(n):
        M[i, i] = 1.0

    name_to_idx = {c: k for k, c in enumerate(cols)}
    custom_controls = list(custom_controls) if custom_controls else []
    custom_idx = [name_to_idx[c] for c in custom_controls if c in name_to_idx]

    for i in range(n):
        for j in range(i + 1, n):
            if policy == "others":
                controls = [k for k in range(n) if k not in (i, j)]
            elif policy == "custom":
                controls = [k for k in custom_idx if k not in (i, j)]
            elif policy == "none":
                controls = []
            else:
                controls = []
            v = _pairwise_partial_value(data, i, j, controls, metric=metric, lag=lag)
            M[i, j] = v
            M[j, i] = v
    return M

def compute_connectivity_variant(
    data,
    variant,
    lag=1,
    control=None,
    *,
    partial_mode: str = "global",
    pairwise_policy: str = "others",
    custom_controls: Optional[List[str]] = None,
):
    """
    variant: ключ из method_mapping.
    control: список колонок для GLOBAL partial (как раньше).
    partial_mode:
      - "global"   : прежнее поведение (control применяется ко всей матрице)
      - "pairwise" : control-set выбирается по паре (для *_partial методов, где это реализовано)
    pairwise_policy (для partial_mode="pairwise"):
      - "others" : контролируем все остальные переменные (по умолчанию)
      - "custom" : контролируем custom_controls (пересечение с остальными)
      - "none"   : без контролей
    """
    try:
        if control is not None and len(control) == 0:
            control = None

        # pairwise-partial (нужно для N=3..4)
        if partial_mode == "pairwise" and isinstance(variant, str) and variant.endswith("_partial"):
            if variant == "correlation_partial":
                return _pairwise_partial_matrix(
                    data,
                    metric="corr",
                    lag=max(1, int(lag or 1)),
                    policy=pairwise_policy,
                    custom_controls=custom_controls,
                )
            if variant == "h2_partial":
                return _pairwise_partial_matrix(
                    data,
                    metric="h2",
                    lag=max(1, int(lag or 1)),
                    policy=pairwise_policy,
                    custom_controls=custom_controls,
                )
            if variant == "mutinf_partial":
                return _pairwise_partial_matrix(
                    data,
                    metric="mi",
                    lag=max(1, int(lag or 1)),
                    policy=pairwise_policy,
                    custom_controls=custom_controls,
                )
            # Остальные partial пока считаем GLOBAL-режимом (через control):
            # корректное "pairwise partial" для TE/AH/Granger требует отдельной
            # условной постановки и другого API.

        if variant in method_mapping:
            return method_mapping[variant](data, lag, control)

        return correlation_matrix(data)
    except Exception as e:
        logging.error(f"[ComputeVariant] Метод {variant} не работает: {e}")
        return None


def powerset(iterable):
    s = list(iterable)
    return chain.from_iterable(combinations(s, r) for r in range(len(s)+1))

##############################################
#  диагностика коэффициентов регрессии
##############################################
def regression_diagnostics(df: pd.DataFrame, target: str, controls: list):
    """
    Рассчитывает линейную регрессию вида: target ~ controls.
    Возвращает строку с R².
    """
    # Если нет контрольных переменных — выходим
    if not controls:
        return f"Нет контрольных переменных для {target}."
    # Иначе строим модель и возвращаем R²
    X = df[controls]
    y = df[target]
    model = LinearRegression().fit(X, y)
    r2 = model.score(X, y)
    return f"{target} ~ {controls}: R² = {r2:.3f}"


##############################################
# Частотный анализ: возвращает пиковые значения
##############################################
def frequency_analysis(series: pd.Series, peak_height_ratio: float = 0.2, *, fs: float = 1.0):
    freqs, amplitude, phase, peaks = plt_fft_analysis(series, fs=fs)
    if freqs.size == 0 or peaks.size == 0:
        return None, None, None
    peak_freqs = freqs[peaks]
    peak_amps = amplitude[peaks]
    periods = 1 / peak_freqs
    return peak_freqs, peak_amps, periods

def sliding_fft_analysis(data: pd.DataFrame, window_size: int, overlap: int) -> dict:
    """Экспериментальный анализ скользящего FFT (по умолчанию отключён)."""
    logging.info("[Sliding FFT] Экспериментальная функция отключена.")
    return {}


def analyze_sliding_windows_with_metric(
    data: pd.DataFrame,
    variant: str,
    window_size: int,
    stride: int,
    *,
    lag: int = 1,
) -> dict:
    """Анализ скользящих окон для заданного window_size.

    Возвращает структуру для HTML-отчёта:
      {
        "best_window": {"start": int, "end": int, "metric": float, "matrix": np.ndarray},
        "curve": {"x": [start_idx...], "y": [metric...]}
      }

    stride — шаг сдвига окна (в точках), lag — лаг (только для lagged-методов).
    """
    if data is None or data.empty:
        return {}

    n = len(data)
    w = int(max(2, min(window_size, n)))
    s = int(max(1, stride))

    xs: List[int] = []
    ys: List[float] = []

    best = {
        "start": 0,
        "end": w,
        "metric": float("-inf"),
        "matrix": None,
    }

    # Ограничиваем количество окон, чтобы не взорваться по времени на длинных рядах.
    # Это не запрещает пользователю уменьшить stride, но даёт безопасный дефолт.
    max_windows = 400
    starts = list(range(0, max(1, n - w + 1), s))
    if len(starts) > max_windows:
        # равномерная подвыборка
        idx = np.linspace(0, len(starts) - 1, max_windows).round().astype(int)
        starts = [starts[i] for i in idx]

    for start in starts:
        end = start + w
        if end > n:
            break
        chunk = data.iloc[start:end]
        try:
            mat = compute_connectivity_variant(chunk, variant, lag=int(max(1, lag)))
            score = _lag_quality(variant, mat)
        except Exception as ex:
            logging.error(f"[SlidingWindow] {variant} win={w} start={start}: {ex}")
            mat = None
            score = float("nan")

        xs.append(int(start))
        ys.append(float(score) if np.isfinite(score) else float("nan"))

        if np.isfinite(score) and float(score) > float(best["metric"]):
            best = {
                "start": int(start),
                "end": int(end),
                "metric": float(score),
                "matrix": mat,
            }

    return {
        "best_window": best,
        "curve": {"x": xs, "y": ys},
    }


def sliding_window_pairwise_analysis(
    data: pd.DataFrame,
    method: str,
    window_size: int,
    overlap: int,
) -> dict:
    """Экспериментальный парный анализ скользящих окон (по умолчанию отключён)."""
    logging.info("[Sliding Pairwise] Экспериментальная функция отключена.")
    return {}

##############################################
# Листы с коэффициентами и частотным анализом
##############################################
def export_coefficients_sheet(tool, wb: Workbook):
    ws = wb.create_sheet("Coefficients & Explanations")
    ws.append(["Описание:", "Лист содержит краткие пояснения коэффициентов регрессий и матриц связей."])
    ws.append(["Например, коэффициенты регрессии показывают, как контрольные переменные влияют на связь между переменными."])
    ws.append([])
    ws.append(["Регрессионная диагностика:"])
    ws.append(["Переменная", "Контроль", "Диагностика"])
    for target in tool.data.columns:
        controls = [c for c in tool.data.columns if c != target]
        diag_str = regression_diagnostics(tool.data, target, controls)
        ws.append([target, str(controls), diag_str])
    ws.append([])
    ws.append(["Матрицы связей:"])
    ws.append(["Метод", "Описание"])
    methods_info = [
        ("correlation_full", "Стандартная корреляционная матрица."),
        ("correlation_partial", "Частичная корреляция (с контролем)."),
        ("mutinf_full", "Полная взаимная информация."),
        ("coherence_full", "Когерентность между переменными.")
    ]
    for m, info in methods_info:
        ws.append([m, info])
    logging.info("[Coefficients] Лист 'Coefficients & Explanations' сформирован.")

def export_frequency_summary_sheet(tool, wb: Workbook):
    ws = wb.create_sheet("Frequency Summary")
    ws.append(["Столбец", "Пиковые частоты", "Пиковые амплитуды", "Периоды", "Пояснение"])
    for col in tool.data.columns:
        s = tool.data[col].dropna()
        freq, amps, periods = frequency_analysis(s)
        if freq is not None:
            freq_str = ", ".join([f"{f:.3f}" for f in freq])
            amps_str = ", ".join([f"{f:.3f}" for f in amps])
            period_str = ", ".join([f"{p:.1f}" for p in periods])
            note = f"Макс. связь на {freq[np.argmax(amps)]:.3f} Hz"
        else:
            freq_str = amps_str = period_str = "Нет пиков"
            note = "Пиковые частоты не выявлены"
        ws.append([col, freq_str, amps_str, period_str, note])
    for col in ws.columns:
        max_length = max(len(str(cell.value)) for cell in col if cell.value is not None)
        ws.column_dimensions[get_column_letter(col[0].column)].width = max_length
    logging.info("[Frequency] Лист 'Frequency Summary' сформирован.")

##############################################
# Новый лист: Индивидуальные АЧХ и ФЧХ (раздельно)
##############################################
def export_individual_ac_ph_sheet(tool, wb: Workbook):
    ws = wb.create_sheet("Individual AC & PH")
    ws.append(["Столбец", "АЧХ", "ФЧХ"])
    plots = plot_individual_ac_ph(tool.data_normalized, "Individual AC & PH")
    for col, imgs in plots.items():
        ws.append([col])
        img_ac = Image(imgs["AC"])
        img_ac.width = 400
        img_ac.height = 300
        ws.add_image(img_ac, f"B{ws.max_row}")
        img_ph = Image(imgs["PH"])
        img_ph.width = 400
        img_ph.height = 300
        ws.add_image(img_ph, f"G{ws.max_row}")
    logging.info("[Individual AC & PH] Лист сформирован.")

##############################################
# Новый лист: Анализ энтропии
##############################################
def export_entropy_sheet(tool, wb: Workbook):
    ws = wb.create_sheet("Entropy Analysis")
    ws.append(["Столбец", "Sample Entropy (sampen)"])
    for col in tool.data.columns:
        s = tool.data[col].dropna()
        ent = compute_sample_entropy(s)
        ws.append([col, f"{ent:.3f}" if not np.isnan(ent) else "N/A"])
    logging.info("[Entropy Analysis] Лист сформирован.")

##############################################
# Новый лист
##############################################
def export_combined_informational_sheet(tool, wb: Workbook):
    ws = wb.create_sheet("Combined Informational Analysis")
    current_row = 1
    ws.cell(row=current_row, column=1, value="Combined Informational Analysis")
    current_row += 2
    ws.cell(row=current_row, column=1, value="Lag Analysis Summary (Aggregated)")
    current_row += 1
    buf_lag = tool.plot_all_methods_lag_comparison(tool.lag_results)
    img_lag = Image(buf_lag)
    img_lag.width = 800
    img_lag.height = 600
    ws.add_image(img_lag, f"A{current_row}")
    current_row += 30
    ws.cell(row=current_row, column=1, value="Sliding Window Analysis Summary (Aggregated)")
    current_row += 1
    sw_res = tool.analyze_sliding_windows(
        "coherence_full",
        window_size=min(50, len(tool.data_normalized) // 2),
        overlap=min(25, len(tool.data_normalized) // 4),
    )
    if sw_res:
        legend_text = "Метод: coherence_full, Окно: 50"
        buf_sw = tool.plot_sliding_window_comparison(sw_res, legend_text=legend_text)
        img_sw = Image(buf_sw)
        img_sw.width = 700
        img_sw.height = 400
        ws.add_image(img_sw, f"A{current_row}")
        current_row += 20
    else:
        ws.append(["Sliding Window Analysis отключён или нет данных."])
        current_row += 2
    ws.cell(row=current_row, column=1, value="Pairwise Lag Analysis (пример для первой пары)")
    current_row += 1
    if len(tool.data.columns) >= 2:
        pair = list(combinations(tool.data.columns, 2))[0]
        col1, col2 = pair
        series1 = tool.data[col1].dropna().values
        series2 = tool.data[col2].dropna().values
        n = min(len(series1), len(series2))
        lag_metrics = {}
        for lag in range(1, 21):
            if n > lag:
                corr = np.corrcoef(series1[lag:], series2[:n-lag])[0, 1] if len(series1[lag:]) > 1 and len(series2[:n-lag]) > 1 else np.nan
                lag_metrics[lag] = corr
        if lag_metrics:
            lags = list(lag_metrics.keys())
            correlations = [lag_metrics[lag] for lag in lags]
            fig, ax = plt.subplots(figsize=(4, 3))
            ax.plot(lags, correlations, marker='o')
            ax.set_title(f"Lag Analysis: {col1}-{col2}")
            legend_text_pair = f"Пара: {col1}-{col2}, Метод: Lag Correlation"
            ax.text(0.5, 0.1, legend_text_pair, transform=ax.transAxes, fontsize=8, 
                    verticalalignment='bottom', bbox=dict(facecolor='white', alpha=0.5))
            buf_plag = BytesIO()
            plt.tight_layout()
            plt.savefig(buf_plag, format="png", dpi=100)
            buf_plag.seek(0)
            plt.close(fig)
            img_plag = Image(buf_plag)
            img_plag.width = 400
            img_plag.height = 300
            ws.add_image(img_plag, f"A{current_row}")
        else:
            ws.append(["Недостаточно данных для Pairwise Lag Analysis."])
        current_row += 20
        # (пример для первой пары)
        ws.cell(row=current_row, column=1, value="Extended Spectral Analysis (пример для первой пары)")
        current_row += 1
        title_es = f"Coherence {col1}-{col2}"
        buf_es = plot_coherence_vs_frequency(tool.data[col1], tool.data[col2], title_es, fs=getattr(tool, "fs", 1.0))
        img_es = Image(buf_es)
        img_es.width = 400
        img_es.height = 300
        ws.add_image(img_es, f"A{current_row}")
        current_row += 20
        # (пример для первой пары)
        ws.cell(row=current_row, column=1, value="Frequency Demonstration (пример для первой пары)")
        current_row += 1
        buf_fd = plot_coherence_vs_frequency(tool.data[col1], tool.data[col2], title_es, fs=getattr(tool, "fs", 1.0))
        img_fd = Image(buf_fd)
        img_fd.width = 400
        img_fd.height = 300
        ws.add_image(img_fd, f"A{current_row}")
        current_row += 20
    else:
        ws.append(["Недостаточно столбцов для Pairwise Lag Analysis или Spectral Analysis."])
        current_row += 60 
    ws.cell(row=current_row, column=1, value="End of Combined Informational Analysis")
    logging.info("[Combined Informational Analysis] Лист сформирован.")

##############################################
#Combined Time Series (агрегированный + индивидуальные графики)
##############################################
def export_combined_ts_sheet(tool, wb: Workbook):
    ws = wb.create_sheet("Combined Time Series")
    ws.append(["Aggregated Time Series: Оригинальные и Нормализованные (на одном графике)"])
    buf_orig = tool.plot_time_series(tool.data, "Aggregated Original Time Series")
    img_orig = Image(buf_orig)
    img_orig.width = 600
    img_orig.height = 300
    ws.add_image(img_orig, "A2")
    
    ws.append([])
    buf_norm = tool.plot_time_series(tool.data_normalized, "Aggregated Normalized Time Series")
    img_norm = Image(buf_norm)
    img_norm.width = 600
    img_norm.height = 300
    ws.add_image(img_norm, "A10")
    
    ws.append(["Individual Time Series Plots:"])
    row = ws.max_row + 2
    for col in tool.data.columns:
        buf_ind = tool.plot_single_time_series(tool.data[col], f"Original: {col}")
        img_ind = Image(buf_ind)
        img_ind.width = 300
        img_ind.height = 200
        ws.add_image(img_ind, f"A{row}")
        buf_ind_norm = tool.plot_single_time_series(tool.data_normalized[col], f"Normalized: {col}")
        img_ind_norm = Image(buf_ind_norm)
        img_ind_norm.width = 300
        img_ind_norm.height = 200
        ws.add_image(img_ind_norm, f"E{row}")
        row += 15
    logging.info("[Combined Time Series] Лист сформирован.")

##############################################
# Combined FFT (агрегированный + индивидуальные графики)
##############################################
def export_all_fft_sheet(tool, wb: Workbook):
    ws = wb.create_sheet("Combined FFT")
    ws.append(["Combined FFT Analysis (Aggregated) - Original"])
    buf_fft_orig = tool.plot_fft(tool.data, "Aggregated Original FFT")
    img_fft_orig = Image(buf_fft_orig)
    img_fft_orig.width = 600
    img_fft_orig.height = 400
    ws.add_image(img_fft_orig, "A2")
    
    ws.append([])
    ws.append(["Combined FFT Analysis (Aggregated) - Normalized"])
    buf_fft_norm = tool.plot_fft(tool.data_normalized, "Aggregated Normalized FFT")
    img_fft_norm = Image(buf_fft_norm)
    img_fft_norm.width = 600
    img_fft_norm.height = 400
    ws.add_image(img_fft_norm, "A20")
    
    ws.append(["Individual FFT Analysis:"])
    row = ws.max_row + 2
    for col in tool.data.columns:
        buf_fft_ind = tool.plot_single_fft(tool.data[col], f"Original FFT: {col}")
        img_fft_ind = Image(buf_fft_ind)
        img_fft_ind.width = 300
        img_fft_ind.height = 200
        ws.add_image(img_fft_ind, f"A{row}")
        buf_fft_ind_norm = tool.plot_single_fft(tool.data_normalized[col], f"Normalized FFT: {col}")
        img_fft_ind_norm = Image(buf_fft_ind_norm)
        img_fft_ind_norm.width = 300
        img_fft_ind_norm.height = 200
        ws.add_image(img_fft_ind_norm, f"E{row}")
        row += 15
    logging.info("[Combined FFT] Лист сформирован.")

##############################################
# Создание оглавления с гиперссылками
##############################################
def create_table_of_contents(wb: Workbook):
    if "Table of Contents" in wb.sheetnames:
        old_sheet = wb["Table of Contents"]
        wb.remove(old_sheet)
    toc = wb.create_sheet("Table of Contents", 0)
    row = 1
    for sheet_name in wb.sheetnames:
        if sheet_name == "Table of Contents":
            continue
        link = f"#{sheet_name}!A1"
        cell = toc.cell(row=row, column=1)
        cell.value = sheet_name
        cell.hyperlink = link
        cell.style = "Hyperlink"
        row += 1

##############################################
# Класс BigMasterTool 
##############################################
class BigMasterTool:
    def __init__(self, data: pd.DataFrame = None, enable_experimental: bool = False) -> None:
        if data is not None:
            # Удаляем полностью константные колонки (точное равенство всем значениям первой строки)
            try:
                data = data.loc[:, (data != data.iloc[0]).any()]
            except Exception:
                pass

            # Жёстко приводим к числам: если колонка строковая, но содержит числа — она станет numeric.
            self.data = data.copy()
            for c in list(self.data.columns):
                self.data[c] = pd.to_numeric(self.data[c], errors="coerce")

            # Выкидываем колонки, где почти всё NaN после приведения
            good_cols = [c for c in self.data.columns if self.data[c].notna().mean() >= 0.2]
            self.data = self.data[good_cols]

            # Если после всего получилось пусто — оставляем пустой DataFrame
            if self.data.shape[1] > 0:
                self.data.columns = [f"c{i+1}" for i in range(self.data.shape[1])]
        else:
            self.data = pd.DataFrame()

        self.data_normalized: pd.DataFrame = pd.DataFrame()
        self.results: dict = {}
        self.lag_results: dict = {}
        self.fft_results: dict = {}
        self.data_type: str = 'unknown'
        self.enable_experimental = enable_experimental
        # Частота дискретизации (Гц) для FFT/когерентности/частотных отчётов.
        # По умолчанию fs=1.0: частоты интерпретируются как "циклы на отсчёт".
        self.fs: float = 1.0
        self.variant_lags: dict = {}
        self.lag_ranges = {v: range(1, 21) for v in method_mapping}
        # Контрольные наборы для partial-методов: стратегия и предохранитель от экспоненты
        # strategy: 'all_others' (по умолчанию), 'none', 'powerset'
        self.control_strategy: str = 'all_others'
        self.max_control_sets: int = 32
        # Коррекция множественных сравнений для p-value методов: 'none'|'fdr_bh'
        self.pvalue_correction: str = 'none'

        self.undirected_methods = [m for m in method_mapping if not get_method_spec(m).directed]
        self.directed_methods = [m for m in method_mapping if get_method_spec(m).directed]

    def _ensure_pairwise_cache(self) -> None:
        """
        Ленивая подготовка тяжёлых pairwise-таблиц (Undirected/Directed Methods).
        """
        if self.data_normalized.empty:
            self.normalize_data()
        if not self.undirected_rows or not self.directed_rows:
            self.prepare_pairs()
        if not self.connectivity_matrices:
            self.compute_all_matrices()

    def _get_cached_matrix(self, method: str, control_set: List[str]) -> Optional[np.ndarray]:
        mats = self.connectivity_matrices.get(method, {})
        key = frozenset(control_set) if control_set else frozenset()
        if key in mats:
            return mats[key]
        # если метод не зависит от control, используем базовую матрицу
        if not is_control_sensitive_method(method):
            return mats.get(frozenset(), None)
        return None

    def load_data_excel(
        self,
        filepath: str,
        *,
        header: str = "auto",
        time_col: str = "auto",
        transpose: str = "auto",
        preprocess: bool = True,
        log_transform: bool = False,
        remove_outliers: bool = True,
        normalize: bool = True,
        fill_missing: bool = True,
        check_stationarity: bool = False,
    ) -> pd.DataFrame:
        """Загружает данные из файла и готовит матрицу time×features для расчётов."""
        self.data = load_or_generate(
            filepath,
            header=header,
            time_col=time_col,
            transpose=transpose,
            preprocess=preprocess,
            log_transform=log_transform,
            remove_outliers=remove_outliers,
            normalize=normalize,
            fill_missing=fill_missing,
            check_stationarity=check_stationarity,
        )
        self.raw_data = self.data.copy()
        self.data_normalized = self.data.copy()  # В текущей архитектуре "normalized" совпадает с предобработанными данными.
        self.data = self.data.fillna(self.data.mean(numeric_only=True))
        self.data_type = 'file'
        logging.info(f"[BigMasterTool] Данные загружены, shape = {self.data.shape}.")
        return self.data

    def normalize_data(self):
        if self.data is None or self.data.empty or self.data.shape[1] == 0:
            logging.warning("[BigMasterTool] normalize_data: нет данных для нормализации.")
            self.data_normalized = pd.DataFrame()
            return

        cols_to_norm = [c for c in self.data.columns if pd.api.types.is_numeric_dtype(self.data[c])]
        if not cols_to_norm:
            self.data_normalized = self.data.copy() 
            logging.warning("[BigMasterTool] normalize_data: нет числовых колонок для нормализации.")
            return

        sc = StandardScaler()
        self.data_normalized = self.data.copy()
        self.data_normalized[cols_to_norm] = sc.fit_transform(self.data[cols_to_norm])
        logging.info("[BigMasterTool] Данные нормализованы.")


    def detect_outliers(self, col: str, thresh: float = 3, use_normalized: bool = False) -> np.ndarray:
        data = self.data_normalized if use_normalized else self.data
        if col not in data.columns or not pd.api.types.is_numeric_dtype(data[col]):
            return np.array([])
        arr = data[col].dropna().values
        if len(arr) == 0:
            return np.array([])
        z_scores = np.abs(stats.zscore(arr))
        return np.where(z_scores > thresh)[0]

    def remove_outliers(self, thresh: float = 5) -> None:
        if self.data.empty: return
        orig_shape = self.data.shape
        all_outl_indices = set()
        for col in self.data.columns:
            outl = self.detect_outliers(col, thresh=thresh, use_normalized=False)
            if len(outl) > 0:
                logging.info(f"[Outliers] '{col}': найдено {len(outl)} выбросов (z>{thresh}).")
                all_outl_indices.update(self.data.index[outl].tolist())
            else:
                logging.info(f"[Outliers] '{col}': выбросов нет.")
        
        if all_outl_indices:
            self.data = self.data.drop(list(all_outl_indices)).reset_index(drop=True)
            new_shape = self.data.shape
            logging.info(f"[Outliers] Размер данных {orig_shape} => {new_shape}.")
            self.normalize_data()
        else:
            logging.info("[Outliers] Данные без изменений.")

    def optimize_lag(self, variant: str, candidate_lags: range = range(1, 21)) -> dict:
        original_data = self.data_normalized 
        lag_metrics = {}
        spec = get_method_spec(variant)
        for lag in tqdm(candidate_lags, desc=f"Optimizing lag for {variant}"):
            try:
                mat = compute_connectivity_variant(original_data, variant, lag)
                metric = _lag_quality(variant, mat)
                lag_metrics[lag] = (metric, mat)
            except Exception as ex:
                logging.error(f"[Lag] Ошибка {variant} lag={lag}: {ex}")
                num_cols = original_data.shape[1] if not original_data.empty else 0
                lag_metrics[lag] = (np.nan, np.full((num_cols, num_cols), np.nan))
        return lag_metrics 

    def analyze_lags(self, variant: str, candidate_lags: range = None) -> dict:
        c_lags = self.lag_ranges.get(variant, range(1, 21)) if candidate_lags is None else candidate_lags
        return self.optimize_lag(variant, c_lags)

    def compute_all_matrices(self):
        if self.data_normalized.empty:
            self.normalize_data()
        if self.data_normalized.empty:
            logging.warning("[Matrices] Нет данных для вычисления матриц.")
            return

        # Готовим пары (и ограниченный список control-set'ов)
        if not hasattr(self, "undirected_rows") or not hasattr(self, "directed_rows"):
            self.prepare_pairs()
        if not getattr(self, "undirected_rows", None) or not getattr(self, "directed_rows", None):
            self.prepare_pairs()

        required = set()
        for _, S in getattr(self, "undirected_rows", []):
            required.add(frozenset(S) if S else frozenset())
        for _, S in getattr(self, "directed_rows", []):
            required.add(frozenset(S) if S else frozenset())

        # предохранитель: если по какой-то причине required пустой — НЕ делаем powerset
        if not required:
            required = {frozenset()}  # без контроля

        self.all_control_sets = [list(s) for s in sorted(required, key=lambda x: (len(x), sorted(list(x))))]

        self.connectivity_matrices = {}
        for method in method_mapping.keys():
            self.connectivity_matrices[method] = {}

            if is_control_sensitive_method(method):
                for S in self.all_control_sets:
                    mat = compute_connectivity_variant(self.data_normalized, method, lag=1, control=S)
                    if mat is None:
                        continue
                    # Коррекция p-values (если включена)
                    if getattr(self, "pvalue_correction", "none") == "fdr_bh" and is_pvalue_method(method):
                        mat = apply_pvalue_correction_matrix(mat, directed=get_method_spec(method).directed)
                    self.connectivity_matrices[method][frozenset(S)] = mat
            else:
                mat = compute_connectivity_variant(self.data_normalized, method, lag=1, control=None)
                if mat is None:
                    continue
                if getattr(self, "pvalue_correction", "none") == "fdr_bh" and is_pvalue_method(method):
                    mat = apply_pvalue_correction_matrix(mat, directed=get_method_spec(method).directed)
                self.connectivity_matrices[method][frozenset()] = mat

        logging.info("[Matrices] Все матрицы вычислены.")

    def prepare_pairs(self):
        """Готовит список пар и допустимых control-set'ов.

        Старое поведение делало полный powerset(others) для каждой пары => экспонента.
        Сейчас есть предохранитель:
        - по умолчанию используем один control-set: все остальные переменные (all_others);
        - если включён режим powerset, ограничиваем число наборов self.max_control_sets.
        """
        if self.data.empty:
            return

        cols = list(self.data.columns)
        self.undirected_pairs = list(combinations(cols, 2))
        self.directed_pairs = list(permutations(cols, 2))

        # настройки (можно менять после создания объекта)
        strategy = getattr(self, "control_strategy", "all_others")
        max_sets = int(getattr(self, "max_control_sets", 32))
        max_sets = max(1, max_sets)

        def _control_sets(others: List[str]) -> List[List[str]]:
            if not others:
                return [[]]
            if strategy == "none":
                return [[]]
            if strategy == "all_others":
                return [list(others)]
            if strategy == "powerset":
                out: List[List[str]] = []
                # обязательно включаем пустой и полный
                out.append([])
                if len(out) < max_sets:
                    out.append(list(others))
                # дальше добавляем подмножества по возрастанию размера, но с лимитом
                for r in range(1, len(others)):
                    for comb in combinations(others, r):
                        out.append(list(comb))
                        if len(out) >= max_sets:
                            return out
                return out
            # fallback
            return [list(others)]

        self.undirected_rows = []
        for pair in self.undirected_pairs:
            others = [c for c in cols if c not in pair]
            for S in _control_sets(others):
                self.undirected_rows.append((pair, S))

        self.directed_rows = []
        for pair in self.directed_pairs:
            others = [c for c in cols if c not in pair]
            for S in _control_sets(others):
                self.directed_rows.append((pair, S))

        if strategy == "powerset" and len(cols) >= 10:
            logging.warning(
                f"[Pairs] powerset режим ограничен max_control_sets={max_sets}. "
                "Полный powerset слишком дорогой."
            )
        logging.info("[Pairs] Пары сформированы.")

    def get_undirected_value(self, mat, var1, var2, indices):
        if mat is None: return np.nan
        i, j = indices[var1], indices[var2]
        return mat[min(i, j), max(i, j)]

    def get_directed_value(self, mat, src, tgt, indices):
        if mat is None: return np.nan
        i, j = indices[src], indices[tgt]
        return mat[i, j]


    def run_all_methods(self, precompute_controls: bool = False, precompute_pairs: bool = False) -> None:
        self.normalize_data()
        if self.data_normalized.empty:
            logging.warning("[RunAll] Нет данных для выполнения анализа.")
            return

        self.results = {}
        self.lag_results = {}
        if self.enable_experimental:
            self.fft_results = sliding_fft_analysis(
                self.data_normalized,
                window_size=min(200, len(self.data_normalized) // 2),
                overlap=min(100, len(self.data_normalized) // 4),
            )
        else:
            self.fft_results = {}
        
                # Базовые результаты на лаге 1 (все методы из method_mapping).
        methods_to_run = list(method_mapping.keys())
        for variant in methods_to_run:
            self.results[variant] = compute_connectivity_variant(self.data_normalized, variant, lag=1)

        # Анализ лагов только для направленных методов, где лаг поддерживается.
        lag_methods = [
            m for m in methods_to_run
            if get_method_spec(m).directed and get_method_spec(m).supports_lag
        ]
        for variant in lag_methods:
            self.lag_results[variant] = self.analyze_lags(
                variant,
                self.lag_ranges.get(variant, range(1, 21)),
            )

        if precompute_controls:
            self.compute_all_matrices()
        if precompute_pairs:
            self.prepare_pairs()
        logging.info("[RunAll] Все методы завершены.")


    def run_selected_methods(
        self,
        variants: List[str],
        *,
        max_lag: int = DEFAULT_MAX_LAG,
        lag_step: int = 1,
        compute_lag_sweep: bool = True,
        pick_best_lag: bool = True,
        # window sweep (для всех методов)
        window_min: Optional[int] = None,
        window_max: Optional[int] = None,
        window_step: int = 100,
        window_stride: Optional[int] = None,
        # выбор переменных/пар и partial-control
        include_vars: Optional[List[str]] = None,
        pair_filter: Optional[List[Tuple[str, str]]] = None,
        partial_mode: str = "global",            # "global" | "pairwise"
        pairwise_policy: str = "others",         # "others" | "custom" | "none"
        partial_controls: Optional[List[str]] = None,  # для global
        custom_controls: Optional[List[str]] = None,   # для pairwise_policy="custom"

        # legacy flags (оставлены для совместимости; предпочтительнее параметры выше)
        sliding_windows: bool = False,
        window_sizes: Optional[List[int]] = None,
        overlap: int = 50,
    ) -> Dict[str, int]:
        """Запускает выбранные методы.

        Возвращает: dict variant -> выбранный lag (или 1).

        Диагностика для отчёта:
          - self.lag_sweep[variant][lag] = {"score": float, "matrix": np.ndarray}
          - self.window_sweep[variant][w] = {"best": {...}, "curve": {...}}
          - self.best_window[variant] = {...}
        """
        self.normalize_data()
        if self.data_normalized.empty:
            logging.warning("[RunSelected] Нет данных.")
            self.results = {}
            self.variant_lags = {}
            return {}


        # Фильтр переменных (например, считать только c1,c3,c4)
        if include_vars:
            keep = [c for c in self.data_normalized.columns if c in set(include_vars)]
            if len(keep) >= 2:
                self.data = self.data[keep].copy()
                self.data_normalized = self.data_normalized[keep].copy()
            else:
                logging.warning("[RunSelected] include_vars отфильтровал <2 колонок — игнорирую.")
        # Фильтр пар (для отчёта/карусели). Если None — все пары.
        self.pair_filter = pair_filter
        self.partial_mode = partial_mode
        self.pairwise_policy = pairwise_policy
        self.partial_controls = partial_controls
        self.custom_controls = custom_controls
        allowed = set(method_mapping.keys())

        variants = [v for v in variants if v in allowed]
        if not variants:
            logging.warning("[RunSelected] Пустой список методов после фильтрации.")
            self.results = {}
            self.variant_lags = {}
            return {}

        self.results = {}
        self.variant_lags = {}
        self.lag_sweep = {}
        self.window_sweep = {}
        self.best_window = {}

        max_lag = max(1, int(max_lag))
        lag_step = max(1, int(lag_step))

        n = len(self.data_normalized)
        # окно: по умолчанию включено для всех методов (как просили)
        if window_min is None:
            window_min = max(20, min(200, n))
        if window_max is None:
            window_max = max(window_min, min(2000, n))
        window_min = max(10, int(window_min))
        window_max = max(window_min, int(window_max))
        window_step = max(1, int(window_step))

        # если пользователь использует старый API (sliding_windows/window_sizes) — маппим
        if sliding_windows and window_sizes:
            # приоритет у явного window_sizes
            _wlist = [int(w) for w in window_sizes if int(w) >= 10]
        else:
            _wlist = list(range(window_min, window_max + 1, window_step))

        # stride: шаг сдвига окна по времени
        if window_stride is None:
            # безопасный дефолт: не слишком густо
            window_stride = max(1, int(min(200, max(5, window_min // 4))))
        window_stride = max(1, int(window_stride))

        for variant in variants:
            spec = get_method_spec(variant)
            supports_lag = bool(getattr(spec, "supports_lag", False))
            chosen_lag = 1

            if supports_lag and (compute_lag_sweep or pick_best_lag):
                sweep = {}
                best_lag = 1
                best_score = float("-inf")

                for lag in range(1, max_lag + 1, lag_step):
                    try:
                        mat = compute_connectivity_variant(self.data_normalized, variant, lag=lag, control=partial_controls, partial_mode=partial_mode, pairwise_policy=pairwise_policy, custom_controls=custom_controls)
                        score = _lag_quality(variant, mat)
                    except Exception as ex:
                        logging.error(f"[LagSweep] {variant} lag={lag}: {ex}")
                        mat = None
                        score = float("nan")
                    sweep[lag] = {"score": float(score), "matrix": mat}
                    if np.isfinite(score) and score > best_score:
                        best_score = float(score)
                        best_lag = lag

                self.lag_sweep[variant] = sweep
                chosen_lag = best_lag if pick_best_lag else 1

            self.variant_lags[variant] = chosen_lag
            self.results[variant] = compute_connectivity_variant(self.data_normalized, variant, lag=chosen_lag, control=partial_controls, partial_mode=partial_mode, pairwise_policy=pairwise_policy, custom_controls=custom_controls)

            # Window sweep (всегда)
            ws_res = {}
            best_global = None  # {"window_size":..., "best":...}
            for w in _wlist:
                w_eff = min(int(w), len(self.data_normalized))
                if w_eff < 10:
                    continue
                try:
                    res = analyze_sliding_windows_with_metric(
                        self.data_normalized,
                        variant,
                        w_eff,
                        int(window_stride),
                        lag=int(chosen_lag),
                    )
                except Exception as ex:
                    logging.error(f"[WinSweep] {variant} w={w_eff}: {ex}")
                    continue
                ws_res[w_eff] = res
                cand = res.get("best_window") or res.get("best") or None
                if cand is not None:
                    if best_global is None or (cand.get("metric", -np.inf) > best_global["best"].get("metric", -np.inf)):
                        best_global = {"window_size": w_eff, "best": cand}

            if ws_res:
                self.window_sweep[variant] = ws_res
            if best_global is not None:
                self.best_window[variant] = best_global

        return dict(self.variant_lags)

    def export_html_report(
        self,
        output_path: str,
        *,
        variants: Optional[List[str]] = None,
        graph_threshold: float = 0.2,
        p_alpha: float = 0.05,
        embed_images: bool = True,
        include_matrix_tables: bool = True,
        include_diagnostics: bool = True,
        diagnostics_max_series: int = 8,
        diagnostics_max_pairs: int = 6,
        include_adf: bool = True,
        include_hurst: bool = True,
        include_seasonality: bool = True,
        include_fft: bool = True,
        include_ac_ph: bool = True,
        include_entropy: bool = True,
        include_frequency_summary: bool = True,
        include_frequency_dependence: bool = True,
    ) -> str:
        """HTML-отчёт с «каруселью» для каждого метода (lag/windows/curves)."""
        if variants is None:
            variants = list(self.results.keys()) if self.results else []

        df = self.data_normalized if not self.data_normalized.empty else self.data
        cols = list(df.columns)
        pair_filter = getattr(self, "pair_filter", None)
        show_cols = cols
        if pair_filter:
            involved = sorted(set([x for p in pair_filter for x in p]), key=lambda s: int(str(s)[1:]) if str(s).startswith("c") and str(s)[1:].isdigit() else str(s))
            if len(involved) >= 2:
                show_cols = [c for c in involved if c in cols]

        def _b64_png(buf: BytesIO) -> str:
            return base64.b64encode(buf.getvalue()).decode("ascii")

        def _plot_matrix(mat: np.ndarray, title: str) -> BytesIO:
            buf = BytesIO()
            fig, ax = plt.subplots(figsize=(5, 4))
            if mat is None or not isinstance(mat, np.ndarray) or mat.size == 0:
                ax.text(0.5, 0.5, "No data", ha="center", va="center")
                ax.axis("off")
            else:
                im = ax.imshow(mat, aspect="auto")
                ax.set_title(title)
                fig.colorbar(im, ax=ax, fraction=0.046, pad=0.04)
                ax.set_xticks(range(len(show_cols)))
                ax.set_yticks(range(len(show_cols)))
                ax.set_xticklabels(cols, rotation=90, fontsize=7)
                ax.set_yticklabels(cols, fontsize=7)
            fig.tight_layout()
            fig.savefig(buf, format="png", dpi=150)
            plt.close(fig)
            buf.seek(0)
            return buf

        def _plot_curve(xs, ys, title: str, xlab: str) -> BytesIO:
            buf = BytesIO()
            fig, ax = plt.subplots(figsize=(5, 3))
            ax.plot(xs, ys, marker="o")
            ax.set_title(title)
            ax.set_xlabel(xlab)
            ax.set_ylabel("quality")
            fig.tight_layout()
            fig.savefig(buf, format="png", dpi=150)
            plt.close(fig)
            buf.seek(0)
            return buf

        def _matrix_table(mat: np.ndarray) -> str:
            if mat is None or not isinstance(mat, np.ndarray) or mat.size == 0:
                return "<div class='muted'>No data</div>"
            rows = []
            rows.append("<table class='matrix'><thead><tr><th></th>" + "".join(f"<th>{_html.escape(c)}</th>" for c in cols) + "</tr></thead><tbody>")
            for i, rname in enumerate(cols):
                rows.append("<tr><th>" + _html.escape(rname) + "</th>" + "".join(f"<td>{mat[i,j]:.4g}</td>" for j in range(len(cols))) + "</tr>")
            rows.append("</tbody></table>")
            return "".join(rows)

        def _carousel(items: List[Tuple[str, str]], cid: str) -> str:
            # items: list of (label, img_b64)
            if not items:
                return "<div class='muted'>No diagnostics</div>"
            tabs = "".join(f"<button class='tab' onclick=\"showSlide('{cid}', {i})\">{_html.escape(lbl)}</button>" for i,(lbl,_) in enumerate(items))
            slides = "".join(
                f"<div class='slide' id='{cid}_s{i}' style='display:{'block' if i==0 else 'none'}'>"
                f"<div class='slideLabel'>{_html.escape(lbl)}</div>"
                f"<img src='data:image/png;base64,{b64}' />"
                f"</div>"
                for i,(lbl,b64) in enumerate(items)
            )
            return f"<div class='carousel'><div class='tabs'>{tabs}</div>{slides}</div>"

        sections = []
        toc = []


        # ----------------------------
        # Initial diagnostics block
        # ----------------------------
        if include_diagnostics:
            diag_cols = show_cols[: max(1, int(diagnostics_max_series))] if diagnostics_max_series else show_cols

            def _safe_id(s: str) -> str:
                import re as _re
                return _re.sub(r'[^a-zA-Z0-9_\-]+', '_', str(s))

            def _plot_series_line(s: pd.Series, title: str) -> str:
                buf = BytesIO()
                fig, ax = plt.subplots(figsize=(6, 2.6))
                ax.plot(np.arange(len(s)), s.values, linewidth=1.0)
                ax.set_title(title)
                ax.set_xlabel("t")
                ax.set_ylabel("value")
                fig.tight_layout()
                fig.savefig(buf, format="png", dpi=150)
                plt.close(fig)
                buf.seek(0)
                return _b64_png(buf)

            
            def _plot_multiseries_line(df_sub: pd.DataFrame, title: str) -> str:
                buf = BytesIO()
                fig, ax = plt.subplots(figsize=(6, 2.8))
                x = np.arange(len(df_sub))
                for c in df_sub.columns:
                    ax.plot(x, df_sub[c].values, linewidth=1.0, label=str(c), alpha=0.9)
                ax.set_title(title)
                ax.set_xlabel("t")
                ax.set_ylabel("value")
                if df_sub.shape[1] <= 12:
                    ax.legend(fontsize=7, ncol=2, loc="upper right", frameon=False)
                fig.tight_layout()
                fig.savefig(buf, format="png", dpi=150)
                plt.close(fig)
                buf.seek(0)
                return _b64_png(buf)

            def _plot_hist(s: pd.Series, title: str) -> str:
                buf = BytesIO()
                fig, ax = plt.subplots(figsize=(6, 2.6))
                ax.hist(s.values, bins=30)
                ax.set_title(title)
                ax.set_xlabel("value")
                ax.set_ylabel("count")
                fig.tight_layout()
                fig.savefig(buf, format="png", dpi=150)
                plt.close(fig)
                buf.seek(0)
                return _b64_png(buf)


            # Overview: all selected series on one plot
            try:
                overview_img = _plot_multiseries_line(df[diag_cols], "Все ряды (совмещённый график)")
                cards = [
                    "<div class='card'>"
                    "<h2 id='diag_overview'>Данные: все ряды</h2>"
                    f"<img class='img' alt='overview' src='data:image/png;base64,{overview_img}'/>"
                    "</div>"
                ]
            except Exception:
                cards = []

            # Per-series cards
            for col in diag_cols:
                if col not in df.columns:
                    continue
                s = pd.to_numeric(df[col], errors="coerce").dropna()
                if len(s) < 3:
                    continue

                # Basic stats
                stats_html = (
                    f"<div class='muted'>n={len(s)} • mean={float(np.mean(s)):.4g} • std={float(np.std(s)):.4g} "
                    f"• min={float(np.min(s)):.4g} • max={float(np.max(s)):.4g}</div>"
                )

                # ADF
                adf_html = ""
                if include_adf:
                    try:
                        adf_stat, adf_pv = self.test_stationarity(col, use_normalized=(df is getattr(self, 'data_normalized', df)))
                        if adf_stat is not None and adf_pv is not None:
                            adf_html = f"<div><b>ADF</b>: stat={adf_stat:.4g}, p={adf_pv:.4g} ({'stationary' if adf_pv < p_alpha else 'non-stationary'})</div>"
                        else:
                            adf_html = "<div><b>ADF</b>: N/A</div>"
                    except Exception as e:
                        adf_html = "<div><b>ADF</b>: error <span class='muted'>(" + _html.escape(str(e)) + ")</span></div>"

                # Hurst
                hurst_html = ""
                if include_hurst:
                    try:
                        H_rs = self.compute_hurst_rs(s)
                        H_dfa = self.compute_hurst_dfa(s)
                        H_av  = self.compute_hurst_aggregated_variance(s)
                        H_wav = self.compute_hurst_wavelet(s)
                        hurst_html = (
                            "<div><b>Hurst</b>: "
                            f"RS={H_rs:.4g} • DFA={H_dfa:.4g} • AggVar={H_av:.4g} • Wavelet={H_wav:.4g}</div>"
                        )
                    except Exception as e:
                        hurst_html = "<div><b>Hurst</b>: error <span class='muted'>(" + _html.escape(str(e)) + ")</span></div>"

                # Seasonality
                seas_html = ""
                if include_seasonality:
                    try:
                        peak_f, period = self.detect_seasonality(s)
                        if peak_f is not None and period is not None:
                            seas_html = f"<div><b>Seasonality</b>: peak_f={float(peak_f):.4g}, period≈{float(period):.4g}</div>"
                        else:
                            seas_html = "<div><b>Seasonality</b>: not detected</div>"
                    except Exception as e:
                        seas_html = "<div><b>Seasonality</b>: error <span class='muted'>(" + _html.escape(str(e)) + ")</span></div>"

                # Entropy
                ent_html = ""
                if include_entropy:
                    try:
                        se = self.compute_sample_entropy(s.values)
                        ent_html = f"<div><b>Entropy</b>: SampleEn={float(se):.4g}</div>"
                    except Exception as e:
                        ent_html = "<div><b>Entropy</b>: error <span class='muted'>(" + _html.escape(str(e)) + ")</span></div>"

                # Frequency summary (top peaks)
                freq_html = ""
                if include_frequency_summary:
                    try:
                        fs = float(getattr(self, "fs", 1.0))
                        freq, amps, periods = frequency_analysis(s, fs=fs)
                        if freq is not None and len(freq) > 0:
                            topk = min(3, len(freq))
                            freq_html = "<div><b>Frequency summary</b>: " + ", ".join(
                                f"{float(freq[i]):.4g}Hz (A={float(amps[i]):.3g}, T={float(periods[i]):.3g})" for i in range(topk)
                            ) + "</div>"
                        else:
                            freq_html = "<div><b>Frequency summary</b>: no peaks</div>"
                    except Exception as e:
                        freq_html = "<div><b>Frequency summary</b>: error <span class='muted'>(" + _html.escape(str(e)) + ")</span></div>"

                # Plots
                img_series = _plot_series_line(s, f"Series: {col}")
                img_hist = _plot_hist(s, f"Histogram: {col}")

                plot_items = [
                    ("series", img_series),
                    ("hist", img_hist),
                ]

                if include_fft:
                    try:
                        ps = self.plot_power_spectrum(s, f"FFT / Power spectrum: {col}")
                        plot_items.append(("fft", _b64_png(ps)))
                    except Exception:
                        pass

                if include_ac_ph:
                    try:
                        ac = self.plot_autocorrelation(s, f"Autocorrelation: {col}")
                        plot_items.append(("ac", _b64_png(ac)))
                    except Exception:
                        pass

                    try:
                        # Use phase/amplitude helper if available
                        plots = plot_individual_ac_ph(df[[col]], f"AC & PH: {col}")
                        if col in plots and "PH" in plots[col]:
                            with open(plots[col]["PH"], "rb") as f:
                                plot_items.append(("ph", base64.b64encode(f.read()).decode("ascii")))
                    except Exception:
                        pass

                carousel_html = _carousel(plot_items, cid=f"diag_{_safe_id(col)}")

                cards.append(
                    "<div class='card'>"
                    f"<h2 id='diag_{_safe_id(col)}'>{_html.escape(col)}</h2>"
                    f"{stats_html}"
                    f"{adf_html}{hurst_html}{seas_html}{ent_html}{freq_html}"
                    f"{carousel_html}"
                    "</div>"
                )

            # Pairwise frequency dependence (coherence)
            pair_cards = ""
            if include_frequency_dependence and len(diag_cols) >= 2:
                pairs = list(combinations(diag_cols, 2))[: max(0, int(diagnostics_max_pairs))]
                rows = []
                plot_items = []
                for (c1, c2) in pairs:
                    s1 = df[c1].dropna()
                    s2 = df[c2].dropna()
                    n = min(len(s1), len(s2))
                    if n < 4:
                        continue
                    s1 = s1.values[:n]
                    s2 = s2.values[:n]
                    try:
                        fs = float(getattr(self, "fs", 1.0))
                        nperseg = int(max(8, min(64, n // 2)))
                        freqs, cxy = coherence(s1, s2, fs=fs, nperseg=nperseg, detrend="constant")
                        if cxy.size == 0:
                            continue
                        max_idx = int(np.argmax(cxy))
                        rows.append(f"<tr><td>{_html.escape(c1)}–{_html.escape(c2)}</td><td>{float(cxy[max_idx]):.3g}</td><td>{float(freqs[max_idx]):.4g}</td></tr>")
                        buf = plot_coherence_vs_frequency(pd.Series(s1), pd.Series(s2), f"Coherence {c1}–{c2}", fs=fs, nperseg=nperseg)
                        plot_items.append((f"{c1}–{c2}", _b64_png(buf)))
                    except Exception:
                        continue

                if rows:
                    table = "<table class='matrix'><thead><tr><th>pair</th><th>max coh</th><th>freq</th></tr></thead><tbody>" + "".join(rows) + "</tbody></table>"
                else:
                    table = "<div class='muted'>No coherence results</div>"
                pair_cards = "<div class='card'><h2 id='freqdep'>Frequency dependence</h2>" + table + _carousel(plot_items, "freqdep_car") + "</div>"

            diag_section = (
                "<section class='card' id='diagnostics'>"
                "<h1>Initial analysis</h1>"
                "<div class='muted'>visual • stationarity • hurst • seasonality • fft • ac/ph • entropy</div>"
                + pair_cards
                + "".join(cards)
                + "</section>"
            )
            toc.insert(0, "<li><a href='#diagnostics'>Initial analysis</a></li>")
            sections.insert(0, diag_section)

        for k, variant in enumerate(variants, start=1):
            toc.append(f"<li><a href='#m_{k}'>{_html.escape(variant)}</a></li>")
            chosen_lag = int(getattr(self, "variant_lags", {}).get(variant, 1))
            base_mat = self.results.get(variant)

            # summary (вшиваем в карточку, чтобы всегда было видно)
            summary_lines = []
            summary_lines.append(f"<div><b>chosen lag:</b> {chosen_lag}</div>")
            sweep_for_summary = getattr(self, "lag_sweep", {}).get(variant) or {}
            if sweep_for_summary:
                # лучший лаг по sweep
                best_lag = max(
                    sweep_for_summary.keys(),
                    key=lambda l: (float(sweep_for_summary[l].get("score", float("-inf"))) if np.isfinite(sweep_for_summary[l].get("score", float("nan"))) else float("-inf")),
                )
                best_score = sweep_for_summary[best_lag].get("score")
                if best_score is not None and np.isfinite(best_score):
                    summary_lines.append(f"<div><b>best lag (sweep):</b> {int(best_lag)} (score {float(best_score):.4g})</div>")

            bw = getattr(self, "best_window", {}).get(variant)
            if bw and bw.get("best"):
                wsize = int(bw.get("window_size", 0) or 0)
                b = bw["best"]
                metric = b.get("metric")
                if metric is not None and np.isfinite(metric):
                    summary_lines.append(f"<div><b>best window size:</b> {wsize} (metric {float(metric):.4g})</div>")
                if b.get("start") is not None and b.get("end") is not None:
                    summary_lines.append(f"<div><b>best window interval:</b> [{int(b['start'])}, {int(b['end'])})</div>")

            summary_html = "<div class='summary'>" + "".join(summary_lines) + "</div>"

            car_items = []

            # lag carousel
            sweep = getattr(self, "lag_sweep", {}).get(variant)
            if sweep:
                lags_sorted = sorted(sweep.keys())
                lag_min = lags_sorted[0]
                lag_max = lags_sorted[-1]
                mat_min = sweep[lag_min].get("matrix")
                mat_max = sweep[lag_max].get("matrix")
                # NOTE: нельзя писать `a or b`, если `a` может быть numpy-массивом:
                # bool(np.array(...)) -> ValueError("truth value is ambiguous")
                _m = sweep.get(chosen_lag, {}).get("matrix", None)
                mat_best = _m if _m is not None else base_mat
                car_items.append((f"Lag min ({lag_min})", _b64_png(_plot_matrix(mat_min, f"{variant} lag={lag_min}"))))
                car_items.append((f"Lag max ({lag_max})", _b64_png(_plot_matrix(mat_max, f"{variant} lag={lag_max}"))))
                car_items.append((f"Lag chosen ({chosen_lag})", _b64_png(_plot_matrix(mat_best, f"{variant} lag={chosen_lag}"))))
                ys = [float(sweep[l]["score"]) for l in lags_sorted]
                car_items.append(("Lag quality", _b64_png(_plot_curve(lags_sorted, ys, f"{variant}: quality vs lag", "lag"))))

            # window carousel
            ws = getattr(self, "window_sweep", {}).get(variant)
            bw = getattr(self, "best_window", {}).get(variant)
            if ws and bw:
                w_sorted = sorted(ws.keys())
                w_min = w_sorted[0]
                w_max = w_sorted[-1]
                best_w = int(bw["window_size"])
                # ожидаемо: res содержит best_window->matrix
                def _best_mat(res):
                    cand = res.get("best_window") or res.get("best") or {}
                    return cand.get("matrix")
                mat_wmin = _best_mat(ws[w_min])
                mat_wmax = _best_mat(ws[w_max])
                mat_wbest = _best_mat(ws[best_w])
                car_items.append((f"Win min ({w_min})", _b64_png(_plot_matrix(mat_wmin, f"{variant} window={w_min}"))))
                car_items.append((f"Win max ({w_max})", _b64_png(_plot_matrix(mat_wmax, f"{variant} window={w_max}"))))
                car_items.append((f"Win best ({best_w})", _b64_png(_plot_matrix(mat_wbest, f"{variant} window={best_w}"))))

                # зависимость качества от window size
                try:
                    xs_w = []
                    ys_w = []
                    for w in w_sorted:
                        cand = (ws[w].get("best_window") or ws[w].get("best") or {})
                        m = cand.get("metric")
                        if m is None or not np.isfinite(m):
                            continue
                        xs_w.append(int(w))
                        ys_w.append(float(m))
                    if len(xs_w) >= 2:
                        car_items.append(("Win size quality", _b64_png(_plot_curve(xs_w, ys_w, f"{variant}: quality vs window size", "window size"))))
                except Exception:
                    pass

                # кривая по окнам для best_w: metric по старту окна
                curve = ws[best_w].get("curve") or ws[best_w].get("window_curve") or None
                if curve and isinstance(curve, dict) and curve.get("x") and curve.get("y"):
                    car_items.append(("Window curve", _b64_png(_plot_curve(curve["x"], curve["y"], f"{variant}: window metric", "window start"))))

            carousel_html = _carousel(car_items, f"c_{k}")

            extra = ""
            if include_matrix_tables:
                extra = "<h4>Matrix (chosen lag)</h4>" + _matrix_table(base_mat)

            sections.append(
                f"<section class='card' id='m_{k}'>"
                f"<h2>{_html.escape(variant)}</h2>"
                f"{summary_html}"
                f"{carousel_html}"
                f"{extra}"
                f"</section>"
            )

        html = f"""<!doctype html>
<html>
<head>
<meta charset='utf-8'/>
<title>TimeSeries Report</title>
<style>
body{{font-family:Arial, sans-serif; margin:0; background:#fafafa;}}
header{{padding:16px 20px; background:#111; color:#fff;}}
main{{display:flex; gap:16px; padding:16px 20px;}}
nav{{width:260px; position:sticky; top:16px; align-self:flex-start; background:#fff; border:1px solid #ddd; border-radius:10px; padding:12px;}}
.card{{background:#fff; border:1px solid #ddd; border-radius:12px; padding:14px; margin-bottom:14px;}}
.muted{{color:#666; font-size:13px;}}
.summary{{color:#222; font-size:13px; line-height:1.35; margin-top:6px; margin-bottom:6px; padding:10px; border:1px solid #eee; border-radius:12px; background:#fcfcfc;}}
.carousel{{border:1px solid #eee; border-radius:12px; padding:10px; margin-top:10px;}}
.tabs{{display:flex; flex-wrap:wrap; gap:6px; margin-bottom:10px;}}
.tab{{border:1px solid #ccc; background:#f6f6f6; border-radius:999px; padding:6px 10px; cursor:pointer; font-size:12px;}}
.slide img{{max-width:100%; border-radius:10px; border:1px solid #eee;}}
.slideLabel{{font-weight:700; margin-bottom:6px;}}
table.matrix{{border-collapse:collapse; font-size:11px; width:100%; overflow:auto; display:block;}}
table.matrix th, table.matrix td{{border:1px solid #eee; padding:4px 6px; text-align:right;}}
table.matrix th{{position:sticky; top:0; background:#fff; text-align:center;}}
</style>
<script>
function showSlide(cid, idx){{
  const slides = document.querySelectorAll('[id^="'+cid+'_s"]');
  slides.forEach((el,i)=>{{ el.style.display = (i===idx?'block':'none'); }});
}}
</script>
</head>
<body>
<header>
  <div style='font-size:18px;font-weight:700;'>Time Series Connectivity Report</div>
  <div class='muted' style='color:#ddd;'>methods: {len(variants)} • vars: {len(cols)} • length: {len(df)}</div>
</header>
<main>
  <nav>
    <div style='font-weight:700;margin-bottom:8px;'>Оглавление</div>
    <ul>{''.join(toc)}</ul>
  </nav>
  <div style='flex:1; min-width:0;'>
    {''.join(sections)}
  </div>
</main>
</body>
</html>"""

        Path(output_path).parent.mkdir(parents=True, exist_ok=True)
        Path(output_path).write_text(html, encoding="utf-8")
        return str(output_path)





def run_selected_methods(
    self,
    variants: List[str],
    *,
    max_lag: int = DEFAULT_MAX_LAG,
    pick_best_lag: bool = False,
) -> Dict[str, int]:
    """
    Запускает только выбранные методы и возвращает словарь variant -> выбранный lag.

    - Для методов без лага используется lag=1.
    - Если pick_best_lag=True, то для directed-методов перебирается lag=1..max_lag и выбирается лучший
      по простой эвристике (p-value минимизируется, остальные максимизируются по среднему |вне диагонали|).
    """
    self.normalize_data()
    if self.data_normalized.empty:
        logging.warning("[RunSelected] Нет данных для выполнения анализа.")
        self.results = {}
        return {}

    # фильтруем методы (учитывая наличие pyinform)
    allowed = set(method_mapping.keys())

    variants = [v for v in variants if v in allowed]
    if not variants:
        logging.warning("[RunSelected] Пустой список методов после фильтрации.")
        self.results = {}
        return {}

    def _score_matrix(mat: np.ndarray, variant: str) -> float:
        if mat is None or not isinstance(mat, np.ndarray) or mat.size == 0:
            return float("-inf")
        n = mat.shape[0]
        mask = ~np.eye(n, dtype=bool)
        vals = mat[mask]
        vals = vals[np.isfinite(vals)]
        if vals.size == 0:
            return float("-inf")
        if get_method_spec(variant).is_p_value:
            # меньше p-value -> лучше
            return -float(np.nanmean(vals))
        return float(np.nanmean(np.abs(vals)))

    self.results = {}
    selected_lags: Dict[str, int] = {}

    for variant in variants:
        spec = get_method_spec(variant)
        if pick_best_lag and spec.supports_lag:
            best_lag = 1
            best_score = float("-inf")
            for lag in range(1, max(1, int(max_lag)) + 1):
                mat = compute_connectivity_variant(self.data_normalized, variant, lag=lag, control=partial_controls, partial_mode=partial_mode, pairwise_policy=pairwise_policy, custom_controls=custom_controls)
                s = _score_matrix(mat, variant)
                if s > best_score:
                    best_score = s
                    best_lag = lag
            selected_lags[variant] = best_lag
            self.results[variant] = compute_connectivity_variant(self.data_normalized, variant, lag=best_lag)
        else:
            selected_lags[variant] = 1
            self.results[variant] = compute_connectivity_variant(self.data_normalized, variant, lag=1)

    # запоминаем для отчёта (необязательно, но удобно)
    self.variant_lags = selected_lags
    return selected_lags

    def analyze_sliding_windows(self, variant: str, window_size: int = 100, overlap: int = 50, threshold: float = 0.2) -> dict:
        if self.data_normalized.empty:
            return {}
        if not self.enable_experimental:
            return {}
        actual_window_size = min(window_size, len(self.data_normalized))
        actual_overlap = min(overlap, actual_window_size // 2) 
        return analyze_sliding_windows_with_metric(self.data_normalized, variant, actual_window_size, actual_overlap)

    def sliding_window_pairwise_analysis(self, method: str, window_size: int = 50, overlap: int = 25) -> dict:
        if self.data_normalized.empty or not self.enable_experimental:
            return {}
        actual_window_size = min(window_size, len(self.data_normalized))
        actual_overlap = min(overlap, actual_window_size // 2)
        return sliding_window_pairwise_analysis(self.data_normalized, method, actual_window_size, actual_overlap)

    def plot_lag_metrics(self, variant: str, lag_results: dict, legend_text: str = "") -> BytesIO:
        valid = [(l, lag_results[l][0]) for l in sorted(lag_results.keys()) if not np.isnan(lag_results[l][0])]
        if not valid:
            return BytesIO()
        lags, metrics = zip(*valid)
        fig, ax = plt.subplots(figsize=(6, 4))
        ax.plot(lags, metrics, marker='o', linestyle='-', color='b', label=legend_text)
        ax.set_title(f"Lag Metrics for {variant.upper()} (используемые лаги: {min(lags)}-{max(lags)})")
        ax.set_xlabel("Lag")
        ax.set_ylabel("Metric")
        ax.grid(True)
        ax.legend()
        buf = BytesIO()
        plt.tight_layout()
        plt.savefig(buf, format="png", dpi=100)
        buf.seek(0)
        plt.close(fig)
        return buf

    def plot_sliding_window_comparison(self, sw_results: dict, legend_text: str = "") -> BytesIO:
        if not sw_results: return BytesIO()
        positions = sorted(sw_results.keys())
        metrics = [sw_results[pos][2] for pos in positions]
        fig, ax = plt.subplots(figsize=(10, 4))
        ax.plot(positions, metrics, marker='o', linestyle='-', color='m')
        ax.set_title("Sliding Window Metric vs. Window Start (окно = 50 точек)")
        ax.set_xlabel("Window Start Index")
        ax.set_ylabel("Metric")
        ax.grid(True)
        if legend_text:
            ax.text(0.05, 0.95, legend_text, transform=ax.transAxes, fontsize=8,
                    verticalalignment='top', bbox=dict(facecolor='white', alpha=0.5))
        buf = BytesIO()
        plt.tight_layout()
        plt.savefig(buf, format="png", dpi=100)
        buf.seek(0)
        plt.close(fig)
        return buf

    def plot_all_methods_lag_comparison(self, lag_results_dict: dict) -> BytesIO:
        fig, ax = plt.subplots(figsize=(10, 6))
        for variant, res in lag_results_dict.items():
            valid = [(l, res[l][0]) for l in sorted(res.keys()) if not np.isnan(res[l][0])]
            if valid:
                lags, metrics = zip(*valid)
                ax.plot(lags, metrics, marker='o', label=variant.upper())
        ax.set_title("Lag Analysis Comparison")
        ax.set_xlabel("Lag")
        ax.set_ylabel("Metric")
        ax.legend()
        ax.grid(True)
        buf = BytesIO()
        plt.tight_layout()
        plt.savefig(buf, format="png", dpi=100)
        buf.seek(0)
        plt.close(fig)
        return buf

    def assess_significance(self,
                        method: str,
                        n_permutations: int = 100,
                        alpha: float = 0.05):
 
        orig = compute_connectivity_variant(self.data_normalized.copy(), method, lag=1)
        if orig is None or orig.shape[0] < 2:
            return None, None

        n = orig.shape[0]
        idx_i, idx_j = np.triu_indices(n, k=1)
        n_pairs = len(idx_i)
        
        perm_vals = np.zeros((n_permutations, n_pairs), dtype=float)

        for k in range(n_permutations):
            permuted = self.data_normalized.apply(np.random.permutation, axis=0)
            pm = compute_connectivity_variant(permuted.copy(), method, lag=1)
            if pm is None or pm.shape != orig.shape:
                perm_vals[k, :] = np.nan
            else:
                perm_vals[k, :] = np.abs(pm[idx_i, idx_j])

        orig_vals = np.abs(orig[idx_i, idx_j])
        
        # Здесь могут быть NaN.
        if np.any(np.isnan(orig_vals)) or n_permutations == 0:
            p_matrix = np.full_like(orig, np.nan)
            sig = np.full_like(orig, False, dtype=bool)
            return sig, p_matrix

        p_vals = (
            (perm_vals >= orig_vals[None, :]).sum(axis=0) + 1
        ) / (n_permutations + 1)

        sig = np.zeros_like(orig, dtype=bool)
        for (i,j), p in zip(zip(idx_i, idx_j), p_vals):
            sig[i, j] = sig[j, i] = (p < alpha)

        np.fill_diagonal(sig, False)

        p_matrix = np.zeros_like(orig, dtype=float)
        for (i,j), p in zip(zip(idx_i, idx_j), p_vals):
            p_matrix[i, j] = p_matrix[j, i] = p

        return sig, p_matrix


    def compute_hurst_rs(self, series) -> float:
        try:
            H, _, _ = compute_Hc(series.dropna().values, kind='change', simplified=True)
            return H
        except Exception as ex:
            logging.error(f"[Hurst RS] Ошибка: {ex}")
            return np.nan

    def compute_hurst_dfa(self, series) -> float:
        try:
            if not _nolds_or_warn():
                return np.nan
            H = nolds.dfa(series.dropna().values)
            return float(H)
        except Exception as ex:
            logging.error(f"[Hurst DFA] Ошибка: {ex}")
            return np.nan

    def compute_hurst_aggregated_variance(self, series, max_n=100) -> float:
        try:
            arr = np.array(series.dropna())
            N = len(arr)
            if N < max_n:
                return np.nan
            m_vals = np.arange(1, min(max_n+1, N//2))
            variances = []
            for m in m_vals:
                nb = N // m
                if nb > 0:
                    reshaped = arr[:nb*m].reshape(nb, m)
                    block_means = reshaped.mean(axis=1)
                    if len(block_means) > 1:
                        variances.append(np.var(block_means))
            if not variances: return np.nan
            log_m = np.log10(m_vals[:len(variances)])
            log_var = np.log10(variances)
            slope, _ = np.polyfit(log_m, log_var, 1)
            H = 1 - slope/2
            return H
        except Exception as ex:
            logging.error(f"[Hurst AggVar] Ошибка: {ex}")
            return np.nan

    def compute_hurst_wavelet(self, series) -> float:
        try:
            arr = np.array(series.dropna())
            N = len(arr)
            if N < 50:
                return np.nan
            yf_arr = fft(arr)
            freqs = np.fft.fftfreq(N)
            psd = np.abs(yf_arr)**2
            idx = freqs > 0
            freqs = freqs[idx]
            psd = psd[idx]
            if len(freqs) < 2: return np.nan
            log_freqs = np.log10(freqs)
            log_psd = np.log10(psd)
            slope, _ = np.polyfit(log_freqs, log_psd, 1)
            H = (1 - slope)/2
            return H
        except Exception as ex:
            logging.error(f"[Hurst Wavelet] Ошибка: {ex}")
            return np.nan
    #графики?

def compute_sample_entropy(self, x) -> float:
    """Sample entropy (устойчиво): принимает array/Series, возвращает float или NaN."""
    try:
        arr = np.asarray(x, dtype=np.float64).reshape(-1)
        arr = arr[np.isfinite(arr)]
        if arr.size < 20 or np.std(arr) < 1e-10:
            return np.nan
        return float(nolds.sampen(arr))
    except Exception as ex:
        logging.error(f"[Sample Entropy] Ошибка: {ex}")
        return np.nan

    def plot_autocorrelation(self, series: pd.Series, title: str, suppress_noise: bool = False, noise_threshold: float = 0.9, legend_text: str = "") -> BytesIO:
        fig, ax = plt.subplots(figsize=(6, 4))
        plot_acf(series.dropna(), ax=ax, lags=min(50, len(series)-1), zero=False, alpha=0.05, fft=True)
        
        ax.set_title(title + (" (с подавлением шума)" if suppress_noise else ""))
        if legend_text:
            ax.text(0.05, 0.95, legend_text, transform=ax.transAxes, fontsize=8,
                    verticalalignment='top', bbox=dict(facecolor='white', alpha=0.5))
        buf  = BytesIO()
        plt.tight_layout()
        plt.savefig(buf, format="png", dpi=100)
        buf.seek(0)
        plt.close(fig)
        return buf

    def plot_power_spectrum(self, series: pd.Series, title: str) -> BytesIO:
        N = len(series)
        if N < 2: return BytesIO()
        arr = _as_float64_1d(series)
        if arr.size < 2:
            return BytesIO()
        yf_arr = fft(arr)
        fs = float(getattr(self, "fs", 1.0))
        fs = fs if np.isfinite(fs) and fs > 0 else 1.0
        freqs = np.fft.fftfreq(int(arr.size), d=1.0 / fs)
        amp = np.abs(yf_arr)**2
        idx = freqs > 0
        freqs = freqs[idx]
        amp = amp[idx]
        if len(freqs) < 2: return BytesIO()
        log_freqs = np.log10(freqs)
        log_amp = np.log10(amp)
        slope, _ = np.polyfit(log_freqs, log_amp, 1)
        fig, ax = plt.subplots(figsize=(6, 4))
        ax.loglog(freqs, amp, label=f"Slope={slope:.2f}")
        ax.set_title(title)
        ax.set_xlabel("Частота (log)")
        ax.set_ylabel("Мощность (log)")
        ax.legend()
        buf = BytesIO()
        plt.tight_layout()
        plt.savefig(buf, format="png")
        buf.seek(0)
        plt.close(fig)
        return buf

    def detect_seasonality(self, series: pd.Series, threshold=0.2):
        N = len(series)
        if N < 2: return None, None
        arr = _as_float64_1d(series)
        if arr.size < 2:
            return None, None
        yf_arr = fft(arr)
        fs = float(getattr(self, "fs", 1.0))
        fs = fs if np.isfinite(fs) and fs > 0 else 1.0
        freqs = np.fft.fftfreq(int(arr.size), d=1.0 / fs)
        amp = np.abs(yf_arr)**2
        idx = freqs > 0
        freqs = freqs[idx]
        amp = amp[idx]
        if amp.size == 0: return None, None
        peaks, props = find_peaks(amp, height=threshold*np.max(amp))
        if len(peaks) > 0:
            peak_freqs = freqs[peaks]
            periods = 1 / peak_freqs
            return peak_freqs, periods
        return None, None

    def export_hurst_sheet(self, wb: Workbook, method_name: str, compute_func, plot_func):
        ws = wb.create_sheet(method_name)
        ws.append(["Столбец", "Hurst Exponent"])
        row_num = 2
        for col in self.data.columns:
            s = self.data[col].dropna()
            if len(s) < 50:
                ws.append([col, "N/A (недостаточно данных)"])
                row_num += 1
                continue
            H = compute_func(s)
            if H is not None and not np.isnan(H):
                ws.append([col, f"{H:.3f}"])
                buf = plot_func(s, f"{method_name} - {col}")
                img = Image(buf)
                img.width = 400
                img.height = 300
                ws.add_image(img, f"D{row_num}")
            else:
                ws.append([col, "N/A"])
            row_num += 1

    def export_seasonality_sheet(self, wb: Workbook):
        ws = wb.create_sheet("Seasonality")
        ws.append(["Столбец", "Пиковые частоты", "Периоды", "Пояснение"])
        for col in self.data.columns:
            s = self.data[col].dropna()
            freq, periods = self.detect_seasonality(s)
            if freq is not None and np.asarray(freq).size > 0 and periods is not None and np.asarray(periods).size > 0:
                freq_str = ", ".join([f"{f:.3f}" for f in freq])
                period_str = ", ".join([f"{p:.1f}" for p in periods])
                explanation = f"Сезонность обнаружена, период ≈ {np.median(periods):.1f}"
                ws.append([col, freq_str, period_str, explanation]) 
            else:
                ws.append([col, "Нет пиков", "Нет пиков", "Сезонность не обнаружена"])
                
    def export_undirected_sheet(self, wb: Workbook, threshold: float, p_value_alpha: float):
        ws = wb.create_sheet("Undirected Methods")
        headers = ["Pair", "Control Set"] + self.undirected_methods
        ws.append(headers)
        indices = {c: i for i, c in enumerate(self.data.columns)}
        fill_green = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
        fill_pink = PatternFill(start_color="FFC0CB", end_color="FFC0CB", fill_type="solid")
        for row_idx, row in enumerate(self.undirected_rows, start=2):
            pair, S = row
            pair_str = "-".join(pair)
            S_str = ",".join(S) if S else "None"
            vals = [pair_str, S_str]
            for method in self.undirected_methods:
                mat = self._get_cached_matrix(method, S)
                if mat is not None:
                    v = self.get_undirected_value(mat, pair[0], pair[1], indices)
                    vals.append(fmt_val(v))
                else:
                    vals.append("N/A")
            ws.append(vals)
            for col_idx, method in enumerate(self.undirected_methods, start=3):
                cell = ws.cell(row=row_idx, column=col_idx)
                if cell.value != "N/A":
                    try:
                        v = float(cell.value)
                        if is_pvalue_method(method):
                            cell.fill = fill_green if v < p_value_alpha else fill_pink
                        else:
                            cell.fill = fill_green if abs(v) >= threshold else fill_pink
                    except:
                        pass
        for col in ws.columns:
            max_length = max(len(str(cell.value)) for cell in col if cell.value is not None)
            ws.column_dimensions[get_column_letter(col[0].column)].width = max_length

    def export_directed_sheet(self, wb: Workbook, threshold: float, p_value_alpha: float):
        ws = wb.create_sheet("Directed Methods")
        headers = ["Directed Pair", "Control Set"] + self.directed_methods
        ws.append(headers)
        indices = {c: i for i, c in enumerate(self.data.columns)}
        fill_green = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
        fill_pink = PatternFill(start_color="FFC0CB", end_color="FFC0CB", fill_type="solid")
        for row_idx, row in enumerate(self.directed_rows, start=2):
            pair, S = row
            pair_str = f"{pair[0]}->{pair[1]}"
            S_str = ",".join(S) if S else "None"
            vals = [pair_str, S_str]
            for method in self.directed_methods:
                mat = self._get_cached_matrix(method, S)
                if mat is not None:
                    v = self.get_directed_value(mat, pair[0], pair[1], indices)
                    vals.append(fmt_val(v))
                else:
                    vals.append("N/A")
            ws.append(vals)
            for col_idx, method in enumerate(self.directed_methods, start=3):
                cell = ws.cell(row=row_idx, column=col_idx)
                if cell.value != "N/A":
                    try:
                        v = float(cell.value)
                        if is_pvalue_method(method):
                            cell.fill = fill_green if v < p_value_alpha else fill_pink
                        else:
                            cell.fill = fill_green if abs(v) >= threshold else fill_pink
                    except:
                        pass
        for col in ws.columns:
            max_length = max(len(str(cell.value)) for cell in col if cell.value is not None)
            ws.column_dimensions[get_column_letter(col[0].column)].width = max_length

    def select_lag_metrics(self, lag_res: dict, variant: str):
        valid = [l for l in lag_res if lag_res[l] is not None and lag_res[l][0] is not None and not np.isnan(lag_res[l][0])]
        if not valid:
            return None, None, None
        metrics = [lag_res[l][0] for l in valid]
        # metric уже приведён к "больше => лучше" для всех методов
        best = max(valid, key=lambda l: lag_res[l][0])
        worst = min(valid, key=lambda l: lag_res[l][0])
        median = min(valid, key=lambda l: abs(lag_res[l][0] - np.median(metrics)))
        return best, median, worst


    def export_method_sheet(self, wb: Workbook, variant: str, threshold: float, window_size: int, overlap: int, p_value_alpha: float = 0.05) -> None:
        directed_flag = is_directed_method(variant)
        is_pval = is_pvalue_method(variant)
        edge_threshold = p_value_alpha if is_pval else threshold
        invert_threshold = True if is_pval else False
        ws = wb.create_sheet(variant.upper() + " Results")
        ws.append([f"Метод: {variant.upper()} (Лаг = 1, окно = {window_size})"])
        if spec.is_p_value:
            ws.append([f"Интерпретация: p-value (меньше = сильнее), типичный порог α={p_value_alpha}"])
        else:
            ws.append([f"Интерпретация: сила связи (|value| больше = сильнее), порог визуализации {threshold}"])
        full_mat = compute_connectivity_variant(self.data_normalized, variant, lag=1)
        if full_mat is None or not hasattr(full_mat, "shape") or (full_mat.shape[0] != self.data_normalized.shape[1] or full_mat.shape[1] != self.data_normalized.shape[1]):
             ws.append([f"{variant.upper()}: Метод не работает для этих данных или вернул матрицу некорректного размера."])
             return
        ws.append(["Полная матрица:"])
        legend_text = f"Лаг=1, Окно={window_size}"
        add_method_to_sheet(ws, ws.max_row + 1, f"{variant.upper()} Full Matrix", full_mat, directed=directed_flag, legend_text=legend_text)
        buf_heat_full = plot_heatmap(full_mat, f"{variant.upper()} Heatmap (Full)", legend_text=legend_text)
        img_heat_full = Image(buf_heat_full)
        img_heat_full.width = 400
        img_heat_full.height = 300
        ws.add_image(img_heat_full, "G2")
        buf_conn_full = plot_connectome(full_mat, f"{variant.upper()} Connectome (Full)",
                                        threshold=edge_threshold, directed=directed_flag, invert_threshold=invert_threshold, legend_text=legend_text)
        img_conn_full = Image(buf_conn_full)
        img_conn_full.width = 400
        img_conn_full.height = 400
        ws.add_image(img_conn_full, "M2")
        ws.append(["--- Медианный и Лучший лаг ---"])
        lag_res = self.lag_results.get(variant, {}) # исп предрассчит
        best_lag, median_lag, worst_lag = self.select_lag_metrics(lag_res, variant)
        if median_lag is not None and lag_res.get(median_lag) and lag_res[median_lag][1] is not None:
            med_mat = lag_res[median_lag][1]
            ws.append([f"Медианный лаг: {median_lag}"])
            legend_text_med = f"Лаг={median_lag}, Окно={window_size}"
            buf_med_heat = plot_heatmap(med_mat, f"{variant.upper()} Heatmap (Median Lag)", legend_text=legend_text_med)
            img_med_heat = Image(buf_med_heat)
            img_med_heat.width = 400
            img_med_heat.height = 300
            ws.add_image(img_med_heat, "A20")
            buf_med_conn = plot_connectome(med_mat, f"{variant.upper()} Connectome (Median Lag)", threshold=edge_threshold, directed=directed_flag, invert_threshold=invert_threshold, legend_text=legend_text_med)
            img_med_conn = Image(buf_med_conn)
            img_med_conn.width = 400
            img_med_conn.height = 400
            ws.add_image(img_med_conn, "G20")
        else:
            ws.append([f"Медианный лаг: N/A"])

        if best_lag is not None and lag_res.get(best_lag) and lag_res[best_lag][1] is not None:
            best_mat = lag_res[best_lag][1]
            ws.append([f"Лучший лаг: {best_lag}"])
            legend_text_best = f"Лаг={best_lag}, Окно={window_size}"
            buf_best_heat = plot_heatmap(best_mat, f"{variant.upper()} Heatmap (Best Lag)", legend_text=legend_text_best)
            img_best_heat = Image(buf_best_heat)
            img_best_heat.width = 400
            img_best_heat.height = 300
            ws.add_image(img_best_heat, "M20")
            buf_best_conn = plot_connectome(best_mat, f"{variant.upper()} Connectome (Best Lag)", threshold=edge_threshold, directed=directed_flag, invert_threshold=invert_threshold, legend_text=legend_text_best)
            img_best_conn = Image(buf_best_conn)
            img_best_conn.width = 400
            img_best_conn.height = 400
            ws.add_image(img_best_conn, "Q20")
        else:
            ws.append([f"Лучший лаг: N/A"])
            
        ws.append(["--- Анализ скользящих окон ---"])
        sw_summary = {}
        for w_size in [50, 100, 500]:
                # не брать слишком большое окно
            current_w_size = min(w_size, len(self.data_normalized) // 2) # надо хотя бы два
            if current_w_size < 10: # минимально разумный
                ws.append([f"Размер окна {w_size}: Недостаточно данных для скользящих окон."])
                continue

            sw_res = analyze_sliding_windows_with_metric(self.data_normalized, variant, window_size=current_w_size, overlap=current_w_size//2)
            legend_text_sw = f"Метод={variant}, Окно={current_w_size}"
            buf_sw = self.plot_sliding_window_comparison(sw_res, legend_text=legend_text_sw)
            img_sw = Image(buf_sw)
            img_sw.width = 400
            img_sw.height = 300
            cell = f"A{ws.max_row + 2}"
            ws.add_image(img_sw, cell)
            if sw_res:
                best_start = max(sw_res.keys(), key=lambda s: sw_res[s][2])
                sw_summary[current_w_size] = (best_start, sw_res[best_start])
                ws.append([f"Размер окна {current_w_size}: лучшее окно = {best_start}, метрика = {sw_res[best_start][2]:.3f}"])
            else:
                 ws.append([f"Размер окна {current_w_size}: Нет результатов для скользящих окон."])

        if sw_summary:
            best_overall = max(sw_summary.items(), key=lambda item: item[1][1][2])
            best_w_size = best_overall[0]
            best_w_start, best_w_data = best_overall[1]
            ws.append(["--- Лучшее окно среди всех ---"])
            ws.append([f"Размер окна: {best_w_size}, старт: {best_w_start}"])
            best_win_mat = best_w_data[1]
            ws.append(["Матрица лучшего окна:"])
            add_method_to_sheet(ws, ws.max_row + 1, f"{variant.upper()} Best Window Matrix", best_win_mat, directed=directed_flag, legend_text=f"Лаг=1, Окно={best_w_size}")
            buf_best_heat = plot_heatmap(best_win_mat, f"{variant.upper()} Heatmap (Best Window)", legend_text=f"Лаг=1, Окно={best_w_size}")
            img_best_heat = Image(buf_best_heat)
            img_best_heat.width = 400
            img_best_heat.height = 300
            ws.add_image(img_best_heat, "A40")
            buf_best_conn = plot_connectome(best_win_mat, f"{variant.upper()} Connectome (Best Window)", threshold, directed=directed_flag, legend_text=f"Лаг=1, Окно={best_w_size}")
            img_best_conn = Image(buf_best_conn)
            img_best_conn.width = 400
            img_best_conn.height = 400
            ws.add_image(img_best_conn, "G40")
        ws.append(["--- Конец листа ---"])

    def export_summary_sheet(self, wb: Workbook, graph_threshold: float = 0.2, p_value_alpha: float = 0.05) -> None:
        ws_summary = wb.create_sheet("Summary")
        full_methods = [m for m in self.lag_results.keys()]        
        undirected_pairs = list(combinations(self.data.columns, 2))
        directed_pairs = list(permutations(self.data.columns, 2))
        
        headers = ["Название метода"]
        for p in undirected_pairs:
            p_str = f"{p[0]}-{p[1]}"
            headers.extend([f"{p_str} (мед)", f"{p_str} (луч)"])
        for p in directed_pairs:
            p_str = f"{p[0]}->{p[1]}"
            headers.extend([f"{p_str} (мед)", f"{p_str} (луч)"])
        headers.extend(["Лучший лаг", "Медианный лаг", "H_RS", "H_DFA", "H_AggVar", "H_Wavelet", "Seasonality Periods"])
        ws_summary.append(headers)
        
        fill_green = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
        fill_pink = PatternFill(start_color="FFC0CB", end_color="FFC0CB", fill_type="solid")
        indices = {c: i for i, c in enumerate(self.data.columns)}
        
        for variant in full_methods:
            spec = get_method_spec(variant); is_pval = is_pvalue_method(variant)
            lag_res = self.lag_results.get(variant, {}) # берём превыч
            
            best_lag, median_lag, _ = self.select_lag_metrics(lag_res, variant)
            
            mat_med = lag_res.get(median_lag, [None, None])[1] if median_lag is not None else None
            mat_best = lag_res.get(best_lag, [None, None])[1] if best_lag is not None else None
            
            row = [variant]
            
            for pair in undirected_pairs:
                var1, var2 = pair
                if mat_med is not None and mat_best is not None:
                    i, j = indices.get(var1, -1), indices.get(var2, -1) 
                    if i != -1 and j != -1:
                        if spec.directed:
                            a_med = mat_med[i, j] if mat_med.shape[0] > i and mat_med.shape[1] > j else np.nan
                            b_med = mat_med[j, i] if mat_med.shape[0] > j and mat_med.shape[1] > i else np.nan
                            a_best = mat_best[i, j] if mat_best.shape[0] > i and mat_best.shape[1] > j else np.nan
                            b_best = mat_best[j, i] if mat_best.shape[0] > j and mat_best.shape[1] > i else np.nan
                            if spec.is_p_value:
                                val_med = np.nanmin([a_med, b_med])
                                val_best = np.nanmin([a_best, b_best])
                            else:
                                val_med = np.nanmax([abs(a_med), abs(b_med)])
                                val_best = np.nanmax([abs(a_best), abs(b_best)])
                        else:
                            mi, mj = min(i, j), max(i, j)
                            val_med = mat_med[mi, mj] if mat_med.shape[0] > mj and mat_med.shape[1] > mj else np.nan
                            val_best = mat_best[mi, mj] if mat_best.shape[0] > mj and mat_best.shape[1] > mj else np.nan
                    else:
                        val_med, val_best = np.nan, np.nan
                else:
                    val_med = np.nan
                    val_best = np.nan
                row.extend([fmt_val(val_med), fmt_val(val_best)])

            for pair in directed_pairs:
                src, tgt = pair
                if mat_med is not None and mat_best is not None:
                    i, j = indices.get(src, -1), indices.get(tgt, -1)
                    if i != -1 and j != -1:
                        val_med = mat_med[i, j] if mat_med.shape[0] > i and mat_med.shape[1] > j else np.nan
                        val_best = mat_best[i, j] if mat_best.shape[0] > i and mat_best.shape[1] > j else np.nan
                    else:
                        val_med, val_best = np.nan, np.nan
                else:
                    val_med = np.nan
                    val_best = np.nan
                row.extend([fmt_val(val_med), fmt_val(val_best)])
            
            row.extend([best_lag if best_lag is not None else 'N/A', median_lag if median_lag is not None else 'N/A'])
            
            hurst_vals_for_avg = []
            season_vals_for_avg = []
            for col in self.data.columns:
                s = self.data[col].dropna()
                H_RS = self.compute_hurst_rs(s)
                H_DFA = self.compute_hurst_dfa(s)
                H_AggVar = self.compute_hurst_aggregated_variance(s)
                H_Wavelet = self.compute_hurst_wavelet(s)
                
                hurst_vals_for_avg.extend([H_RS, H_DFA, H_AggVar, H_Wavelet])
                
                freq, periods = self.detect_seasonality(s)
                season_text = f"Есть сезонность, период ≈ {np.median(periods):.1f}" if freq is not None and periods is not None and len(periods) > 0 else "Нет сезонности"
                season_vals_for_avg.append(season_text)

            avg_hurst_values = [h for h in hurst_vals_for_avg if h is not None and not np.isnan(h)]
            avg_hurst_results = [fmt_val(np.mean(avg_hurst_values)) if avg_hurst_values else "N/A"] * 4
            
            avg_season_text = "; ".join(sorted(list(set(season_vals_for_avg)))) 
            
            row.extend(avg_hurst_results)
            row.append(avg_season_text)
            ws_summary.append(row)
            
            cur_row = ws_summary.max_row
            num_pair_cols = (len(undirected_pairs) + len(directed_pairs)) * 2
            for idx in range(2, 2 + num_pair_cols):
                cell = ws_summary.cell(row=cur_row, column=idx)
                if cell.value != "N/A":
                    try:
                        v = float(cell.value)
                        if is_pvalue_method(variant):
                            cell.fill = fill_green if v < p_value_alpha else fill_pink
                        else:
                            cell.fill = fill_green if abs(v) > graph_threshold else fill_pink
                    except:
                        pass
        
        for i, _ in enumerate(headers, start=1):
            max_len = 0
            for r in range(1, ws_summary.max_row + 1):
                cell_value = ws_summary.cell(row=r, column=i).value
                if cell_value is not None:
                    max_len = max(max_len, len(str(cell_value)))
            ws_summary.column_dimensions[get_column_letter(i)].width = max_len + 2

    def export_coefficients_sheet(self, wb: Workbook):
        export_coefficients_sheet(self, wb)

    def export_frequency_summary_sheet(self, wb: Workbook):
        export_frequency_summary_sheet(self, wb)

    def export_frequency_dependence_sheet(self, wb: Workbook):
        ws = wb.create_sheet("Frequency Dependence")
        ws.append(["Пара", "Макс. когерентность", "Частота (Hz)"])
        for pair in combinations(self.data.columns, 2):
            col1, col2 = pair
            s1 = self.data[col1].dropna()
            s2 = self.data[col2].dropna()
            n = min(len(s1), len(s2))
            if n == 0:
                ws.append([f"{col1}-{col2}", "N/A", "N/A"])
                continue
            s1 = s1.values[:n]
            s2 = s2.values[:n]
            fs = float(getattr(self, "fs", 1.0))
            fs = fs if np.isfinite(fs) and fs > 0 else 1.0
            nperseg = int(max(8, min(64, n // 2)))
            freqs, cxy = coherence(s1, s2, fs=fs, nperseg=nperseg, detrend="constant")
            if cxy.size == 0:
                ws.append([f"{col1}-{col2}", "N/A", "N/A"])
                continue
            max_idx = np.argmax(cxy)
            max_coh = cxy[max_idx]
            max_freq = freqs[max_idx]
            ws.append([f"{col1}-{col2}", f"{max_coh:.3f}", f"{max_freq:.3f}"])
            buf = plot_coherence_vs_frequency(self.data[col1], self.data[col2], f"Coherence {col1}-{col2}", fs=fs)
            img = Image(buf)
            img.width = 400
            img.height = 300
            cell = f"D{ws.max_row}"
            ws.add_image(img, cell)

    def export_big_excel(self, save_path: str = "AllMethods_Full.xlsx", threshold: float = 0.2, p_value_alpha: float = DEFAULT_PVALUE_ALPHA, window_size: int = 100, overlap: int = 50,
                           log_transform=False, remove_outliers=True, normalize=True, fill_missing=True, check_stationarity=False, include_pairwise_sheets: bool = True) -> str:
        wb = Workbook()
        wb.remove(wb.active)
        add_raw_data_sheet(wb, self.data) 
        self.export_summary_sheet(wb, graph_threshold=threshold, p_value_alpha=p_value_alpha)
        if include_pairwise_sheets:
            self._ensure_pairwise_cache()
            self.export_undirected_sheet(wb, threshold=threshold, p_value_alpha=p_value_alpha)
            self.export_directed_sheet(wb, threshold=threshold, p_value_alpha=p_value_alpha)
        ws_diag = wb.create_sheet("Data & Diagnostics")
        ws_diag.append(["Диагностика (Original) с информацией о сезонности"])
        for c in self.data.columns:
            adf_stat, adf_pv = self.test_stationarity(c, use_normalized=False)
            outl = self.detect_outliers(c, thresh=3, use_normalized=False)
            freq, periods = self.detect_seasonality(self.data[c].dropna())
            season_text = f"Период ≈ {np.median(periods):.1f}" if freq is not None and periods is not None and len(periods) > 0 else "Нет сезонности"
            ws_diag.append([f"Столбец: {c}"])
            if adf_stat is not None and adf_pv is not None:
                ws_diag.append([f"ADF: {adf_stat:.6f} (p={adf_pv:.6f})"])
            else:
                ws_diag.append(["ADF: N/A"])
            ws_diag.append([f"Выбросов: {len(outl)}, Пример: {list(outl)[:10]}"])
            ws_diag.append([f"Сезонность: {season_text}"])
            buf_ps = self.plot_power_spectrum(self.data[c].dropna(), f"Power Spectrum {c}")
            img_ps = Image(buf_ps)
            img_ps.width = 400
            img_ps.height = 300
            cell = f"D{ws_diag.max_row - 1}"
            ws_diag.add_image(img_ps, cell)
            ws_diag.append([])
        
        # Combined Time Series – агрегированный и индивидуальные графики
        export_combined_ts_sheet(self, wb)
        
        # Combined FFT – агрегированный и индивидуальные графики
        export_all_fft_sheet(self, wb)
        
        # Объединённый информационный лист
        export_combined_informational_sheet(self, wb)
        
        # --- ЛИСТЫ методов отдельно ---
        for variant in method_mapping.keys():
            self.export_method_sheet(wb, variant, threshold=threshold, window_size=window_size, overlap=overlap, p_value_alpha=p_value_alpha)
        
        self.export_hurst_sheet(wb, "Hurst_RS", self.compute_hurst_rs, self.plot_autocorrelation)
        self.export_hurst_sheet(wb, "Hurst_DFA", self.compute_hurst_dfa, self.plot_power_spectrum)
        self.export_hurst_sheet(wb, "Hurst_AggVar", self.compute_hurst_aggregated_variance, self.plot_autocorrelation)
        self.export_hurst_sheet(wb, "Hurst_Wavelet", self.compute_hurst_wavelet, self.plot_power_spectrum)
        self.export_seasonality_sheet(wb)
        self.export_coefficients_sheet(wb)
        self.export_frequency_summary_sheet(wb)
        self.export_frequency_dependence_sheet(wb)
        export_individual_ac_ph_sheet(self, wb)
        export_entropy_sheet(self, wb)
        
        if "Combined Time Series1" in wb.sheetnames:
            ws_to_remove = wb["Combined Time Series1"]
            wb.remove(ws_to_remove)
        
        create_table_of_contents(wb)
        
        wb.save(save_path)
        logging.info(f"[Export] Excel файл сохранён: {save_path}")
        return save_path

    def test_stationarity(self, col: str, use_normalized: bool = False):
        data = self.data_normalized if use_normalized else self.data
        if col not in data.columns or not pd.api.types.is_numeric_dtype(data[col]):
            return None, None
        arr = data[col].dropna()
        if len(arr) < 2:
            return None, None
        try:
            adf_res = adfuller(arr)
            logging.debug(f"[Stationarity] {col}: ADF = {adf_res[0]:.6f}, p = {adf_res[1]:.6f}")
            return adf_res[0], adf_res[1]
        except Exception as ex:
            logging.error(f"Ошибка ADF для {col}: {ex}")
            return None, None

    def evaluate_noise(self, col: str, use_normalized: bool = False):
        return 0, 0  # ЗАГЛУШКА

    def plot_time_series(self, data: pd.DataFrame, title: str) -> BytesIO:
        fig, ax = plt.subplots(figsize=(8, 4))
        for c in data.columns:
            if pd.api.types.is_numeric_dtype(data[c]):
                ax.plot(data[c].dropna(), label=c)
        ax.set_title(title)
        ax.legend()
        buf = BytesIO()
        plt.tight_layout()
        plt.savefig(buf, format="png")
        buf.seek(0)
        plt.close(fig)
        return buf

    def plot_single_time_series(self, series: pd.Series, title: str) -> BytesIO:
        fig, ax = plt.subplots(figsize=(4, 3))
        if pd.api.types.is_numeric_dtype(series):
            ax.plot(series.dropna(), label=series.name)
        ax.set_title(title)
        ax.legend()
        buf = BytesIO()
        plt.tight_layout()
        plt.savefig(buf, format="png")
        buf.seek(0)
        plt.close(fig)
        return buf

    def plot_fft(self, data: pd.DataFrame, title: str) -> BytesIO:
        fig, ax = plt.subplots(figsize=(8, 4))
        for c in data.columns:
            if pd.api.types.is_numeric_dtype(data[c]):
                freqs, amplitude, _, _ = plt_fft_analysis(data[c])
                if freqs.size > 0:
                    ax.plot(freqs, amplitude, label=c)
        ax.set_title(title)
        ax.set_xlabel("Частота")
        ax.set_ylabel("Амплитуда")
        ax.legend()
        buf = BytesIO()
        plt.tight_layout()
        plt.savefig(buf, format="png", dpi=100)
        buf.seek(0)
        plt.close(fig)
        return buf

    def plot_single_fft(self, series: pd.Series, title: str) -> BytesIO:
        freqs, amplitude, phase, peaks = plt_fft_analysis(series)
        fig, ax = plt.subplots(figsize=(4, 3))
        if freqs.size > 0:
            ax.plot(freqs, amplitude, label=series.name)
            ax.plot(freqs[peaks], amplitude[peaks], "x", label="Peaks")
        ax.set_title(title)
        ax.set_xlabel("Частота")
        ax.set_ylabel("Амплитуда")
        ax.legend()
        buf = BytesIO()
        plt.tight_layout()
        plt.savefig(buf, format="png")
        buf.seek(0)
        plt.close(fig)
        return buf

    def get_heatmaps_for_summary(self, methods_to_get: dict, annotate: bool = False, lag_params: dict = None) -> dict:
        """ Модифицированная версия с опцией аннотации и кастомными лагами. """
        if self.data_normalized.empty: self.normalize_data()
        if lag_params is None:
            lag_params = {}

        generated_heatmaps = {}
        df_for_analysis = self.data_normalized[[c for c in self.data_normalized.columns if c.startswith('c')]]
        if df_for_analysis.empty: return {k: None for k in methods_to_get}

        for report_label, method_variant in methods_to_get.items():
            # ПРЕДЗАДАННЫЕ КОНТРОЛЬНЫЕ для этих там
            control_vars = None
            
            lag = lag_params.get(method_variant, DEFAULT_MAX_LAG)
            
            matrix = compute_connectivity_variant(df_for_analysis, method_variant, lag=lag, control=control_vars)
            
            title = f"{report_label}"
            legend_text = f"Lag={lag}"
            image_buffer = plot_heatmap(matrix, title, legend_text=legend_text, annotate=annotate)
            generated_heatmaps[report_label] = image_buffer
        return generated_heatmaps
    def get_connectomes_for_summary(self, methods_to_get: dict, lag_params: dict = None) -> dict:
      
        if self.data_normalized.empty: self.normalize_data()
        if lag_params is None:
            lag_params = {}

        generated_connectomes = {}
        df_for_analysis = self.data_normalized[[c for c in self.data_normalized.columns if c.startswith('c')]]
        if df_for_analysis.empty: return {k: None for k in methods_to_get}

        for report_label, method_variant in methods_to_get.items():
            control_vars = None
            
            lag = lag_params.get(method_variant, DEFAULT_MAX_LAG)
            
            matrix = compute_connectivity_variant(df_for_analysis, method_variant, lag=lag, control=control_vars)

            if matrix is not None:
                is_directed = is_directed_method(method_variant)
                is_pval = is_pvalue_method(method_variant)
                threshold = 0.05 if is_pval else 0.2
                invert_threshold = True if is_pval else False

                title = f"{report_label}"
                legend_text = f"Lag={lag}"
                image_buffer = plot_connectome(matrix, title, threshold=threshold, directed=is_directed, invert_threshold=invert_threshold, legend_text=legend_text)
                generated_connectomes[report_label] = image_buffer
            else:
                generated_connectomes[report_label] = None
        return generated_connectomes



    # -----------------------------
    # HTML / Site report
    # -----------------------------
    def export_html_report_legacy_v2(
        self,
        output_path: str,
        variants: Optional[List[str]] = None,
        graph_threshold: float = 0.2,
        p_alpha: float = 0.05,
        embed_images: bool = True,
        include_matrix_tables: bool = True,
        max_edges: int = 25,
    ) -> str:
        if not self.results:
            self.run_all_methods()
        df = self.data_normalized if not self.data_normalized.empty else self.data
        variants = variants or [v for v in STABLE_METHODS if v in method_mapping]

        def _b64_png(buf: BytesIO) -> str:
            return base64.b64encode(buf.getvalue()).decode("ascii")

        def _matrix_to_html(mat: np.ndarray, cols: List[str]) -> str:
            if mat is None or not isinstance(mat, np.ndarray) or mat.size == 0:
                return "<div class='muted'>No data</div>"
            out = ["<table class='matrix'><thead><tr><th></th>"]
            for c in cols:
                out.append(f"<th>{_html.escape(str(c))}</th>")
            out.append("</tr></thead><tbody>")
            for i, rname in enumerate(cols):
                out.append(f"<tr><th>{_html.escape(str(rname))}</th>")
                for j in range(len(cols)):
                    v = mat[i, j]
                    if np.isnan(v):
                        s = "NaN"
                    else:
                        s = f"{v:.4g}"
                    out.append(f"<td>{_html.escape(s)}</td>")
                out.append("</tr>")
            out.append("</tbody></table>")
            return "".join(out)

        def _top_edges(mat: np.ndarray, cols: List[str], variant: str) -> List[Tuple[str, str, float]]:
            if mat is None or not isinstance(mat, np.ndarray) or mat.size == 0:
                return []
            edges = []
            n = len(cols)
            directed = is_directed_method(variant)
            pval = is_pvalue_method(variant)
            for i in range(n):
                for j in range(n):
                    if i == j:
                        continue
                    if not directed and j <= i:
                        continue
                    v = mat[i, j]
                    if np.isnan(v):
                        continue
                    edges.append((cols[i], cols[j], float(v)))
            if pval:
                # Для p-value методов сортируем по "силе" связи, но пороги интерпретируем по p.
                edges.sort(key=lambda x: p_to_score(x[2]), reverse=True)
            else:
                edges.sort(key=lambda x: abs(x[2]), reverse=True)
            return edges[:max_edges]

        def _section_anchor_title(variant: str, default_title: str) -> Tuple[str, str]:
            """Возвращает якорь/заголовок для секции метода в HTML-отчёте."""
            v = variant.lower()
            if v == "correlation_partial":
                return "pcorr_block", "Partial correlation (КПК)"
            if v.startswith("correlation"):
                return "corr_block", "Correlation"
            if v == "mutinf_full":
                return "mi_block", "Mutual information"
            if v == "mutinf_partial":
                return "cmi_block", "Conditional mutual information (КЧК)"
            if v.startswith("te_"):
                return "te_block", "Transfer entropy"
            return variant, default_title

        cols = list(df.columns)
        pair_filter = getattr(self, "pair_filter", None)
        show_cols = cols
        if pair_filter:
            involved = sorted(set([x for p in pair_filter for x in p]), key=lambda s: int(str(s)[1:]) if str(s).startswith("c") and str(s)[1:].isdigit() else str(s))
            if len(involved) >= 2:
                show_cols = [c for c in involved if c in cols]
        toc_items = []
        sections = [
            (
                "<section id='data_summary'>"
                "<h2 id='data_summary'>Data summary</h2>"
                f"<div class='muted'>Rows: {len(df)} • Columns: {len(cols)} • Variants: {len(variants)}</div>"
                "</section>"
            )
        ]
        used_section_ids = {"data_summary"}
        for v in variants:
            info = METHOD_INFO.get(v, {"title": v, "meaning": ""})
            section_id, section_title = _section_anchor_title(v, info["title"])
            if section_id in used_section_ids:
                i = 2
                while f"{section_id}_{i}" in used_section_ids:
                    i += 1
                section_id = f"{section_id}_{i}"
            used_section_ids.add(section_id)
            toc_items.append(f"<li><a href='#{_html.escape(section_id)}'>{_html.escape(section_title)}</a></li>")

            lag_block = ""
            if v in self.lag_results and self.lag_results[v]:
                best_lag, median_lag, _ = self.select_lag_metrics(self.lag_results[v], v)
                lag_block = (
                    f"<div class='kv'>"
                    f"<div><b>Best lag</b>: {best_lag}</div>"
                    f"<div><b>Median lag</b>: {median_lag}</div>"
                    f"</div>"
                )

            mat1 = self.results.get(v)
            if mat1 is None:
                mat1 = compute_connectivity_variant(df, v, lag=1, control=None)

            heat = plot_heatmap(mat1, f"{v} heatmap", legend_text="Lag=1")
            thr = p_alpha if is_pvalue_method(v) else graph_threshold
            conn = plot_connectome(
                mat1,
                f"{v} connectome",
                threshold=thr,
                directed=is_directed_method(v),
                invert_threshold=is_pvalue_method(v),
                legend_text=f"Lag=1; thr={thr}",
            )
            if embed_images:
                heat_html = f"<img class='img' src='data:image/png;base64,{_b64_png(heat)}'/>"
                conn_html = f"<img class='img' src='data:image/png;base64,{_b64_png(conn)}'/>"
            else:
                heat_html = "<div class='muted'>embed_images=False</div>"
                conn_html = "<div class='muted'>embed_images=False</div>"

            edges = _top_edges(mat1, cols, v)
            edges_html = ["<ol class='edges'>"]
            for a, b, val in edges:
                if is_pvalue_method(v):
                    score = p_to_score(val)
                    line = f"{_html.escape(a)} → {_html.escape(b)}: p = {val:.4g}, score = {score:.3f}"
                    edges_html.append(f"<li>{line}</li>")
                else:
                    edges_html.append(f"<li>{_html.escape(a)} → {_html.escape(b)}: {val:.4g}</li>" if is_directed_method(v)
                                     else f"<li>{_html.escape(a)} — {_html.escape(b)}: {val:.4g}</li>")
            edges_html.append("</ol>")

            table_html = _matrix_to_html(mat1, cols) if include_matrix_tables else "<div class='muted'>matrix table disabled</div>"
            te_params_block = ""
            if v.lower().startswith("te_"):
                lag_max = max(self.lag_ranges.get(v, range(1, 2)))
                te_k = DEFAULT_BINS
                te_l = 1
                n_rows = len(df)
                te_params_block = (
                    f"<p><b>TE parameters:</b> k={te_k}, l={te_l}, lag_max={lag_max}, N={n_rows}</p>"
                )

            sections.append(
                f"<section id='{_html.escape(section_id)}'>"
                f"<h2 id='{_html.escape(section_id)}'>{_html.escape(section_title)}</h2>"
                f"<div class='muted'>{_html.escape(info.get('meaning',''))}</div>"
                f"{lag_block}"
                f"{te_params_block}"
                f"<h3>Matrix (heatmap)</h3>{heat_html}"
                f"<h3>Connectome graph</h3>{conn_html}"
                f"<h3>Top edges</h3>{''.join(edges_html)}"
                f"<h3>Matrix (Lag=1)</h3>{table_html}"
                f"</section>"
            )

        html_doc = f"""<!doctype html>
<html lang="ru">
<head>
  <meta charset="utf-8"/>
  <meta name="viewport" content="width=device-width,initial-scale=1"/>
  <title>Time Series Connectivity Report</title>
  <style>
    body{{font-family:system-ui,-apple-system,Segoe UI,Roboto,Arial; margin:0; padding:0; background:#0b0b0c; color:#eaeaea;}}
    header{{padding:20px 22px; border-bottom:1px solid #222; position:sticky; top:0; background:#0b0b0c; z-index:10;}}
    main{{display:grid; grid-template-columns: 320px 1fr; gap:18px; padding:18px 22px;}}
    nav{{border:1px solid #222; border-radius:12px; padding:14px; height:calc(100vh - 110px); overflow:auto;}}
    nav ul{{margin:0; padding-left:18px;}}
    a{{color:#8ab4f8; text-decoration:none;}}
    a:hover{{text-decoration:underline;}}
    section{{border:1px solid #222; border-radius:12px; padding:16px; margin-bottom:16px; background:#111;}}
    .muted{{color:#b7b7b7; font-size:13px;}}
    .grid{{display:grid; grid-template-columns: 1fr 1fr; gap:10px; margin-top:10px;}}
    .img{{max-width:100%; border-radius:10px; border:1px solid #222; background:#0b0b0c;}}
    .matrix{{border-collapse:collapse; width:100%; overflow:auto; display:block;}}
    .matrix th,.matrix td{{border:1px solid #2a2a2a; padding:6px 8px; font-size:12px; white-space:nowrap;}}
    .matrix th{{position:sticky; left:0; background:#141414;}}
    .edges{{font-size:13px;}}
    .kv{{display:flex; gap:16px; margin-top:8px; font-size:13px;}}
    @media (max-width: 1000px){{ main{{grid-template-columns: 1fr;}} nav{{height:auto;}} .grid{{grid-template-columns:1fr;}} }}
  </style>
</head>
<body>
  <div style="position:fixed; right:20px; top:20px; width:260px;
              background:#f7f7f7; color:#111; border:1px solid #ccc; padding:12px;
              font-family:Arial; font-size:14px; z-index:9999; border-radius:8px;">
    <b>Contents</b><br>
    <a href="#data_summary">Data summary</a><br>
    <a href="#corr_block">Correlation</a><br>
    <a href="#mi_block">Mutual information</a><br>
    <a href="#pcorr_block">Partial correlation (КПК)</a><br>
    <a href="#cmi_block">Conditional MI (КЧК)</a><br>
    <a href="#te_block">Transfer entropy</a><br>
  </div>
  <header>
    <div style="font-size:18px; font-weight:700;">Time Series Connectivity Report</div>
    <div class="muted">Методы: {len(variants)} • Переменные: {len(cols)} • Длина ряда: {len(df)}</div>
  </header>
  <main>
    <nav>
      <div style="font-weight:700; margin-bottom:8px;">Оглавление</div>
      <ul>{''.join(toc_items)}</ul>
    </nav>
    <div>
      {''.join(sections)}
    </div>
  </main>
</body>
</html>"""

        out = Path(output_path)
        out.parent.mkdir(parents=True, exist_ok=True)
        out.write_text(html_doc, encoding="utf-8")
        logging.info(f"[Export] HTML report saved: {str(out)}")
        return str(out)

    def export_site_report(
        self,
        out_dir: str,
        variants: Optional[List[str]] = None,
        graph_threshold: float = 0.2,
        p_alpha: float = 0.05,
        include_matrix_tables: bool = True,
        max_edges: int = 25,
        zip_path: Optional[str] = None,
    ) -> str:
        """
        Мини-сайт: index.html + assets/*.png. Опционально упаковывает в zip_path.
        """
        if not self.results:
            self.run_all_methods()
        df = self.data_normalized if not self.data_normalized.empty else self.data
        variants = variants or [v for v in STABLE_METHODS if v in method_mapping]
        outp = Path(out_dir)
        assets = outp / "assets"
        assets.mkdir(parents=True, exist_ok=True)

        cols = list(df.columns)
        pair_filter = getattr(self, "pair_filter", None)
        show_cols = cols
        if pair_filter:
            involved = sorted(set([x for p in pair_filter for x in p]), key=lambda s: int(str(s)[1:]) if str(s).startswith("c") and str(s)[1:].isdigit() else str(s))
            if len(involved) >= 2:
                show_cols = [c for c in involved if c in cols]

        def _save_png(buf: BytesIO, name: str) -> str:
            p = assets / name
            p.write_bytes(buf.getvalue())
            return f"assets/{name}"

        def _matrix_to_html_legacy(mat: np.ndarray, cols_: List[str]) -> str:
            if mat is None or not isinstance(mat, np.ndarray) or mat.size == 0:
                return "<div class='muted'>No data</div>"
            out = ["<table class='matrix'><thead><tr><th></th>"]
            for c in cols_:
                out.append(f"<th>{_html.escape(str(c))}</th>")
            out.append("</tr></thead><tbody>")
            for i, rname in enumerate(cols_):
                out.append(f"<tr><th>{_html.escape(str(rname))}</th>")
                for j in range(len(cols_)):
                    v = mat[i, j]
                    s = "NaN" if np.isnan(v) else f"{v:.4g}"
                    out.append(f"<td>{_html.escape(s)}</td>")
                out.append("</tr>")
            out.append("</tbody></table>")
            return "".join(out)

        def _top_edges(mat: np.ndarray, cols_: List[str], variant: str) -> List[Tuple[str, str, float]]:
            if mat is None or not isinstance(mat, np.ndarray) or mat.size == 0:
                return []
            edges = []
            n = len(cols_)
            directed = is_directed_method(variant)
            pval = is_pvalue_method(variant)
            for i in range(n):
                for j in range(n):
                    if i == j:
                        continue
                    if not directed and j <= i:
                        continue
                    v = mat[i, j]
                    if np.isnan(v):
                        continue
                    edges.append((cols_[i], cols_[j], float(v)))
            if pval:
                # Для p-value методов сортируем по -log10(p), чтобы Top edges были визуально сопоставимыми.
                edges.sort(key=lambda x: p_to_score(x[2]), reverse=True)
            else:
                edges.sort(key=lambda x: abs(x[2]), reverse=True)
            return edges[:max_edges]

        def _section_anchor_title(variant: str, default_title: str) -> Tuple[str, str]:
            """Возвращает якорь/заголовок для секции метода в mini-site отчёте."""
            v = variant.lower()
            if v == "correlation_partial":
                return "pcorr_block", "Partial correlation (КПК)"
            if v.startswith("correlation"):
                return "corr_block", "Correlation"
            if v == "mutinf_full":
                return "mi_block", "Mutual information"
            if v == "mutinf_partial":
                return "cmi_block", "Conditional mutual information (КЧК)"
            if v.startswith("te_"):
                return "te_block", "Transfer entropy"
            return variant, default_title

        toc_items = []
        sections = [
            (
                "<section id='data_summary'>"
                "<h2 id='data_summary'>Data summary</h2>"
                f"<div class='muted'>Rows: {len(df)} • Columns: {len(cols)} • Variants: {len(variants)}</div>"
                "</section>"
            )
        ]
        used_section_ids = {"data_summary"}
        for v in variants:
            info = METHOD_INFO.get(v, {"title": v, "meaning": ""})
            section_id, section_title = _section_anchor_title(v, info["title"])
            if section_id in used_section_ids:
                i = 2
                while f"{section_id}_{i}" in used_section_ids:
                    i += 1
                section_id = f"{section_id}_{i}"
            used_section_ids.add(section_id)
            toc_items.append(f"<li><a href='#{_html.escape(section_id)}'>{_html.escape(section_title)}</a></li>")

            mat1 = self.results.get(v)
            if mat1 is None:
                mat1 = compute_connectivity_variant(df, v, lag=1, control=None)

            heat = plot_heatmap(mat1, f"{v} heatmap", legend_text="Lag=1")
            thr = p_alpha if is_pvalue_method(v) else graph_threshold
            conn = plot_connectome(
                mat1,
                f"{v} connectome",
                threshold=thr,
                directed=is_directed_method(v),
                invert_threshold=is_pvalue_method(v),
                legend_text=f"Lag=1; thr={thr}",
            )
            heat_path = _save_png(heat, f"{v}_heatmap.png")
            conn_path = _save_png(conn, f"{v}_connectome.png")

            edges = _top_edges(mat1, cols, v)
            edges_html = ["<ol class='edges'>"]
            for a, b, val in edges:
                if is_pvalue_method(v):
                    score = p_to_score(val)
                    line = f"{_html.escape(a)} → {_html.escape(b)}: p = {val:.4g}, score = {score:.3f}"
                    edges_html.append(f"<li>{line}</li>")
                else:
                    edges_html.append(f"<li>{_html.escape(a)} → {_html.escape(b)}: {val:.4g}</li>" if is_directed_method(v)
                                     else f"<li>{_html.escape(a)} — {_html.escape(b)}: {val:.4g}</li>")
            edges_html.append("</ol>")

            table_html = _matrix_to_html(mat1, cols) if include_matrix_tables else "<div class='muted'>matrix table disabled</div>"
            te_params_block = ""
            if v.lower().startswith("te_"):
                lag_max = max(self.lag_ranges.get(v, range(1, 2)))
                te_k = DEFAULT_BINS
                te_l = 1
                n_rows = len(df)
                te_params_block = (
                    f"<p><b>TE parameters:</b> k={te_k}, l={te_l}, lag_max={lag_max}, N={n_rows}</p>"
                )
            sections.append(
                f"<section id='{_html.escape(section_id)}'>"
                f"<h2 id='{_html.escape(section_id)}'>{_html.escape(section_title)}</h2>"
                f"<div class='muted'>{_html.escape(info.get('meaning',''))}</div>"
                f"{te_params_block}"
                f"<h3>Matrix (heatmap)</h3><img class='img' src='{heat_path}'/>"
                f"<h3>Connectome graph</h3><img class='img' src='{conn_path}'/>"
                f"<h3>Top edges</h3>{''.join(edges_html)}"
                f"<h3>Matrix (Lag=1)</h3>{table_html}"
                f"</section>"
            )

        index_html = f"""<!doctype html>
<html lang="ru">
<head>
  <meta charset="utf-8"/>
  <meta name="viewport" content="width=device-width,initial-scale=1"/>
  <title>Time Series Connectivity Report</title>
  <style>
    body{{font-family:system-ui,-apple-system,Segoe UI,Roboto,Arial; margin:0; padding:0; background:#0b0b0c; color:#eaeaea;}}
    header{{padding:20px 22px; border-bottom:1px solid #222; position:sticky; top:0; background:#0b0b0c; z-index:10;}}
    main{{display:grid; grid-template-columns: 320px 1fr; gap:18px; padding:18px 22px;}}
    nav{{border:1px solid #222; border-radius:12px; padding:14px; height:calc(100vh - 110px); overflow:auto;}}
    nav ul{{margin:0; padding-left:18px;}}
    a{{color:#8ab4f8; text-decoration:none;}}
    a:hover{{text-decoration:underline;}}
    section{{border:1px solid #222; border-radius:12px; padding:16px; margin-bottom:16px; background:#111;}}
    .muted{{color:#b7b7b7; font-size:13px;}}
    .grid{{display:grid; grid-template-columns: 1fr 1fr; gap:10px; margin-top:10px;}}
    .img{{max-width:100%; border-radius:10px; border:1px solid #222; background:#0b0b0c;}}
    .matrix{{border-collapse:collapse; width:100%; overflow:auto; display:block;}}
    .matrix th,.matrix td{{border:1px solid #2a2a2a; padding:6px 8px; font-size:12px; white-space:nowrap;}}
    .matrix th{{position:sticky; left:0; background:#141414;}}
    .edges{{font-size:13px;}}
    @media (max-width: 1000px){{ main{{grid-template-columns: 1fr;}} nav{{height:auto;}} .grid{{grid-template-columns:1fr;}} }}
  </style>
</head>
<body>
  <div style="position:fixed; right:20px; top:20px; width:260px;
              background:#f7f7f7; color:#111; border:1px solid #ccc; padding:12px;
              font-family:Arial; font-size:14px; z-index:9999; border-radius:8px;">
    <b>Contents</b><br>
    <a href="#data_summary">Data summary</a><br>
    <a href="#corr_block">Correlation</a><br>
    <a href="#mi_block">Mutual information</a><br>
    <a href="#pcorr_block">Partial correlation (КПК)</a><br>
    <a href="#cmi_block">Conditional MI (КЧК)</a><br>
    <a href="#te_block">Transfer entropy</a><br>
  </div>
  <header>
    <div style="font-size:18px; font-weight:700;">Time Series Connectivity Report</div>
    <div class="muted">Методы: {len(variants)} • Переменные: {len(cols)} • Длина ряда: {len(df)}</div>
  </header>
  <main>
    <nav>
      <div style="font-weight:700; margin-bottom:8px;">Оглавление</div>
      <ul>{''.join(toc_items)}</ul>
    </nav>
    <div>
      {''.join(sections)}
    </div>
  </main>
</body>
</html>"""

        (outp / "index.html").write_text(index_html, encoding="utf-8")

        if zip_path:
            zp = Path(zip_path)
            zp.parent.mkdir(parents=True, exist_ok=True)
            base = str(zp).removesuffix(".zip")
            shutil.make_archive(base, "zip", root_dir=str(outp))
            return str(Path(base + ".zip"))
        return str(outp)



    # ---------------------------
    # Fixed + extended execution helpers
    # ---------------------------

    def run_selected_methods_legacy_simple(
        self,
        variants: List[str],
        *,
        max_lag: int = DEFAULT_MAX_LAG,
        compute_lag_sweep: bool = True,
        pick_best_lag: bool = True,
        sliding_windows: bool = False,
        window_sizes: Optional[List[int]] = None,
        overlap: int = 50,
    ) -> Dict[str, int]:
        """
        Run only selected methods and (optionally) optimize lag / run sliding-window diagnostics.

        Returns: dict variant -> chosen_lag (or 1 for non-lag methods)

        Side effects (for reporting):
          - self.results[variant]               : final matrix (chosen lag)
          - self.variant_lags[variant]          : chosen lag
          - self.lag_sweep[variant][lag]        : dict(score=float, matrix=np.ndarray)
          - self.window_sweep[variant]          : dict(window_size -> dict(start_metrics, best, etc.))
          - self.best_window[variant]           : dict with best window across sizes
        """
        self.normalize_data()
        if self.data_normalized.empty:
            logging.warning("[RunSelected] No data to run analysis.")
            self.results = {}
            return {}

        allowed = set(method_mapping.keys())

        variants = [v for v in variants if v in allowed]
        if not variants:
            logging.warning("[RunSelected] Empty method list after filtering.")
            self.results = {}
            return {}

        def _score_matrix(mat: np.ndarray, variant: str) -> float:
            if mat is None or not isinstance(mat, np.ndarray) or mat.size == 0:
                return float("-inf")
            n = mat.shape[0]
            mask = ~np.eye(n, dtype=bool)
            vals = mat[mask]
            vals = vals[np.isfinite(vals)]
            if vals.size == 0:
                return float("-inf")
            if get_method_spec(variant).is_p_value:
                # smaller p-values are better -> invert
                return -float(np.nanmean(vals))
            return float(np.nanmean(np.abs(vals)))

        self.results = {}
        self.variant_lags = {}
        self.lag_sweep = {}
        self.window_sweep = {}
        self.best_window = {}

        max_lag = max(1, int(max_lag))

        for variant in variants:
            spec = get_method_spec(variant)

            # ---- Lag sweep / optimization
            chosen_lag = 1
            if spec.supports_lag and (compute_lag_sweep or pick_best_lag):
                sweep: Dict[int, Dict[str, Any]] = {}
                best_lag = 1
                best_score = float("-inf")
                for lag in range(1, max_lag + 1):
                    mat = compute_connectivity_variant(self.data_normalized, variant, lag=lag, control=partial_controls, partial_mode=partial_mode, pairwise_policy=pairwise_policy, custom_controls=custom_controls)
                    s = _score_matrix(mat, variant)
                    sweep[lag] = {"score": s, "matrix": mat}
                    if s > best_score:
                        best_score = s
                        best_lag = lag
                self.lag_sweep[variant] = sweep
                chosen_lag = best_lag if pick_best_lag else 1
            else:
                self.lag_sweep[variant] = {}

            self.variant_lags[variant] = chosen_lag
            self.results[variant] = compute_connectivity_variant(self.data_normalized, variant, lag=chosen_lag, control=partial_controls, partial_mode=partial_mode, pairwise_policy=pairwise_policy, custom_controls=custom_controls)

            # ---- Sliding windows diagnostics (optional)
            if sliding_windows:
                wsizes = window_sizes or [50, 100, 500]
                # normalize sizes to sane values
                wsizes = [int(w) for w in wsizes if w and int(w) > 5]
                if not wsizes:
                    wsizes = [50, 100, 500]

                per_size: Dict[int, Dict[str, Any]] = {}
                best_overall = {"window_size": None, "start": None, "score": float("-inf"), "matrix": None}

                for w in wsizes:
                    # analyze_sliding_windows_with_metric returns dict with per-window metrics + best window
                    res = analyze_sliding_windows_with_metric(self.data_normalized, variant, min(w, len(self.data_normalized)), min(overlap, max(1, w // 2)))
                    # Expected keys: 'window_metrics' (list or dict), 'best' (start, metric, matrix)
                    per_size[w] = res

                    # infer "best" record robustly
                    best = None
                    if isinstance(res, dict):
                        best = res.get("best") or res.get("best_window") or res.get("best_result")
                    if isinstance(best, dict):
                        s = float(best.get("metric", best.get("score", float("-inf"))))
                        if s > best_overall["score"]:
                            best_overall = {
                                "window_size": w,
                                "start": int(best.get("start", best.get("window_start", 0))),
                                "score": s,
                                "matrix": best.get("matrix"),
                            }

                self.window_sweep[variant] = per_size
                self.best_window[variant] = best_overall

        return dict(self.variant_lags)

    # ---------------------------
    # Carousel-style HTML report additions
    # ---------------------------

    def export_html_report_legacy_v3(
        self,
        output_path: str,
        variants: Optional[List[str]] = None,
        graph_threshold: float = 0.2,
        p_alpha: float = 0.05,
        embed_images: bool = True,
        include_matrix_tables: bool = True,
        max_edges: int = 25,
        *,
        carousel: bool = True,
    ) -> str:
        """Export a single self-contained HTML report.

        If carousel=True, each method section includes a slide carousel:
          - lag=1 heatmap, lag=max heatmap (if lag sweep exists)
          - chosen lag heatmap
          - lag quality curve
          - min window size heatmap, max window size heatmap (if windows exist)
          - best window heatmap
          - window quality curve (for best window size)
        """
        if not self.results:
            self.run_all_methods()

        df = self.data_normalized if not self.data_normalized.empty else self.data
        cols = list(df.columns)
        pair_filter = getattr(self, "pair_filter", None)
        show_cols = cols
        if pair_filter:
            involved = sorted(set([x for p in pair_filter for x in p]), key=lambda s: int(str(s)[1:]) if str(s).startswith("c") and str(s)[1:].isdigit() else str(s))
            if len(involved) >= 2:
                show_cols = [c for c in involved if c in cols]

        variants = variants or [v for v in STABLE_METHODS if v in method_mapping]

        def _b64_png(buf: BytesIO) -> str:
            return base64.b64encode(buf.getvalue()).decode("ascii")

        def _heatmap_png(mat: np.ndarray, title: str) -> Optional[str]:
            if mat is None or not isinstance(mat, np.ndarray) or mat.size == 0:
                return None
            fig = plt.figure(figsize=(5.2, 4.6))
            ax = fig.add_subplot(111)
            im = ax.imshow(mat, aspect="auto")
            ax.set_title(title, fontsize=10)
            ax.set_xticks(range(len(show_cols)))
            ax.set_yticks(range(len(show_cols)))
            ax.set_xticklabels(cols, rotation=90, fontsize=7)
            ax.set_yticklabels(cols, fontsize=7)
            fig.colorbar(im, ax=ax, fraction=0.046, pad=0.04)
            buf = BytesIO()
            fig.tight_layout()
            fig.savefig(buf, format="png", dpi=140)
            plt.close(fig)
            buf.seek(0)
            return _b64_png(buf)

        def _curve_png(xs: List[int], ys: List[float], title: str, xlabel: str) -> Optional[str]:
            if not xs or not ys:
                return None
            fig = plt.figure(figsize=(5.2, 3.3))
            ax = fig.add_subplot(111)
            ax.plot(xs, ys, marker="o")
            ax.set_title(title, fontsize=10)
            ax.set_xlabel(xlabel)
            ax.set_ylabel("quality")
            ax.grid(True, alpha=0.3)
            buf = BytesIO()
            fig.tight_layout()
            fig.savefig(buf, format="png", dpi=140)
            plt.close(fig)
            buf.seek(0)
            return _b64_png(buf)

        # Keep the original implementation by calling the old generator if present.
        # We reuse a simplified HTML skeleton here to avoid breaking other parts.
        html_parts: List[str] = []
        html_parts.append("""<!doctype html>
<html lang="en">
<head>
<meta charset="utf-8" />
<meta name="viewport" content="width=device-width, initial-scale=1" />
<title>Time Series Connectivity Report</title>
<style>
body{font-family:system-ui,-apple-system,Segoe UI,Roboto,Arial,sans-serif;margin:0;background:#0b0b0f;color:#eaeaf0}
main{max-width:1200px;margin:0 auto;padding:24px}
h1{font-size:22px;margin:0 0 14px}
h2{font-size:18px;margin:26px 0 10px}
.card{background:#13131a;border:1px solid #242433;border-radius:12px;padding:14px 14px;margin:14px 0}
.muted{color:#a7a7b7}
hr{border:0;border-top:1px solid #242433;margin:18px 0}
.carousel{position:relative;border:1px solid #242433;border-radius:12px;overflow:hidden;background:#0f0f15}
.slides{display:flex;transition:transform .25s ease}
.slide{min-width:100%;padding:12px;box-sizing:border-box}
.slide img{max-width:100%;height:auto;border-radius:10px;border:1px solid #242433;background:#0b0b0f}
.slide .cap{margin-top:8px;font-size:12px;color:#c9c9d6}
.car-ctrl{display:flex;gap:8px;align-items:center;justify-content:space-between;margin-top:10px}
button{background:#1b1b25;border:1px solid #2c2c3d;color:#eaeaf0;border-radius:10px;padding:8px 10px;cursor:pointer}
button:hover{background:#232332}
.pill{font-size:12px;color:#c9c9d6}
.matrix{border-collapse:collapse;width:100%;font-size:12px}
.matrix th,.matrix td{border:1px solid #2c2c3d;padding:4px 6px;text-align:right}
.matrix th{background:#171722;color:#eaeaf0;position:sticky;top:0}
</style>
</head>
<body><main>
""")

        html_parts.append(f"<h1>Connectivity report</h1><div class='muted'>n={len(df)} series={len(cols)}</div><hr/>")

        # JS for carousels
        html_parts.append("""<script>
function carPrev(id){
  const root=document.getElementById(id);
  const slides=root.querySelector('.slides');
  const n=parseInt(root.dataset.n,10);
  let i=parseInt(root.dataset.i,10);
  i=(i-1+n)%n;
  root.dataset.i=i;
  slides.style.transform=`translateX(${-100*i}%)`;
  root.querySelector('.pill').textContent = `${i+1}/${n}`;
}
function carNext(id){
  const root=document.getElementById(id);
  const slides=root.querySelector('.slides');
  const n=parseInt(root.dataset.n,10);
  let i=parseInt(root.dataset.i,10);
  i=(i+1)%n;
  root.dataset.i=i;
  slides.style.transform=`translateX(${-100*i}%)`;
  root.querySelector('.pill').textContent = `${i+1}/${n}`;
}
</script>""")

        for variant in variants:
            if variant not in self.results:
                continue

            chosen_lag = getattr(self, "variant_lags", {}).get(variant, 1)
            mat_chosen = self.results.get(variant)

            html_parts.append(f"<h2 id='{_html.escape(variant)}'>{_html.escape(variant)}</h2>")
            html_parts.append("<div class='card'>")

            # Build carousel slides
            slides = []

            # Lag slides
            sweep = getattr(self, "lag_sweep", {}).get(variant) or {}
            if sweep:
                min_lag = min(sweep.keys())
                max_lag = max(sweep.keys())
                m1 = sweep[min_lag].get("matrix")
                m2 = sweep[max_lag].get("matrix")
                p1 = _heatmap_png(m1, f"{variant} (lag={min_lag})")
                p2 = _heatmap_png(m2, f"{variant} (lag={max_lag})")
                if p1:
                    slides.append((p1, f"Lag min: {min_lag}"))
                if p2:
                    slides.append((p2, f"Lag max: {max_lag}"))
                # curve
                xs = sorted(sweep.keys())
                ys = [float(sweep[x].get("score", float("nan"))) for x in xs]
                pc = _curve_png(xs, ys, f"{variant}: lag sweep", "lag")
                if pc:
                    slides.append((pc, "Quality vs lag"))

            # Chosen lag matrix
            pc = _heatmap_png(mat_chosen, f"{variant} (chosen lag={chosen_lag})")
            if pc:
                slides.append((pc, f"Chosen lag: {chosen_lag}"))

            # Window slides
            win = getattr(self, "window_sweep", {}).get(variant) or {}
            if win:
                wsizes = sorted([int(k) for k in win.keys()])
                if wsizes:
                    wmin, wmax = wsizes[0], wsizes[-1]

                    def _best_matrix_for_w(w: int):
                        r = win.get(w, {}) or {}
                        best = None
                        if isinstance(r, dict):
                            best = r.get("best") or r.get("best_window") or r.get("best_result")
                        if isinstance(best, dict):
                            return best.get("matrix"), float(best.get("metric", best.get("score", float("nan")))), int(best.get("start", best.get("window_start", 0)))
                        return None, float("nan"), 0

                    mmin, smin, stmin = _best_matrix_for_w(wmin)
                    mmax, smax, stmax = _best_matrix_for_w(wmax)
                    pmin = _heatmap_png(mmin, f"{variant} (window={wmin}, best start={stmin})")
                    pmax = _heatmap_png(mmax, f"{variant} (window={wmax}, best start={stmax})")
                    if pmin:
                        slides.append((pmin, f"Window min: {wmin} (best start {stmin}, score {smin:.3g})"))
                    if pmax:
                        slides.append((pmax, f"Window max: {wmax} (best start {stmax}, score {smax:.3g})"))
                    # curve for the best window size (metric vs start)
                    best_overall = getattr(self, "best_window", {}).get(variant) or {}
                    bw = best_overall.get("window_size", wmin)
                    r = win.get(int(bw), {}) or {}
                    metrics = r.get("window_metrics") or r.get("metrics") or r.get("per_window") or None
                    xs=[]
                    ys=[]
                    if isinstance(metrics, list):
                        # list of dicts: {start, metric}
                        for it in metrics:
                            if not isinstance(it, dict):
                                continue
                            xs.append(int(it.get("start", it.get("window_start", 0))))
                            ys.append(float(it.get("metric", it.get("score", float("nan")))))
                    elif isinstance(metrics, dict):
                        # dict start->metric
                        for k,v in metrics.items():
                            xs.append(int(k))
                            ys.append(float(v))
                    if xs and ys:
                        order=np.argsort(xs)
                        xs=[xs[i] for i in order]
                        ys=[ys[i] for i in order]
                        pw = _curve_png(xs, ys, f"{variant}: window sweep (window={bw})", "window start")
                        if pw:
                            slides.append((pw, "Quality vs window start"))

                    # best window heatmap
                    bmat = best_overall.get("matrix")
                    if bmat is not None:
                        pb = _heatmap_png(bmat, f"{variant} (best window={best_overall.get('window_size')}, start={best_overall.get('start')})")
                        if pb:
                            slides.append((pb, f"Best window overall: size {best_overall.get('window_size')}, start {best_overall.get('start')}"))

            if carousel and slides:
                car_id = f"car_{re.sub(r'[^a-zA-Z0-9_]+','_',variant)}"
                html_parts.append(f"<div class='carousel' id='{car_id}' data-i='0' data-n='{len(slides)}'>")
                html_parts.append("<div class='slides'>")
                for b64, cap in slides:
                    html_parts.append("<div class='slide'>")
                    html_parts.append(f"<img src='data:image/png;base64,{b64}' alt='plot' />")
                    html_parts.append(f"<div class='cap'>{_html.escape(cap)}</div>")
                    html_parts.append("</div>")
                html_parts.append("</div>")  # slides
                html_parts.append("<div class='car-ctrl'>")
                html_parts.append(f"<button onclick=\"carPrev('{car_id}')\">◀</button>")
                html_parts.append(f"<span class='pill'>1/{len(slides)}</span>")
                html_parts.append(f"<button onclick=\"carNext('{car_id}')\">▶</button>")
                html_parts.append("</div>")
                html_parts.append("</div>")
            else:
                html_parts.append("<div class='muted'>No carousel data. Run with compute_lag_sweep/sliding_windows enabled.</div>")

            # Optional numeric matrix table
            if include_matrix_tables and isinstance(mat_chosen, np.ndarray) and mat_chosen.size:
                html_parts.append("<hr/><div class='muted'>Matrix (chosen lag)</div>")
                # use existing helper if available in module scope
                try:
                    html_parts.append(_matrix_to_html(mat_chosen, cols))  # type: ignore[name-defined]
                except Exception:
                    # fallback: simple table
                    html_parts.append("<table class='matrix'><thead><tr><th></th>" + "".join(f"<th>{_html.escape(str(c))}</th>" for c in cols) + "</tr></thead><tbody>")
                    for i,rn in enumerate(cols):
                        html_parts.append(f"<tr><th>{_html.escape(str(rn))}</th>" + "".join(f"<td>{mat_chosen[i,j]:.4g}</td>" for j in range(len(cols))) + "</tr>")
                    html_parts.append("</tbody></table>")

            html_parts.append("</div>")  # card

        html_parts.append("</main></body></html>")
        Path(output_path).parent.mkdir(parents=True, exist_ok=True)
        Path(output_path).write_text("\n".join(html_parts), encoding="utf-8")
        return str(output_path)


if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="Compute connectivity measures for multivariate time series."
    )
    parser.add_argument(
        "input_file",
        help="Path to input CSV or Excel file with time series data",
    )
    parser.add_argument(
        "--lags",
        type=int,
        default=DEFAULT_MAX_LAG,
        help="Lag or model order (for Granger, TE, etc.)",
    )
    parser.add_argument(
        "--pvalue-alpha",
        type=float,
        default=DEFAULT_PVALUE_ALPHA,
        help="Alpha for p-value methods (Granger full/directed)",
    )
    parser.add_argument(
        "--log",
        action="store_true",
        help="Apply logarithm transform to data (for positive-valued data)",
    )
    parser.add_argument(
        "--no-outliers",
        action="store_true",
        help="Disable outlier removal",
    )
    parser.add_argument(
        "--no-normalize",
        action="store_true",
        help="Disable normalization of data",
    )
    parser.add_argument(
        "--no-stationarity-check",
        action="store_true",
        help="Disable stationarity check (ADF test)",
    )
    parser.add_argument(
        "--graph-threshold",
        type=float,
        default=0.5,
        help="Threshold for graph edges (weight >= threshold)",
    )
    parser.add_argument(
        "--output",
        default=None,
        help="Output Excel file path (defaults to TimeSeriesAnalysis/AllMethods_Full.xlsx)",
    )
    parser.add_argument(
        "--quiet-warnings",
        action="store_true",
        help="Suppress most warnings for cleaner CLI output.",
    )
    parser.add_argument(
        "--experimental",
        action="store_true",
        help="Enable experimental sliding-window analyses.",
    )
    args = parser.parse_args()

    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s %(levelname)s %(message)s",
    )
    configure_warnings(quiet=args.quiet_warnings)

    filepath = os.path.abspath(args.input_file)
    output_path = args.output or os.path.join(save_folder, "AllMethods_Full.xlsx")
    output_dir = os.path.dirname(output_path)
    if output_dir:
        os.makedirs(output_dir, exist_ok=True)

    tool = BigMasterTool(enable_experimental=args.experimental)
    tool.lag_ranges = {v: range(1, args.lags + 1) for v in method_mapping}
    tool.load_data_excel(
        filepath,
        log_transform=args.log,
        remove_outliers=not args.no_outliers,
        normalize=not args.no_normalize,
        fill_missing=True,
        check_stationarity=not args.no_stationarity_check,
    )
    tool.run_all_methods(precompute_controls=not args.no_excel, precompute_pairs=not args.no_excel)
    do_excel = not args.no_excel
    do_report = bool(args.report_html)
    if not do_excel and not do_report:
        do_excel = True

    if do_excel:
        tool.export_big_excel(
            output_path,
            threshold=args.graph_threshold,
            p_value_alpha=args.pvalue_alpha,
            window_size=100,
            overlap=50,
            log_transform=args.log,
            remove_outliers=not args.no_outliers,
            normalize=not args.no_normalize,
            fill_missing=True,
            check_stationarity=not args.no_stationarity_check,
        )

    if do_report:
        report_path = os.path.abspath(args.report_html)
        report_dir = os.path.dirname(report_path)
        if report_dir:
            os.makedirs(report_dir, exist_ok=True)
        tool.export_html_report(
            report_path,
            graph_threshold=args.graph_threshold,
            p_value_alpha=args.pvalue_alpha,
        )

    print("Анализ завершён, результаты сохранены в:", output_path)


############################################################
# HOTFIX: missing BigMasterTool methods
#
# A previous patch introduced an indentation regression that
# effectively "closed" the BigMasterTool class too early.
# The HTML diagnostics expects these methods to exist.
#
# We restore them here in a backward-compatible way via
# monkey-patching: if the attribute is missing, we add it.
############################################################


def _bmt_test_stationarity(self, col: str, use_normalized: bool = False):
    """ADF test: returns (statistic, pvalue) or (None, None)."""
    data = self.data_normalized if use_normalized else self.data
    if data is None or data.empty:
        return None, None
    if col not in data.columns or not pd.api.types.is_numeric_dtype(data[col]):
        return None, None
    arr = pd.to_numeric(data[col], errors="coerce").dropna().values
    if arr.size < 3:
        return None, None
    try:
        stat, pval = adfuller(arr)[:2]
        return float(stat), float(pval)
    except Exception as ex:
        logging.error(f"[ADF] {col}: {ex}")
        return None, None


def _bmt_compute_sample_entropy(self, x) -> float:
    """Sample entropy (nolds.sampen). Returns NaN on failure."""
    try:
        if not _nolds_or_warn():
            return float("nan")
        arr = np.asarray(x, dtype=np.float64).reshape(-1)
        arr = arr[np.isfinite(arr)]
        if arr.size < 20:
            return float("nan")
        if np.nanstd(arr) < 1e-12:
            return float("nan")
        return float(nolds.sampen(arr))
    except Exception as ex:
        logging.error(f"[Sample Entropy] {ex}")
        return float("nan")


def _bmt_detect_seasonality(self, series: pd.Series, threshold: float = 0.2):
    """FFT peak search: returns (peak_freqs, periods) or (None, None)."""
    try:
        s = pd.to_numeric(series, errors="coerce").dropna().values.astype(np.float64)
        if s.size < 8:
            return None, None
        # remove mean to reduce DC
        s = s - np.mean(s)
        yf = fft(s)
        # fs can accidentally be stored as an array/scalar-like; make it a clean float.
        fs_raw = getattr(self, "fs", 1.0)
        try:
            fs_arr = np.asarray(fs_raw)
            if fs_arr.ndim == 0:
                fs = float(fs_arr)
            elif fs_arr.size == 1:
                fs = float(fs_arr.reshape(-1)[0])
            else:
                fs = 1.0
        except Exception:
            fs = 1.0
        fs = fs if np.isfinite(fs) and fs > 0 else 1.0
        freqs = np.fft.fftfreq(int(s.size), d=1.0 / fs)
        amp = np.abs(yf) ** 2
        idx = freqs > 0
        freqs = freqs[idx]
        amp = amp[idx]
        if amp.size == 0 or not np.isfinite(amp).any():
            return None, None
        h = threshold * np.nanmax(amp)
        peaks, props = find_peaks(amp, height=h)
        if peaks.size == 0:
            return None, None
        peak_freqs = freqs[peaks]
        # avoid division by 0
        peak_freqs = peak_freqs[np.isfinite(peak_freqs) & (peak_freqs > 0)]
        if peak_freqs.size == 0:
            return None, None
        periods = 1.0 / peak_freqs
        return peak_freqs, periods
    except Exception as ex:
        logging.error(f"[Seasonality] {ex}")
        return None, None


def _bmt__coerce_1d_numeric(series_like):
    """Robustly coerce input to a 1D float64 numpy array."""
    try:
        s = pd.to_numeric(series_like, errors="coerce")
        if isinstance(s, pd.DataFrame):
            # take first column
            s = s.iloc[:, 0]
        if isinstance(s, (pd.Series, pd.Index)):
            arr = s.to_numpy()
        else:
            arr = np.asarray(s)
        arr = np.asarray(arr, dtype=np.float64)
        if arr.ndim == 0:
            arr = arr.reshape(1)
        else:
            arr = arr.reshape(-1)
        # drop NaNs
        arr = arr[np.isfinite(arr)]
        return arr
    except Exception:
        arr = np.asarray(series_like)
        arr = np.asarray(arr, dtype=np.float64).reshape(-1)
        arr = arr[np.isfinite(arr)]
        return arr


def _bmt_compute_hurst_rs(self, series) -> float:
    """Hurst via rescaled range (R/S)."""
    try:
        arr = _bmt__coerce_1d_numeric(series)
        if arr.size < 20:
            return np.nan
        # Prefer `hurst` package if available.
        try:
            H, _, _ = compute_Hc(arr, kind="change", simplified=True)
            return float(H)
        except Exception:
            # Fallback: nolds.hurst_rs
            if not _nolds_or_warn():
                return np.nan
            return float(nolds.hurst_rs(arr))
    except Exception as ex:
        logging.error(f"[Hurst RS] {ex}")
        return np.nan


def _bmt_compute_hurst_dfa(self, series) -> float:
    """Hurst via DFA."""
    try:
        arr = _bmt__coerce_1d_numeric(series)
        if arr.size < 20:
            return np.nan
        if not _nolds_or_warn():
            return np.nan
        return float(nolds.dfa(arr))
    except Exception as ex:
        logging.error(f"[Hurst DFA] {ex}")
        return np.nan


def _bmt_compute_hurst_aggvar(self, series, max_n: int = 100) -> float:
    """Hurst via aggregated variance method."""
    try:
        arr = _bmt__coerce_1d_numeric(series)
        N = int(arr.size)
        if N < 50:
            return np.nan
        max_n = int(max_n) if max_n is not None else 100
        max_n = max(10, min(max_n, N // 2))
        m_vals = np.arange(1, max_n + 1)
        variances = []
        used_m = []
        for m in m_vals:
            nb = N // m
            if nb <= 1:
                continue
            reshaped = arr[: nb * m].reshape(nb, m)
            block_means = reshaped.mean(axis=1)
            if block_means.size <= 1:
                continue
            v = np.var(block_means)
            if np.isfinite(v) and v > 0:
                variances.append(v)
                used_m.append(m)
        if len(variances) < 2:
            return np.nan
        log_m = np.log10(np.asarray(used_m, dtype=np.float64))
        log_var = np.log10(np.asarray(variances, dtype=np.float64))
        slope, _ = np.polyfit(log_m, log_var, 1)
        H = 1.0 - float(slope) / 2.0
        return float(H)
    except Exception as ex:
        logging.error(f"[Hurst AggVar] {ex}")
        return np.nan


def _bmt_compute_hurst_wavelet(self, series) -> float:
    """Hurst proxy via log-log PSD slope (rough wavelet-like heuristic)."""
    try:
        arr = _bmt__coerce_1d_numeric(series)
        N = int(arr.size)
        if N < 50:
            return np.nan
        arr = arr - np.mean(arr)
        yf = fft(arr)
        freqs = np.fft.fftfreq(N)
        psd = np.abs(yf) ** 2
        idx = freqs > 0
        freqs = freqs[idx]
        psd = psd[idx]
        if freqs.size < 2:
            return np.nan
        # Avoid zeros
        keep = (psd > 0) & np.isfinite(psd) & np.isfinite(freqs)
        freqs = freqs[keep]
        psd = psd[keep]
        if freqs.size < 2:
            return np.nan
        slope, _ = np.polyfit(np.log10(freqs), np.log10(psd), 1)
        # heuristic mapping; keep it bounded-ish
        H = (1.0 - float(slope)) / 2.0
        return float(H)
    except Exception as ex:
        logging.error(f"[Hurst Wavelet] {ex}")
        return np.nan


def _bmt_compute_hurst_rs(self, series) -> float:
    """Compatibility wrapper for Hurst RS. Uses `hurst.compute_Hc` if available."""
    try:
        s = pd.to_numeric(series, errors="coerce").dropna().values.astype(np.float64)
        if s.size < 16:
            return float("nan")
        try:
            # Prefer the library used elsewhere in the codebase.
            H, _, _ = compute_Hc(s, kind="change", simplified=True)
            return float(H)
        except Exception:
            # Fallback: RS via nolds if present.
            if hasattr(nolds, "hurst_rs"):
                return float(nolds.hurst_rs(s))
            return float("nan")
    except Exception as ex:
        logging.error(f"[Hurst RS] {ex}")
        return float("nan")


# Patch if missing (safe for newer versions).
try:
    if "BigMasterTool" in globals():
        if not hasattr(BigMasterTool, "test_stationarity"):
            BigMasterTool.test_stationarity = _bmt_test_stationarity
        if not hasattr(BigMasterTool, "compute_sample_entropy"):
            BigMasterTool.compute_sample_entropy = _bmt_compute_sample_entropy
        if not hasattr(BigMasterTool, "detect_seasonality"):
            BigMasterTool.detect_seasonality = _bmt_detect_seasonality
        # Hurst: ensure at least 4 variants are always available.
        if not hasattr(BigMasterTool, "compute_hurst_rs"):
            BigMasterTool.compute_hurst_rs = _bmt_compute_hurst_rs
        if not hasattr(BigMasterTool, "compute_hurst_dfa"):
            BigMasterTool.compute_hurst_dfa = _bmt_compute_hurst_dfa
        if not hasattr(BigMasterTool, "compute_hurst_aggregated_variance"):
            BigMasterTool.compute_hurst_aggregated_variance = _bmt_compute_hurst_aggvar
        # Some codepaths might call it differently.
        if not hasattr(BigMasterTool, "compute_hurst_aggvar"):
            BigMasterTool.compute_hurst_aggvar = _bmt_compute_hurst_aggvar
        if not hasattr(BigMasterTool, "compute_hurst_wavelet"):
            BigMasterTool.compute_hurst_wavelet = _bmt_compute_hurst_wavelet
except Exception as _patch_ex:
    logging.error(f"[HOTFIX] Failed to patch BigMasterTool: {_patch_ex}")

